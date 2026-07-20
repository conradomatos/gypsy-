"""
Mapeia todas as abas da planilha FAZZER ZIP.xlsm
Gera relatório Excel em intermediarios/MAPA_ABAS_FAZZER.xlsx
NÃO extrai dados — apenas documenta estrutura, propósito e dependências.
"""

import openpyxl
from openpyxl.utils import get_column_letter
import re
import os
from collections import defaultdict

# Caminhos
BASE = r"D:\00 - CLAUDE_CODE\01. PROJETOS\04. ORCAMENTACAO_POR_MODELAGEM_FINANCEIRA\01_DESTILACAO"
FONTE = os.path.join(BASE, "fontes", "FAZZER ZIP.xlsm")
SAIDA = os.path.join(BASE, "intermediarios", "MAPA_ABAS_FAZZER.xlsx")

print("Abrindo arquivo (com fórmulas)...")
wb = openpyxl.load_workbook(FONTE, data_only=False, read_only=False)
print(f"Abas: {len(wb.sheetnames)}")

# Também abrir com data_only=True para ver valores calculados
print("Abrindo arquivo (valores calculados)...")
wb_val = openpyxl.load_workbook(FONTE, data_only=True, read_only=True)

all_sheet_names = wb.sheetnames

def find_real_data_range(ws, max_scan=200):
    """Encontra o range real de dados (ignora linhas/colunas vazias)."""
    first_row = None
    last_row = 0
    first_col = None
    last_col = 0

    scan_rows = min(ws.max_row or 1, max_scan)
    scan_cols = min(ws.max_column or 1, 50)

    for r in range(1, scan_rows + 1):
        for c in range(1, scan_cols + 1):
            cell = ws.cell(r, c)
            if cell.value is not None:
                if first_row is None:
                    first_row = r
                last_row = r
                if first_col is None or c < first_col:
                    first_col = c
                if c > last_col:
                    last_col = c

    return first_row, last_row, first_col, last_col

def extract_sheet_refs(ws, max_scan_rows=500, max_scan_cols=50):
    """Extrai referências a outras abas nas fórmulas."""
    refs_to = set()  # abas que esta aba referencia
    formulas_found = []

    scan_rows = min(ws.max_row or 1, max_scan_rows)
    scan_cols = min(ws.max_column or 1, max_scan_cols)

    for r in range(1, scan_rows + 1):
        for c in range(1, scan_cols + 1):
            cell = ws.cell(r, c)
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                # Encontrar referências a outras abas: 'Nome Aba'!
                matches = re.findall(r"'([^']+)'!", val)
                for m in matches:
                    refs_to.add(m)
                # Também sem aspas: NomeAba!
                matches2 = re.findall(r"([A-Za-z0-9_]+)!", val)
                for m in matches2:
                    if m not in ('SUM', 'IF', 'VLOOKUP', 'INDEX', 'MATCH', 'ROUND',
                                 'SUMPRODUCT', 'IFERROR', 'CONCATENATE', 'TEXT',
                                 'LEFT', 'RIGHT', 'MID', 'LEN', 'TRIM', 'AVERAGE',
                                 'MAX', 'MIN', 'COUNT', 'COUNTA', 'COUNTIF',
                                 'SUMIF', 'SUMIFS', 'COUNTIFS', 'INDIRECT',
                                 'OFFSET', 'ROW', 'COLUMN', 'CEILING', 'FLOOR',
                                 'ABS', 'MOD', 'INT', 'VALUE', 'SEARCH', 'FIND',
                                 'SUBSTITUTE', 'REPLACE', 'UPPER', 'LOWER',
                                 'PROPER', 'REPT', 'EXACT', 'T', 'N',
                                 'AND', 'OR', 'NOT', 'TRUE', 'FALSE',
                                 'DATE', 'TODAY', 'NOW', 'YEAR', 'MONTH', 'DAY',
                                 'HOUR', 'MINUTE', 'SECOND', 'NETWORKDAYS',
                                 'EOMONTH', 'EDATE', 'DATEDIF',
                                 'HYPERLINK', 'CHOOSE', 'SWITCH',
                                 'ROUNDUP', 'ROUNDDOWN', 'SMALL', 'LARGE',
                                 'RANK', 'PERCENTILE', 'MEDIAN', 'MODE',
                                 'STDEV', 'VAR', 'CORREL', 'TREND', 'FORECAST',
                                 'TYPE', 'ISBLANK', 'ISERROR', 'ISNA', 'ISNUMBER',
                                 'ISTEXT', 'ISLOGICAL', 'ISREF', 'ISEVEN', 'ISODD'):
                        refs_to.add(m)

                if len(formulas_found) < 15:
                    formulas_found.append(f"{get_column_letter(c)}{r}: {val[:120]}")

    return refs_to, formulas_found

def get_headers(ws, max_header_rows=5):
    """Extrai possíveis headers das primeiras linhas."""
    headers = []
    scan_cols = min(ws.max_column or 1, 50)
    for r in range(1, max_header_rows + 1):
        row_vals = []
        for c in range(1, scan_cols + 1):
            val = ws.cell(r, c).value
            if val is not None:
                row_vals.append(str(val).strip()[:80])
        if row_vals:
            headers.append(f"L{r}: {' | '.join(row_vals)}")
    return headers

def count_non_empty(ws, max_scan=500):
    """Conta linhas com pelo menos um valor não-vazio."""
    count = 0
    scan_rows = min(ws.max_row or 1, max_scan)
    scan_cols = min(ws.max_column or 1, 50)
    for r in range(1, scan_rows + 1):
        for c in range(1, scan_cols + 1):
            if ws.cell(r, c).value is not None:
                count += 1
                break
    return count

def sample_data(ws_val, start_row, max_rows=5, max_cols=15):
    """Amostra de dados (valores calculados)."""
    samples = []
    scan_cols = min(ws_val.max_column or 1, max_cols)
    end_row = min((ws_val.max_row or 1), start_row + max_rows)
    for r in range(start_row, end_row + 1):
        row_vals = []
        for c in range(1, scan_cols + 1):
            val = ws_val.cell(r, c).value
            if val is not None:
                row_vals.append(str(val)[:60])
            else:
                row_vals.append("")
        samples.append(" | ".join(row_vals))
    return samples

# Analisar cada aba
print("\nAnalisando abas...\n")
results = []

# Primeiro passo: coletar todas as referências para calcular "quem puxa de mim"
all_refs = {}  # sheet_name -> set of sheets it references

for sheet_name in all_sheet_names:
    ws = wb[sheet_name]
    refs_to, _ = extract_sheet_refs(ws)
    all_refs[sheet_name] = refs_to

# Calcular referências inversas
refs_from = defaultdict(set)  # sheet_name -> set of sheets that reference it
for sheet_name, refs in all_refs.items():
    for ref in refs:
        if ref in all_sheet_names:
            refs_from[ref].add(sheet_name)

# Segundo passo: análise detalhada
for idx, sheet_name in enumerate(all_sheet_names, 1):
    print(f"  [{idx:2d}/39] {sheet_name}...")
    ws = wb[sheet_name]
    ws_v = wb_val[sheet_name]

    # Dimensões
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # Range real
    fr, lr, fc, lc = find_real_data_range(ws)

    # Contagem
    non_empty = count_non_empty(ws)

    # Headers
    headers = get_headers(ws)

    # Fórmulas e referências
    refs_to, formulas = extract_sheet_refs(ws)

    # Referências inversas (quem puxa dados desta aba)
    pulled_by = refs_from.get(sheet_name, set())

    # Amostra de dados
    data_start = fr if fr else 1
    samples = sample_data(ws_v, data_start, max_rows=8, max_cols=12)

    results.append({
        'idx': idx,
        'nome_aba': sheet_name,
        'max_row': max_row,
        'max_col': max_col,
        'real_range': f"L{fr or '?'}:L{lr or '?'} C{fc or '?'}:C{lc or '?'}",
        'linhas_com_dados': non_empty,
        'headers': "\n".join(headers),
        'formulas_amostra': "\n".join(formulas),
        'refs_para': ", ".join(sorted(refs_to)) if refs_to else "(nenhuma)",
        'refs_de': ", ".join(sorted(pulled_by)) if pulled_by else "(nenhuma)",
        'amostra_dados': "\n".join(samples),
    })

wb_val.close()

# Agora fazer análise qualitativa de cada aba
# Baseado nos headers, fórmulas e amostras, classificar
print("\nClassificando abas...\n")

for r in results:
    name = r['nome_aba']
    headers = r['headers'].lower()
    formulas = r['formulas_amostra'].lower()
    amostra = r['amostra_dados'].lower()
    refs_para = r['refs_para']
    refs_de = r['refs_de']
    linhas = r['linhas_com_dados']

    # Classificação por nome/conteúdo
    descricao = ""
    inputs_usuario = ""
    o_que_calcula = ""
    tabela_costai = ""
    prioridade = ""

    if "(1) Atividades" in name:
        descricao = "Lista mestre de atividades/serviços do orçamento. Estrutura hierárquica com códigos de atividade."
        inputs_usuario = "Código, descrição das atividades, unidade"
        o_que_calcula = "Nada — é cadastro base que alimenta outras abas"
        tabela_costai = "Composições (referência de serviços)"
        prioridade = "ALTA"
    elif "(2) Base Dados" in name:
        descricao = "Base de dados de insumos/materiais com códigos, descrições, unidades e preços."
        inputs_usuario = "Código, descrição, unidade, preço unitário"
        o_que_calcula = "Nada — é cadastro base de preços de insumos"
        tabela_costai = "Insumos"
        prioridade = "ALTA"
    elif name == "(3) Plan. Quant.":
        descricao = "Planilha de quantitativos — detalhamento de quantidades por atividade/serviço."
        inputs_usuario = "Quantidades de cada item por atividade"
        o_que_calcula = "Quantitativos totais por item/atividade. Referencia Base Dados para preços."
        tabela_costai = "Composições + ItensComposicao (quantitativos)"
        prioridade = "ALTA"
    elif "Plan. Quant. (Resumo)" in name:
        descricao = "Resumo dos quantitativos — agrupamento consolidado da Plan. Quant. detalhada."
        inputs_usuario = "Nenhum — é resumo automático"
        o_que_calcula = "Totalização dos quantitativos por atividade"
        tabela_costai = "Referência (validação de totais)"
        prioridade = "BAIXA"
    elif "Tab. Dinâm" in name or "Din" in name:
        descricao = "Tabela dinâmica — resumo pivot dos quantitativos para análise rápida."
        inputs_usuario = "Nenhum — tabela dinâmica automática"
        o_que_calcula = "Resumo cruzado dos dados de quantitativos"
        tabela_costai = "Nenhuma — é visualização"
        prioridade = "IGNORAR"
    elif "Profissionais" in name or "Sal" in name:
        descricao = "Cadastro de profissionais com funções e salários base. Define custo da mão de obra."
        inputs_usuario = "Função, salário mensal, tipo (CLT/PJ), periculosidade"
        o_que_calcula = "Base salarial para composições de MO"
        tabela_costai = "Mão de Obra"
        prioridade = "ALTA"
    elif name == "Equipes":
        descricao = "Composição de equipes — define quais profissionais compõem cada equipe de trabalho e quantos de cada."
        inputs_usuario = "Composição de cada equipe (quais funções, quantos de cada)"
        o_que_calcula = "Custo de equipe por composição de profissionais"
        tabela_costai = "Composições (equipes) + ItensComposicao"
        prioridade = "ALTA"
    elif name == "Histograma":
        descricao = "Histograma de mão de obra — distribuição temporal de profissionais ao longo do cronograma."
        inputs_usuario = "Períodos, alocação de equipes"
        o_que_calcula = "Distribuição de HH ao longo do tempo, picos de MO"
        tabela_costai = "Nenhuma — é planejamento operacional"
        prioridade = "IGNORAR"
    elif name == "Cronograma":
        descricao = "Cronograma físico da obra — duração e sequência de atividades."
        inputs_usuario = "Datas de início/fim, duração de atividades"
        o_que_calcula = "Gantt/cronograma do projeto"
        tabela_costai = "Nenhuma — é planejamento"
        prioridade = "IGNORAR"
    elif "MOI - MOD" in name:
        descricao = "Resumo comparativo de custos MOI (mão de obra indireta) e MOD (mão de obra direta)."
        inputs_usuario = "Nenhum — é resumo calculado"
        o_que_calcula = "Totais e proporções de MOI vs MOD"
        tabela_costai = "Referência (validação de custos MO)"
        prioridade = "MEDIA"
    elif "Histograma - 01" in name:
        descricao = "Histograma detalhado cenário 01 — distribuição de HH por período e função."
        inputs_usuario = "Nenhum — calculado a partir de Equipes e Cronograma"
        o_que_calcula = "HH por função/período no cenário 1"
        tabela_costai = "Nenhuma — é planejamento operacional"
        prioridade = "IGNORAR"
    elif "Histograma - 02" in name:
        descricao = "Histograma detalhado cenário 02 — variação do cenário 1."
        inputs_usuario = "Nenhum — calculado"
        o_que_calcula = "HH por função/período no cenário 2"
        tabela_costai = "Nenhuma — é planejamento operacional"
        prioridade = "IGNORAR"
    elif "Histograma - 03" in name:
        descricao = "Histograma detalhado cenário 03 — variação do cenário 1."
        inputs_usuario = "Nenhum — calculado"
        o_que_calcula = "HH por função/período no cenário 3"
        tabela_costai = "Nenhuma — é planejamento operacional"
        prioridade = "IGNORAR"
    elif "Resumo Geral - HH_01" in name:
        descricao = "Resumo geral de HH cenário 01 — totalização de homens-hora por função e período."
        inputs_usuario = "Nenhum — calculado dos histogramas"
        o_que_calcula = "Total HH por função no cenário 1"
        tabela_costai = "Referência (validação HH)"
        prioridade = "BAIXA"
    elif "Resumo Geral - HH_02" in name:
        descricao = "Resumo geral de HH cenário 02."
        inputs_usuario = "Nenhum — calculado"
        o_que_calcula = "Total HH por função no cenário 2"
        tabela_costai = "Referência"
        prioridade = "BAIXA"
    elif "Resumo Geral - HH_03" in name:
        descricao = "Resumo geral de HH cenário 03."
        inputs_usuario = "Nenhum — calculado"
        o_que_calcula = "Total HH por função no cenário 3"
        tabela_costai = "Referência"
        prioridade = "BAIXA"
    elif name == "(7) Resumo Geral - HH":
        descricao = "Consolidação final dos 3 cenários de HH — resumo comparativo."
        inputs_usuario = "Nenhum — consolidação automática"
        o_que_calcula = "Comparação dos 3 cenários de HH total"
        tabela_costai = "Referência (total HH consolidado)"
        prioridade = "BAIXA"
    elif "Encargos Sociais" in name:
        descricao = "Tabela de encargos sociais trabalhistas — percentuais por grupo (A, B, C, D)."
        inputs_usuario = "Alíquotas de encargos (INSS, FGTS, férias, 13º, etc.)"
        o_que_calcula = "Percentual total de encargos sobre salário"
        tabela_costai = "EncargosBDI"
        prioridade = "ALTA"
    elif name == "Planilha1":
        descricao = "Aba auxiliar/rascunho — provavelmente cálculos temporários ou testes."
        inputs_usuario = "Variável — pode conter dados auxiliares"
        o_que_calcula = "Indefinido — verificar conteúdo"
        tabela_costai = "A definir após inspeção"
        prioridade = "BAIXA"
    elif "Incidências  MOI" in name or "Incid" in name and "MOI" in name:
        descricao = "Incidências sobre MOI — percentuais adicionais sobre custo de mão de obra indireta (periculosidade, insalubridade, etc.)."
        inputs_usuario = "Percentuais de incidências por tipo"
        o_que_calcula = "Fator multiplicador total de incidências sobre MOI"
        tabela_costai = "EncargosBDI (incidências MO)"
        prioridade = "ALTA"
    elif "Composição MOI" in name or "Composi" in name and "MOI" in name:
        descricao = "Composição de custo de MOI — detalhamento do custo de cada função indireta (salário + encargos + benefícios)."
        inputs_usuario = "Nenhum — composição automática de Profissionais + Encargos + Incidências"
        o_que_calcula = "Custo/hora completo por função MOI (com todos os encargos)"
        tabela_costai = "Composições (MO indireta) + ItensComposicao"
        prioridade = "ALTA"
    elif "Incidências  MOD" in name or "Incid" in name and "MOD" in name:
        descricao = "Incidências sobre MOD — percentuais adicionais sobre custo de mão de obra direta."
        inputs_usuario = "Percentuais de incidências por tipo"
        o_que_calcula = "Fator multiplicador total de incidências sobre MOD"
        tabela_costai = "EncargosBDI (incidências MO)"
        prioridade = "ALTA"
    elif "Composição MOD" in name or "Composi" in name and "MOD" in name:
        descricao = "Composição de custo de MOD — detalhamento do custo de cada função direta (salário + encargos + benefícios)."
        inputs_usuario = "Nenhum — composição automática de Profissionais + Encargos + Incidências"
        o_que_calcula = "Custo/hora completo por função MOD (com todos os encargos)"
        tabela_costai = "Composições (MO direta) + ItensComposicao"
        prioridade = "ALTA"
    elif name == "Macro1":
        descricao = "Aba de macro VBA — contém código de automação, não dados."
        inputs_usuario = "Nenhum — é código VBA"
        o_que_calcula = "Automação de processos da planilha"
        tabela_costai = "Nenhuma"
        prioridade = "IGNORAR"
    elif "Valores Custo-Venda_HH" in name:
        descricao = "Tabela de valores de custo e venda por HH — resume preço de custo e de venda por função."
        inputs_usuario = "BDI/markup por função ou global"
        o_que_calcula = "Preço de custo HH e preço de venda HH por função"
        tabela_costai = "Mão de Obra (custo_hora + custo_hora_com_encargos) + referência de markup"
        prioridade = "ALTA"
    elif "Resumo Preço Custo HH_01" in name:
        descricao = "Resumo de preço de custo HH cenário 01 — valoração do histograma a preço de custo."
        inputs_usuario = "Nenhum — calculado de Histograma × Custo HH"
        o_que_calcula = "Custo total de MO por período/função no cenário 1"
        tabela_costai = "Referência (custo MO cenário 1)"
        prioridade = "BAIXA"
    elif "Resumo Preço Custo HH_02" in name:
        descricao = "Resumo de preço de custo HH cenário 02."
        inputs_usuario = "Nenhum — calculado"
        o_que_calcula = "Custo total de MO por período/função no cenário 2"
        tabela_costai = "Referência"
        prioridade = "BAIXA"
    elif "Resumo Preço Custo HH_03" in name:
        descricao = "Resumo de preço de custo HH cenário 03."
        inputs_usuario = "Nenhum — calculado"
        o_que_calcula = "Custo total de MO por período/função no cenário 3"
        tabela_costai = "Referência"
        prioridade = "BAIXA"
    elif "Resumo Preço Venda HH_01" in name:
        descricao = "Resumo de preço de venda HH cenário 01 — valoração do histograma a preço de venda."
        inputs_usuario = "Nenhum — calculado de Histograma × Venda HH"
        o_que_calcula = "Receita total de MO por período/função no cenário 1"
        tabela_costai = "Referência (preço venda cenário 1)"
        prioridade = "BAIXA"
    elif "Resumo Preço Venda HH_02" in name:
        descricao = "Resumo de preço de venda HH cenário 02."
        inputs_usuario = "Nenhum — calculado"
        o_que_calcula = "Receita total de MO no cenário 2"
        tabela_costai = "Referência"
        prioridade = "BAIXA"
    elif "Resumo Preço Venda HH_03" in name:
        descricao = "Resumo de preço de venda HH cenário 03."
        inputs_usuario = "Nenhum — calculado"
        o_que_calcula = "Receita total de MO no cenário 3"
        tabela_costai = "Referência"
        prioridade = "BAIXA"
    elif "Equipe Mobilização" in name or "Mobiliza" in name and "Equipe" in name:
        descricao = "Composição da equipe de mobilização — define quem participa e custo da mobilização."
        inputs_usuario = "Composição da equipe de mobilização, dias"
        o_que_calcula = "Custo de mobilização por função e total"
        tabela_costai = "Composições (mobilização)"
        prioridade = "MEDIA"
    elif "Mob" in name and "Desm" in name:
        descricao = "Custos de mobilização e desmobilização — transporte, hospedagem, logística."
        inputs_usuario = "Itens de custo, quantidades, valores unitários"
        o_que_calcula = "Custo total de mob/desmob"
        tabela_costai = "Composições (mob/desmob)"
        prioridade = "MEDIA"
    elif "Despesas Operacionais" in name:
        descricao = "Despesas operacionais da obra — custos fixos mensais e variáveis."
        inputs_usuario = "Itens de despesa, valores mensais, duração"
        o_que_calcula = "Custo total de despesas operacionais por período"
        tabela_costai = "Composições (despesas operacionais) / EncargosBDI"
        prioridade = "MEDIA"
    elif "Máq" in name and "Ferramentas" in name or "Ferr" in name and "quinas" in name.lower():
        descricao = "Máquinas e ferramentas — cadastro e custo de equipamentos leves e ferramentas."
        inputs_usuario = "Item, quantidade, valor unitário, depreciação"
        o_que_calcula = "Custo total e depreciado de ferramentas"
        tabela_costai = "Equipamentos"
        prioridade = "ALTA"
    elif "(24) Equipamentos" in name:
        descricao = "Equipamentos pesados — cadastro e custo de equipamentos maiores (guindastes, caminhões, etc.)."
        inputs_usuario = "Equipamento, tipo (próprio/alugado), custo mensal/hora"
        o_que_calcula = "Custo total de equipamentos por período"
        tabela_costai = "Equipamentos"
        prioridade = "ALTA"
    elif "Preço de Custo Total" in name:
        descricao = "Resumo geral de preço de custo — consolida todos os custos (MO, materiais, equipamentos, despesas)."
        inputs_usuario = "Nenhum — é consolidação"
        o_que_calcula = "Preço de custo total do orçamento por atividade"
        tabela_costai = "Referência (validação de custo total)"
        prioridade = "MEDIA"
    elif "Memória Cálculo" in name or "Mem" in name and "PV" in name:
        descricao = "Memória de cálculo do preço de venda — BDI, impostos, lucro e formação do preço de venda."
        inputs_usuario = "BDI, percentual de lucro, impostos"
        o_que_calcula = "Fator de markup e composição do preço de venda"
        tabela_costai = "EncargosBDI (BDI/markup)"
        prioridade = "ALTA"
    elif "Preço de Venda Total" in name:
        descricao = "Resumo geral de preço de venda — custo + markup = preço de venda por atividade."
        inputs_usuario = "Nenhum — é consolidação"
        o_que_calcula = "Preço de venda total do orçamento"
        tabela_costai = "Referência (preço de venda final)"
        prioridade = "MEDIA"
    else:
        descricao = f"Aba não classificada automaticamente. Headers: {r['headers'][:200]}"
        inputs_usuario = "A definir"
        o_que_calcula = "A definir"
        tabela_costai = "A definir"
        prioridade = "A DEFINIR"

    r['descricao'] = descricao
    r['inputs_usuario'] = inputs_usuario
    r['o_que_calcula'] = o_que_calcula
    r['tabela_costai'] = tabela_costai
    r['prioridade'] = prioridade

# Gerar Excel de saída
print("\nGerando Excel de saída...")
wb_out = openpyxl.Workbook()
ws_out = wb_out.active
ws_out.title = "Mapa de Abas"

# Headers
colunas = [
    "Nº", "Nome da Aba", "Descrição / Propósito", "Inputs do Usuário",
    "O que Calcula/Gera", "Puxa Dados De (fórmulas)", "Alimenta Abas",
    "Tabela CostAI", "Prioridade CostAI",
    "Linhas Max", "Colunas Max", "Linhas com Dados", "Range Real",
    "Headers (primeiras linhas)", "Fórmulas (amostra)", "Dados (amostra)"
]

for c, header in enumerate(colunas, 1):
    cell = ws_out.cell(1, c, header)
    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    cell.fill = openpyxl.styles.PatternFill("solid", fgColor="2F5496")
    cell.alignment = openpyxl.styles.Alignment(horizontal="center", wrap_text=True)

# Dados
for row_idx, r in enumerate(results, 2):
    ws_out.cell(row_idx, 1, r['idx'])
    ws_out.cell(row_idx, 2, r['nome_aba'])
    ws_out.cell(row_idx, 3, r['descricao'])
    ws_out.cell(row_idx, 4, r['inputs_usuario'])
    ws_out.cell(row_idx, 5, r['o_que_calcula'])
    ws_out.cell(row_idx, 6, r['refs_para'])
    ws_out.cell(row_idx, 7, r['refs_de'])
    ws_out.cell(row_idx, 8, r['tabela_costai'])
    ws_out.cell(row_idx, 9, r['prioridade'])
    ws_out.cell(row_idx, 10, r['max_row'])
    ws_out.cell(row_idx, 11, r['max_col'])
    ws_out.cell(row_idx, 12, r['linhas_com_dados'])
    ws_out.cell(row_idx, 13, r['real_range'])
    ws_out.cell(row_idx, 14, r['headers'])
    ws_out.cell(row_idx, 15, r['formulas_amostra'])
    ws_out.cell(row_idx, 16, r['amostra_dados'])

    # Wrap text em colunas textuais
    for c in (3, 4, 5, 6, 7, 8, 14, 15, 16):
        ws_out.cell(row_idx, c).alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")

    # Colorir por prioridade
    pri = r['prioridade']
    if pri == "ALTA":
        fill = openpyxl.styles.PatternFill("solid", fgColor="C6EFCE")
    elif pri == "MEDIA":
        fill = openpyxl.styles.PatternFill("solid", fgColor="FFEB9C")
    elif pri == "BAIXA":
        fill = openpyxl.styles.PatternFill("solid", fgColor="D9E2F3")
    elif pri == "IGNORAR":
        fill = openpyxl.styles.PatternFill("solid", fgColor="F2F2F2")
    else:
        fill = openpyxl.styles.PatternFill("solid", fgColor="FFC7CE")
    ws_out.cell(row_idx, 9).fill = fill

# Largura das colunas
col_widths = [5, 30, 50, 35, 45, 40, 40, 35, 15, 10, 10, 12, 18, 60, 60, 60]
for i, w in enumerate(col_widths, 1):
    ws_out.column_dimensions[get_column_letter(i)].width = w

# Filtro automático
ws_out.auto_filter.ref = f"A1:{get_column_letter(len(colunas))}{len(results)+1}"

# Congelar primeira linha
ws_out.freeze_panes = "A2"

# === Segunda aba: Resumo de dependências ===
ws_dep = wb_out.create_sheet("Dependências")
ws_dep.cell(1, 1, "Aba").font = openpyxl.styles.Font(bold=True)
ws_dep.cell(1, 2, "Puxa dados de").font = openpyxl.styles.Font(bold=True)
ws_dep.cell(1, 3, "É puxada por").font = openpyxl.styles.Font(bold=True)
ws_dep.cell(1, 4, "Nº refs para").font = openpyxl.styles.Font(bold=True)
ws_dep.cell(1, 5, "Nº refs de").font = openpyxl.styles.Font(bold=True)

for row_idx, r in enumerate(results, 2):
    ws_dep.cell(row_idx, 1, r['nome_aba'])
    ws_dep.cell(row_idx, 2, r['refs_para'])
    ws_dep.cell(row_idx, 3, r['refs_de'])
    rt = r['refs_para'].split(", ") if r['refs_para'] != "(nenhuma)" else []
    rd = r['refs_de'].split(", ") if r['refs_de'] != "(nenhuma)" else []
    ws_dep.cell(row_idx, 4, len(rt))
    ws_dep.cell(row_idx, 5, len(rd))

ws_dep.column_dimensions['A'].width = 30
ws_dep.column_dimensions['B'].width = 60
ws_dep.column_dimensions['C'].width = 60
ws_dep.column_dimensions['D'].width = 12
ws_dep.column_dimensions['E'].width = 12

# === Terceira aba: Resumo CostAI ===
ws_costai = wb_out.create_sheet("Resumo CostAI")
headers_costai = ["Tabela CostAI", "Abas Fonte (prioridade ALTA)", "Abas Fonte (MEDIA)", "Abas Fonte (BAIXA)"]
for c, h in enumerate(headers_costai, 1):
    ws_costai.cell(1, c, h).font = openpyxl.styles.Font(bold=True)

# Agrupar por tabela CostAI
costai_map = defaultdict(lambda: {"ALTA": [], "MEDIA": [], "BAIXA": []})
for r in results:
    tabela = r['tabela_costai']
    pri = r['prioridade']
    if pri in ("ALTA", "MEDIA", "BAIXA"):
        costai_map[tabela][pri].append(r['nome_aba'])

row = 2
for tabela in sorted(costai_map.keys()):
    vals = costai_map[tabela]
    ws_costai.cell(row, 1, tabela)
    ws_costai.cell(row, 2, ", ".join(vals["ALTA"]) if vals["ALTA"] else "-")
    ws_costai.cell(row, 3, ", ".join(vals["MEDIA"]) if vals["MEDIA"] else "-")
    ws_costai.cell(row, 4, ", ".join(vals["BAIXA"]) if vals["BAIXA"] else "-")
    row += 1

for col in range(1, 5):
    ws_costai.column_dimensions[get_column_letter(col)].width = 50

# Salvar
wb_out.save(SAIDA)
print(f"\nSalvo em: {SAIDA}")

# Resumo
print("\n=== RESUMO ===")
for pri in ("ALTA", "MEDIA", "BAIXA", "IGNORAR"):
    abas = [r['nome_aba'] for r in results if r['prioridade'] == pri]
    print(f"  {pri}: {len(abas)} abas")
    for a in abas:
        print(f"    - {a}")

wb.close()
print("\nDone!")
