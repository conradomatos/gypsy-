"""
Extração: Mão de Obra
Fontes:
  - Tabela de Entrada de Preços (base salarial — todas as funções)
  - M.O.I (classifica funções indiretas + Hxh Venda)
  - M.O.D. Elétrica (classifica funções diretas + periculosidade)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import re
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
FONTE = BASE_DIR / "fontes" / "Or xxx - Planilhas Pequenas Obras_Lucro Presumido.xlsx"
SAIDA = BASE_DIR / "intermediarios" / "EXTRACAO_Pequenas_Obras_LP.xlsx"
REVISAR = BASE_DIR / "validacao" / "REVISAR_Pequenas_Obras_LP_MaoDeObra.xlsx"

NOME_FONTE = FONTE.name
ENCARGOS_PERCENTUAL = 0.786785606926407  # Extraído de M.O.I L56 col C e M.O.D. L56 col C


def normalizar_texto(raw):
    if raw is None:
        return ""
    txt = str(raw).strip()
    txt = re.sub(r"\s+", " ", txt)
    return txt.upper()


def eh_item_mo(vals):
    """Verifica se a linha é um item de MO (tem número no col1 e descrição no col2)."""
    item = vals.get(1)
    desc = vals.get(2)
    if item is None or desc is None:
        return False
    if isinstance(item, (int, float)) and item > 0:
        return True
    return False


def extrair_tabela_entrada(ws_val, ws_form):
    """Extrai funções da aba Tabela de Entrada de Preços."""
    registros = []
    grupo_atual = ""

    for i, row in enumerate(ws_val.iter_rows(min_row=6, max_row=63, values_only=False), 6):
        vals = {c.column: c.value for c in row if c.value is not None}
        if not vals:
            continue

        # Detectar headers de grupo
        desc = vals.get(2)
        if desc and isinstance(desc, str):
            desc_upper = desc.strip().upper()
            if desc_upper in (
                "GERÊNCIA DO CONTRATO", "GERENCIA DO CONTRATO",
                "CONSTRUÇÃO ELETROMECÂNICA", "CONSTRUCAO ELETROMECANICA",
                "INSTRUMENTAÇÃO", "INSTRUMENTACAO",
                "CONSTRUÇÃO CIVIL", "CONSTRUCAO CIVIL",
                "OUTRAS FUNÇÕES", "OUTRAS FUNCOES",
                "VALE ALIMENTAÇÃO (CONFORME ACORDO COLETIVO)",
            ):
                grupo_atual = desc_upper
                continue
            # Detectar por padrões parciais (encoding)
            for g in ["GER", "CONSTRU", "INSTRUMENT", "OUTRAS", "VALE ALIMENT"]:
                if desc_upper.startswith(g) and 1 not in vals:
                    grupo_atual = desc_upper
                    break

        if not eh_item_mo(vals):
            continue

        # Pular a linha de vale alimentação (é referência, não função)
        if "VALE ALIMENT" in grupo_atual.upper() if grupo_atual else False:
            continue
        if "TICKET" in str(desc).upper() if desc else False:
            continue

        funcao = normalizar_texto(vals.get(2, ""))
        unidade = str(vals.get(3, "")).strip().lower() if vals.get(3) else ""
        salario_mensal = vals.get(4)
        tipo = str(vals.get(5, "")).strip().upper() if vals.get(5) else ""
        custo_hora = vals.get(6)
        peric_raw = vals.get(7)
        dissidio_info = vals.get(8)
        sal_com_peric = vals.get(8)

        # Periculosidade
        periculosidade = 0.0
        if isinstance(peric_raw, str) and peric_raw.strip().upper() in ("SIM", "S"):
            periculosidade = 0.30
        elif isinstance(peric_raw, str) and "INSALUB" in peric_raw.upper():
            periculosidade = 0.20  # Insalubridade padrão

        # Salário mensal
        sal_mensal = None
        if isinstance(salario_mensal, (int, float)):
            sal_mensal = round(salario_mensal, 2)

        # Custo hora
        custo_h = None
        if isinstance(custo_hora, (int, float)):
            custo_h = round(custo_hora, 4)

        # Custo hora com encargos (só para CLT)
        custo_hora_enc = None
        if tipo == "CLT" and custo_h:
            base_peric = custo_h * (1 + periculosidade)
            custo_hora_enc = round(base_peric * (1 + ENCARGOS_PERCENTUAL), 4)

        # Fórmulas
        row_form = {}
        for _, frow in enumerate(ws_form.iter_rows(min_row=i, max_row=i, values_only=False)):
            for c in frow:
                if c.value and isinstance(c.value, str) and c.value.startswith("="):
                    row_form[c.column] = c.value
        formula_parts = []
        if 4 in row_form:
            formula_parts.append(f"salario: {row_form[4]}")
        if 6 in row_form:
            formula_parts.append(f"custo_hora: {row_form[6]}")

        # Dissídio
        dissidio_ref = ""
        if isinstance(dissidio_info, str) and "DISSIDIO" in dissidio_info.upper():
            dissidio_ref = dissidio_info.strip()
        elif isinstance(dissidio_info, str) and "PERICULOSIDADE" in dissidio_info.upper():
            dissidio_ref = "DISSIDIO 2020"  # da coluna H header
        else:
            dissidio_ref = "DISSIDIO 2020"

        # Motivos revisão
        motivos = []
        if not funcao:
            motivos.append("SEM FUNCAO")
        if not tipo:
            motivos.append("SEM TIPO CLT/PJ")
        if sal_mensal is None:
            motivos.append("SEM SALARIO")

        registro = {
            "codigo_fonte": "",
            "funcao": funcao,
            "tipo": tipo,
            "grupo": normalizar_texto(grupo_atual),
            "salario_mensal": sal_mensal,
            "custo_hora": custo_h,
            "periculosidade": periculosidade if periculosidade > 0 else None,
            "encargos_percentual": ENCARGOS_PERCENTUAL if tipo == "CLT" else None,
            "custo_hora_com_encargos": custo_hora_enc,
            "dissidio_referencia": dissidio_ref,
            "classificacao": "",  # será preenchido cruzando com MOI/MOD
            "hxh_venda": None,  # será preenchido pelo MOI
            "_fonte": NOME_FONTE,
            "_aba_origem": "Tabela de Entrada de Preços",
            "_linha_origem": i,
            "_confianca": "media",
            "_revisar": "SIM" if motivos else "",
            "_formula_original": " | ".join(formula_parts) if formula_parts else "",
            "_observacao": " | ".join(motivos) if motivos else "",
        }
        registros.append(registro)

    return registros


def extrair_moi(ws_val):
    """Extrai dados do M.O.I para enriquecer os registros."""
    funcoes_moi = {}
    for i, row in enumerate(ws_val.iter_rows(min_row=11, max_row=20, values_only=False), 11):
        vals = {c.column: c.value for c in row if c.value is not None}
        funcao = normalizar_texto(vals.get(2, ""))
        hxh_venda = vals.get(11)
        custo_hora = vals.get(6)
        if funcao and isinstance(custo_hora, (int, float)):
            funcoes_moi[funcao] = {
                "custo_hora": round(custo_hora, 4),
                "hxh_venda": round(hxh_venda, 4) if isinstance(hxh_venda, (int, float)) else None,
                "classificacao": "MOI",
                "linha_moi": i,
            }
    return funcoes_moi


def extrair_mod(ws_val):
    """Extrai dados do M.O.D para enriquecer os registros."""
    funcoes_mod = {}
    for i, row in enumerate(ws_val.iter_rows(min_row=11, max_row=20, values_only=False), 11):
        vals = {c.column: c.value for c in row if c.value is not None}
        funcao = normalizar_texto(vals.get(2, ""))
        custo_hora = vals.get(6)
        if funcao and isinstance(custo_hora, (int, float)):
            funcoes_mod[funcao] = {
                "custo_hora": round(custo_hora, 4),
                "classificacao": "MOD",
                "linha_mod": i,
            }
    return funcoes_mod


def cruzar_dados(registros, funcoes_moi, funcoes_mod):
    """Enriquece registros da Tabela de Entrada com dados de MOI/MOD."""
    for reg in registros:
        funcao = reg["funcao"]

        # Buscar no MOI
        if funcao in funcoes_moi:
            moi = funcoes_moi[funcao]
            reg["classificacao"] = "MOI"
            if moi["hxh_venda"]:
                reg["hxh_venda"] = moi["hxh_venda"]

        # Buscar no MOD
        if funcao in funcoes_mod:
            mod = funcoes_mod[funcao]
            if reg["classificacao"] == "MOI":
                reg["classificacao"] = "MOI/MOD"
            else:
                reg["classificacao"] = "MOD"

        # Funções que não estão em nenhuma aba MOI/MOD
        if not reg["classificacao"]:
            # Inferir pelo grupo
            grupo = reg["grupo"]
            if any(x in grupo for x in ["ELETROMEC", "INSTRUMENT"]):
                reg["classificacao"] = "MOD"
            elif any(x in grupo for x in ["GER", "ADMINISTR"]):
                reg["classificacao"] = "MOI"
            elif "CIVIL" in grupo:
                reg["classificacao"] = "MOD"
            elif "OUTRAS" in grupo:
                reg["classificacao"] = "MOD"
            else:
                reg["classificacao"] = "NAO_CLASSIFICADO"
                reg["_revisar"] = "SIM"
                obs = reg["_observacao"]
                reg["_observacao"] = (obs + " | " if obs else "") + "NAO ENCONTRADO EM MOI/MOD"


def extrair():
    print(f"Abrindo {FONTE.name}...")

    wb_val = openpyxl.load_workbook(str(FONTE), read_only=True, data_only=True)
    wb_form = openpyxl.load_workbook(str(FONTE), read_only=True, data_only=False)

    aba_entrada = "Tabela de Entrada de Pe\xe7os"
    aba_moi = "M.O.I"
    aba_mod = "M.O.D. El\xe9trica"

    # 1. Extrair base da Tabela de Entrada
    print("Extraindo Tabela de Entrada de Preços...")
    registros = extrair_tabela_entrada(wb_val[aba_entrada], wb_form[aba_entrada])
    print(f"  {len(registros)} funções extraídas")

    # 2. Extrair MOI e MOD
    print("Extraindo M.O.I...")
    funcoes_moi = extrair_moi(wb_val[aba_moi])
    print(f"  {len(funcoes_moi)} funções MOI")

    print("Extraindo M.O.D. Elétrica...")
    funcoes_mod = extrair_mod(wb_val[aba_mod])
    print(f"  {len(funcoes_mod)} funções MOD")

    # 3. Cruzar
    print("Cruzando dados...")
    cruzar_dados(registros, funcoes_moi, funcoes_mod)

    wb_val.close()
    wb_form.close()

    # 4. Salvar
    salvar_excel(registros)
    gerar_relatorio(registros, funcoes_moi, funcoes_mod)

    # 5. Revisão
    revisar = [r for r in registros if r["_revisar"] == "SIM"]
    if revisar:
        salvar_revisao(revisar)


def salvar_excel(registros):
    """Adiciona aba MaoDeObra ao Excel de extração existente ou cria novo."""
    colunas = [
        "codigo_fonte", "funcao", "tipo", "grupo", "classificacao",
        "salario_mensal", "custo_hora", "periculosidade",
        "encargos_percentual", "custo_hora_com_encargos", "hxh_venda",
        "dissidio_referencia",
        "_fonte", "_aba_origem", "_linha_origem", "_confianca",
        "_revisar", "_formula_original", "_observacao",
    ]

    # Abrir Excel existente ou criar novo
    try:
        if SAIDA.exists():
            wb = openpyxl.load_workbook(str(SAIDA))
            if "MaoDeObra" in wb.sheetnames:
                del wb["MaoDeObra"]
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
    except PermissionError:
        print(f"AVISO: {SAIDA.name} está aberto. Criando arquivo separado.")
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet("MaoDeObra")

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    audit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    revisar_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Header
    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Dados
    for row_idx, reg in enumerate(registros, 2):
        for col_idx, col_name in enumerate(colunas, 1):
            val = reg.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_name.startswith("_"):
                cell.fill = audit_fill
            if reg["_revisar"] == "SIM":
                cell.fill = revisar_fill

    # Larguras
    larguras = {
        "codigo_fonte": 14, "funcao": 40, "tipo": 8, "grupo": 30,
        "classificacao": 14, "salario_mensal": 16, "custo_hora": 12,
        "periculosidade": 14, "encargos_percentual": 16,
        "custo_hora_com_encargos": 20, "hxh_venda": 12,
        "dissidio_referencia": 18,
        "_fonte": 45, "_aba_origem": 25, "_linha_origem": 12,
        "_confianca": 10, "_revisar": 8, "_formula_original": 50,
        "_observacao": 40,
    }
    for col_idx, col_name in enumerate(colunas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larguras.get(col_name, 12)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(registros) + 1}"
    ws.freeze_panes = "A2"

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(SAIDA))
        print(f"Salvo: {SAIDA} (aba MaoDeObra)")
    except PermissionError:
        fallback = SAIDA.parent / "EXTRACAO_Pequenas_Obras_LP_MaoDeObra.xlsx"
        wb.save(str(fallback))
        print(f"Salvo (fallback): {fallback}")


def salvar_revisao(registros):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revisar MO"

    colunas = [
        "funcao", "tipo", "grupo", "classificacao",
        "salario_mensal", "custo_hora", "_linha_origem", "_revisar", "_observacao",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")

    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, reg in enumerate(registros, 2):
        for col_idx, col_name in enumerate(colunas, 1):
            ws.cell(row=row_idx, column=col_idx, value=reg.get(col_name, ""))

    larguras = {"funcao": 40, "tipo": 8, "grupo": 30, "classificacao": 14,
                "salario_mensal": 16, "custo_hora": 12, "_linha_origem": 12,
                "_revisar": 8, "_observacao": 50}
    for col_idx, col_name in enumerate(colunas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larguras.get(col_name, 12)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(registros) + 1}"
    ws.freeze_panes = "A2"

    REVISAR.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(REVISAR))
    print(f"Salvo: {REVISAR}")


def gerar_relatorio(registros, funcoes_moi, funcoes_mod):
    print("\n" + "=" * 60)
    print("RELATÓRIO DE EXTRAÇÃO — Mão de Obra")
    print("=" * 60)
    print(f"Fonte: {NOME_FONTE}")
    print(f"Abas:  Tabela de Entrada de Preços + M.O.I + M.O.D. Elétrica")
    print(f"Data:  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("-" * 60)
    print(f"Total funções extraídas:  {len(registros)}")

    # Por tipo
    tipos = {}
    for r in registros:
        t = r["tipo"] or "SEM TIPO"
        tipos[t] = tipos.get(t, 0) + 1
    print(f"Por tipo: {tipos}")

    # Por classificação
    classif = {}
    for r in registros:
        c = r["classificacao"] or "NAO_CLASSIFICADO"
        classif[c] = classif.get(c, 0) + 1
    print(f"Por classificação: {classif}")

    # Por grupo
    grupos = {}
    for r in registros:
        g = r["grupo"] or "SEM GRUPO"
        grupos[g] = grupos.get(g, 0) + 1
    print(f"Por grupo:")
    for g, c in sorted(grupos.items(), key=lambda x: -x[1]):
        print(f"  {g:40s} {c:>3d}")

    # Revisão
    revisar = [r for r in registros if r["_revisar"] == "SIM"]
    print(f"Marcados para revisão: {len(revisar)}")
    for r in revisar:
        print(f"  L{r['_linha_origem']}: {r['funcao'][:30]} — {r['_observacao']}")

    # Cruzamento MOI/MOD
    moi_nomes = set(funcoes_moi.keys())
    mod_nomes = set(funcoes_mod.keys())
    entrada_nomes = set(r["funcao"] for r in registros)

    moi_sem_match = moi_nomes - entrada_nomes
    mod_sem_match = mod_nomes - entrada_nomes

    if moi_sem_match:
        print(f"\nFunções no MOI sem match na Tabela de Entrada: {moi_sem_match}")
    if mod_sem_match:
        print(f"Funções no MOD sem match na Tabela de Entrada: {mod_sem_match}")

    # Faixa salarial
    salarios = [r["salario_mensal"] for r in registros if r["salario_mensal"]]
    if salarios:
        print(f"\nFaixa salarial: R${min(salarios):,.2f} — R${max(salarios):,.2f}")
        print(f"Média: R${sum(salarios)/len(salarios):,.2f}")

    print("=" * 60)


if __name__ == "__main__":
    extrair()
