"""
Extração Passo 7: BASE DADOS PAINEIS → Insumos Siemens
30.557 itens da lista de preços Siemens (Jan/2019).
Preços líquidos (sem PIS, COFINS, ICMS).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import re
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
FONTE = BASE_DIR / "fontes" / "Or xxx - Planilhas Pequenas Obras_Lucro Presumido.xlsx"
SAIDA = BASE_DIR / "intermediarios" / "EXTRACAO_Pequenas_Obras_LP.xlsx"
NOME_FONTE = FONTE.name


def normalizar_texto(raw):
    if raw is None:
        return ""
    return re.sub(r"\s+", " ", str(raw).strip()).upper()


def extrair():
    print("Passo 7: BASE DADOS PAINEIS (Siemens)...")
    print("  Carregando planilha (30k+ linhas)...")

    wb = openpyxl.load_workbook(str(FONTE), read_only=True, data_only=True)
    ws = wb["BASE DADOS PAINEIS"]

    registros = []
    sem_preco = 0
    sem_desc = 0

    for i, row in enumerate(ws.iter_rows(min_row=6, values_only=False), 6):
        vals = {c.column: c.value for c in row if c.value is not None}
        if not vals or 2 not in vals:
            continue

        material_code = str(vals.get(2, "")).strip()
        product_number = str(vals.get(3, "")).strip() if vals.get(3) else ""
        descricao = normalizar_texto(vals.get(4, ""))
        grupo = normalizar_texto(vals.get(5, ""))
        ncm_raw = vals.get(6)
        disponibilidade = normalizar_texto(vals.get(7, ""))
        preco = vals.get(8) if isinstance(vals.get(8), (int, float)) else None

        if not descricao:
            sem_desc += 1
            continue

        if preco is None or preco <= 0:
            sem_preco += 1

        # NCM: pode ser numérico
        ncm = ""
        if ncm_raw:
            ncm = str(int(ncm_raw)) if isinstance(ncm_raw, (int, float)) else str(ncm_raw).strip()
            # Formatar NCM: 85015290 → 8501.52.90
            if len(ncm) == 8 and ncm.isdigit():
                ncm = f"{ncm[:4]}.{ncm[4:6]}.{ncm[6:]}"

        registros.append({
            "codigo_fonte": material_code,
            "codigo_mlfb": product_number,
            "nome": descricao,
            "unidade": "UN",
            "categoria": grupo,
            "preco_unitario": round(preco, 2) if preco else None,
            "data_referencia": "2019-01-07",
            "fabricante": "SIEMENS",
            "ncm": ncm,
            "disponibilidade": disponibilidade,
            "_fonte": NOME_FONTE,
            "_aba_origem": "BASE DADOS PAINEIS",
            "_linha_origem": i,
            "_confianca": "baixa",
            "_revisar": "",
            "_inclui_imposto": "NAO",
            "_observacao": "Preços líquidos Siemens Jan/2019 (sem PIS/COFINS/ICMS). Dados defasados.",
        })

        if len(registros) % 5000 == 0:
            print(f"  ... {len(registros)} registros processados")

    wb.close()
    print(f"  Total: {len(registros)} itens Siemens")
    print(f"  Sem preço: {sem_preco} | Sem descrição (ignorados): {sem_desc}")

    salvar(registros)

    # Relatório
    print(f"\n{'='*60}")
    print("RELATÓRIO — Base Dados Painéis (Siemens)")
    print(f"{'='*60}")
    print(f"Total registros: {len(registros)}")
    print(f"Sem preço: {sem_preco}")

    # Top 10 categorias
    cats = {}
    for r in registros:
        c = r["categoria"] or "SEM CATEGORIA"
        cats[c] = cats.get(c, 0) + 1
    print(f"Top categorias ({len(cats)} total):")
    for c, n in sorted(cats.items(), key=lambda x: -x[1])[:15]:
        print(f"  {c:30s} {n:>6d}")

    # Faixa de preço
    precos = [r["preco_unitario"] for r in registros if r["preco_unitario"]]
    if precos:
        print(f"Faixa preço: R${min(precos):,.2f} — R${max(precos):,.2f}")
        print(f"Média: R${sum(precos)/len(precos):,.2f}")
    print(f"{'='*60}")


def salvar(registros):
    colunas = [
        "codigo_fonte", "codigo_mlfb", "nome", "unidade", "categoria",
        "preco_unitario", "data_referencia", "fabricante", "ncm",
        "disponibilidade",
        "_fonte", "_aba_origem", "_linha_origem", "_confianca",
        "_revisar", "_inclui_imposto", "_observacao",
    ]

    # Para 30k+ registros, criar arquivo separado (evitar travar o principal)
    saida_paineis = SAIDA.parent / "EXTRACAO_Pequenas_Obras_LP_Paineis.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "InsumosSiemens"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    audit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, cn in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for ri, reg in enumerate(registros, 2):
        for ci, cn in enumerate(colunas, 1):
            val = reg.get(cn, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            # Sem border individual para performance em 30k rows
            if cn.startswith("_"):
                cell.fill = audit_fill

    larguras = {
        "codigo_fonte": 22, "codigo_mlfb": 24, "nome": 55, "unidade": 8,
        "categoria": 16, "preco_unitario": 14, "data_referencia": 14,
        "fabricante": 12, "ncm": 14, "disponibilidade": 16,
        "_fonte": 45, "_aba_origem": 20, "_linha_origem": 12,
        "_confianca": 10, "_revisar": 8, "_inclui_imposto": 14,
        "_observacao": 55,
    }
    for ci, cn in enumerate(colunas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = larguras.get(cn, 12)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(registros) + 1}"
    ws.freeze_panes = "A2"

    saida_paineis.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(saida_paineis))
    print(f"Salvo: {saida_paineis}")


if __name__ == "__main__":
    extrair()
