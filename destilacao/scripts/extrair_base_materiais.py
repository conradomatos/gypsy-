"""
Extração: Base de Materiais → Insumos
Fonte: Or xxx - Planilhas Pequenas Obras_Lucro Presumido.xlsx
Aba: Base de Materiais

Schema v2: familia / subfamilia / classe (hierarquia 3 níveis)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import re
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
FONTE = BASE_DIR / "fontes" / "Or xxx - Planilhas Pequenas Obras_Lucro Presumido.xlsx"
SAIDA = BASE_DIR / "intermediarios" / "EXTRACAO_Pequenas_Obras_LP.xlsx"
REVISAR = BASE_DIR / "validacao" / "REVISAR_Pequenas_Obras_LP_BaseMateriais.xlsx"

NOME_FONTE = FONTE.name
NOME_ABA = "Base de Materiais"

UNIDADE_MAP = {
    "pç": "PC", "pc": "PC", "ud": "UD", "un": "UN",
    "kg": "KG", "m": "M", "m²": "M2", "m³": "M3",
    "m2": "M2", "m3": "M3", "vb": "VB", "gb": "GB",
    "gl": "GL", "l": "L", "lt": "L", "br": "BR",
    "rl": "RL", "sc": "SC", "t": "T", "cj": "CJ",
    "dias": "H",
}

# --- Normalização de família (coluna A → familia nível 1) ---
FAMILIA_MAP = {
    "PÓRTICOS": "PORTICO",
    "P\xd3RTICOS": "PORTICO",
    "SUPORTES P/ EQUIP.": "SUPORTE",
    "FERRAGENS P/ SUPORTE": "FERRAGEM",
    "MAT. DE SEGURANÇA": "SEGURANCA",
    "MAT. DE SEGURAN\xc7A": "SEGURANCA",
    "CADEIA DE ANCORAGEM": "CADEIA ANCORAGEM",
}

# --- Regex para extrair subfamilia e classe da descrição ---

# Diâmetros: ø1/2", ø3/4", 1", 2.1/2", etc.
RE_DIAMETRO = re.compile(
    r'[øØ∅]?\s*(\d+(?:\.\d+)?(?:/\d+)?")', re.IGNORECASE
)

# Bitola mm²: 16mm², 120mm², 2.5mm², etc.
RE_BITOLA = re.compile(r'(\d+(?:[.,]\d+)?\s*mm[²2])', re.IGNORECASE)

# Dimensões LxAxC: 200x100x3000mm, 50x100mm, etc.
RE_DIMENSOES = re.compile(r'(\d+x\d+(?:x\d+)?)\s*mm', re.IGNORECASE)

# Tensão kV: 15KV, 69kV, 138KV, etc.
RE_TENSAO = re.compile(r'(\d+(?:\.\d+)?)\s*[kK][vV]')

# Potência W: 250W, 400W, 32W, etc.
RE_POTENCIA = re.compile(r'(\d+)\s*[wW]')

# Schedule: SCH 40, SCH 80
RE_SCH = re.compile(r'SCH\s*(\d+)', re.IGNORECASE)


def normalizar_unidade(raw):
    if raw is None:
        return ("", 1)
    txt = str(raw).strip().lower()
    if txt in UNIDADE_MAP:
        unidade = UNIDADE_MAP[txt]
        fator = 8 if txt == "dias" else 1
        return (unidade, fator)
    return (str(raw).strip().upper(), 1)


def normalizar_texto(raw):
    if raw is None:
        return ""
    txt = str(raw).strip()
    txt = re.sub(r"\s+", " ", txt)
    return txt.upper()


def normalizar_familia(raw):
    if raw is None:
        return ""
    txt = str(raw).strip().upper()
    return FAMILIA_MAP.get(txt, txt)


def extrair_subfamilia_classe(familia, descricao):
    """Extrai subfamilia (nível 2) e classe (nível 3) da descrição."""
    desc = descricao.upper() if descricao else ""
    subfamilia = ""
    classe = ""

    if familia == "ELETRODUTO":
        # Subfamilia: tipo de eletroduto
        if "GALVANIZADO A FOGO" in desc or "GALV" in desc:
            if "PESADO" in desc:
                subfamilia = "GALVANIZADO A FOGO PESADO"
            elif "LEVE" in desc or "SEMI" in desc:
                subfamilia = "GALVANIZADO A FOGO LEVE"
            else:
                subfamilia = "GALVANIZADO A FOGO"
        elif "PVC" in desc:
            if "RIGIDO" in desc or "RÍGIDO" in desc:
                subfamilia = "PVC RIGIDO"
            elif "CORRUGADO" in desc or "FLEX" in desc:
                subfamilia = "PVC CORRUGADO"
            else:
                subfamilia = "PVC"
        elif "FLEX" in desc:
            if "METAL" in desc or "SEALTUB" in desc:
                subfamilia = "METALICO FLEXIVEL"
            else:
                subfamilia = "FLEXIVEL"
        elif "ALUMIN" in desc:
            subfamilia = "ALUMINIO"
        # Classe: diâmetro
        m = RE_DIAMETRO.search(desc)
        if m:
            classe = m.group(1)

    elif familia == "CABO":
        # Subfamilia: tipo de cabo
        if "COBRE NU" in desc:
            subfamilia = "COBRE NU"
        elif "MEDIA" in desc or "MÉDIA" in desc or "M\xc9DIA" in desc:
            subfamilia = "MEDIA TENSAO"
        elif "UNIPOLAR" in desc:
            if "750" in desc:
                subfamilia = "UNIPOLAR 750V"
            elif "1KV" in desc or "1 KV" in desc:
                subfamilia = "UNIPOLAR 1KV"
            else:
                subfamilia = "UNIPOLAR"
        elif "TRIPOLAR" in desc:
            if "BLIND" in desc:
                subfamilia = "TRIPOLAR BLINDADO"
            else:
                subfamilia = "TRIPOLAR"
        elif "TETRAPOLAR" in desc or "4X" in desc or "4 X" in desc:
            subfamilia = "TETRAPOLAR"
        elif "CONTROLE" in desc:
            subfamilia = "CONTROLE"
        elif "FLEX" in desc:
            subfamilia = "COBRE FLEXIVEL"
        elif "SINGELO" in desc:
            subfamilia = "SINGELO"
        # Classe: bitola
        m = RE_BITOLA.search(desc)
        if m:
            classe = m.group(1).replace(" ", "").upper()
        else:
            # Tentar tensão para cabos MT
            m2 = RE_TENSAO.search(desc)
            if m2:
                classe = f"{m2.group(1)}KV"

    elif familia == "LEITO":
        if "GALVANIZADO" in desc or "GALV" in desc:
            subfamilia = "GALVANIZADO A FOGO"
        elif "FIBRA" in desc:
            subfamilia = "FIBRA DE VIDRO"
        elif "INOX" in desc:
            subfamilia = "ACO INOX"
        elif "ALUM" in desc:
            subfamilia = "ALUMINIO"
        # Tipo de peça
        if "CURVA" in desc:
            if "HORIZ" in desc:
                subfamilia += " CURVA HORIZONTAL" if subfamilia else "CURVA HORIZONTAL"
            elif "VERT" in desc:
                subfamilia += " CURVA VERTICAL" if subfamilia else "CURVA VERTICAL"
            else:
                subfamilia += " CURVA" if subfamilia else "CURVA"
        elif "REDU" in desc:
            subfamilia += " REDUCAO" if subfamilia else "REDUCAO"
        elif "CRUZETA" in desc or "DERIVACAO" in desc or "DERIV" in desc or "TEE" in desc or "T " in desc:
            subfamilia += " DERIVACAO" if subfamilia else "DERIVACAO"
        elif "TAMPA" in desc:
            subfamilia += " TAMPA" if subfamilia else "TAMPA"
        elif "TRECHO RETO" in desc or "RETO" in desc:
            subfamilia += " TRECHO RETO" if subfamilia else "TRECHO RETO"
        # Classe: dimensões
        m = RE_DIMENSOES.search(desc)
        if m:
            classe = m.group(1) + "MM"

    elif familia == "ELETROCALHA":
        if "GALVANIZADO" in desc or "GALV" in desc:
            subfamilia = "GALVANIZADA A FOGO"
        elif "INOX" in desc:
            subfamilia = "ACO INOX"
        # Tipo
        if "CURVA" in desc:
            subfamilia += " CURVA" if subfamilia else "CURVA"
        elif "REDU" in desc:
            subfamilia += " REDUCAO" if subfamilia else "REDUCAO"
        elif "DERIV" in desc or "TEE" in desc:
            subfamilia += " DERIVACAO" if subfamilia else "DERIVACAO"
        elif "TAMPA" in desc:
            subfamilia += " TAMPA" if subfamilia else "TAMPA"
        elif "RETO" in desc:
            subfamilia += " TRECHO RETO" if subfamilia else "TRECHO RETO"
        m = RE_DIMENSOES.search(desc)
        if m:
            classe = m.group(1) + "MM"

    elif familia == "CONDULETE":
        if "ALUM" in desc:
            subfamilia = "ALUMINIO"
        elif "PVC" in desc or "PLAST" in desc:
            subfamilia = "PVC"
        elif "FERRO" in desc:
            subfamilia = "FERRO"
        # Tipo (C, T, E, LB, LL, X, etc.)
        tipo_match = re.search(r'TIPO\s+"?([A-Z]{1,3})"?', desc)
        if tipo_match:
            subfamilia += f" TIPO {tipo_match.group(1)}" if subfamilia else f"TIPO {tipo_match.group(1)}"
        m = RE_DIAMETRO.search(desc)
        if m:
            classe = m.group(1)

    elif familia == "CONECTORES":
        if "PRENSA" in desc and "CABO" in desc:
            if "ALUM" in desc:
                subfamilia = "PRENSA-CABO ALUMINIO"
            elif "PVC" in desc:
                subfamilia = "PRENSA-CABO PVC"
            elif "INOX" in desc:
                subfamilia = "PRENSA-CABO INOX"
            else:
                subfamilia = "PRENSA-CABO"
        elif "TERMINAL" in desc or "COMPRESS" in desc:
            subfamilia = "TERMINAL COMPRESSAO"
        elif "SPLIT" in desc or "PARALEL" in desc:
            subfamilia = "SPLIT BOLT"
        elif "CUNHA" in desc:
            subfamilia = "CONECTOR CUNHA"
        m = RE_DIAMETRO.search(desc)
        if m:
            classe = m.group(1)
        else:
            m2 = RE_BITOLA.search(desc)
            if m2:
                classe = m2.group(1).replace(" ", "").upper()

    elif familia == "FIXADORES":
        if "ABRA" in desc and "VERGALH" in desc:
            subfamilia = "ABRAÇADEIRA VERGALHAO"
        elif "ABRA" in desc:
            subfamilia = "ABRAÇADEIRA"
        elif "BUCHA" in desc:
            subfamilia = "BUCHA"
        elif "CHUMBADOR" in desc:
            subfamilia = "CHUMBADOR"
        elif "PARAFUSO" in desc:
            subfamilia = "PARAFUSO"
        elif "PORCA" in desc:
            subfamilia = "PORCA"
        elif "ARRUELA" in desc:
            subfamilia = "ARRUELA"
        elif "BARRA ROSCA" in desc or "BARRA ROSCADA" in desc:
            subfamilia = "BARRA ROSCADA"
        m = RE_DIAMETRO.search(desc)
        if m:
            classe = m.group(1)

    elif familia == "TERMINAIS":
        if "POLIM" in desc:
            subfamilia = "MUFLA POLIMERICA"
        elif "TERMOC" in desc or "TERMOCONTR" in desc:
            subfamilia = "MUFLA TERMOCONTRATIL"
        elif "TERM" in desc and ("OLHAL" in desc or "COMPRES" in desc):
            subfamilia = "TERMINAL COMPRESSAO"
        m_tensao = RE_TENSAO.search(desc)
        m_bitola = RE_BITOLA.search(desc)
        partes = []
        if m_tensao:
            partes.append(f"{m_tensao.group(1)}KV")
        if m_bitola:
            partes.append(m_bitola.group(1).replace(" ", "").upper())
        if partes:
            classe = " ".join(partes)

    elif familia == "ILUMINACAO" or "ILUMINA" in familia:
        if "FLUORESC" in desc:
            subfamilia = "FLUORESCENTE"
        elif "PROJETOR" in desc:
            subfamilia = "PROJETOR"
        elif "ARANDELA" in desc:
            subfamilia = "ARANDELA"
        elif "INDUSTRIAL" in desc:
            subfamilia = "INDUSTRIAL"
        elif "PUBLIC" in desc or "PUBLICA" in desc:
            subfamilia = "PUBLICA"
        m = RE_POTENCIA.search(desc)
        if m:
            classe = f"{m.group(1)}W"

    elif familia == "PORTICO":
        if "METAL" in desc and "TRELIC" in desc:
            subfamilia = "METALICO TRELICADO"
        elif "COLUNA MET" in desc:
            subfamilia = "COLUNA METALICA"
        elif "VIGA MET" in desc:
            subfamilia = "VIGA METALICA"
        elif "CONCRETO" in desc:
            if "COLUNA" in desc:
                subfamilia = "COLUNA CONCRETO"
            elif "VIGA" in desc:
                subfamilia = "VIGA CONCRETO"
            elif "ANEL" in desc:
                subfamilia = "ANEL CONCRETO"
            else:
                subfamilia = "CONCRETO"
        m = RE_TENSAO.search(desc)
        if m:
            classe = f"{m.group(1)}KV"

    elif familia == "SUPORTE":
        if "COLUNA" in desc:
            subfamilia = "COLUNA"
        elif "CAPITEL" in desc:
            subfamilia = "CAPITEL"
        elif "BASE" in desc:
            subfamilia = "BASE"
        m = RE_TENSAO.search(desc)
        if m:
            classe = f"{m.group(1)}KV"

    elif familia == "ATERRAMENTO":
        if "HASTE" in desc:
            subfamilia = "HASTE"
        elif "CAIXA" in desc:
            subfamilia = "CAIXA INSPECAO"
        elif "SOLDA" in desc or "CADWELD" in desc:
            subfamilia = "SOLDA EXOTERMICA"
        elif "CONECTOR" in desc:
            subfamilia = "CONECTOR"
        m = RE_DIAMETRO.search(desc)
        if m:
            classe = m.group(1)

    elif familia == "FERRAGEM":
        if "BARRA CHATA" in desc:
            subfamilia = "BARRA CHATA"
        elif "PERFIL" in desc:
            subfamilia = "PERFIL"
        elif "TUBO" in desc:
            subfamilia = "TUBO"
        elif "CANTONEIRA" in desc:
            subfamilia = "CANTONEIRA"
        elif "CHAPA" in desc:
            subfamilia = "CHAPA"

    elif familia == "TUBOS":
        if "ALUM" in desc:
            subfamilia = "ALUMINIO"
        elif "PVC" in desc:
            subfamilia = "PVC"
        elif "A\xc7O" in desc or "ACO" in desc:
            subfamilia = "ACO"
        m = RE_DIAMETRO.search(desc)
        if m:
            classe = m.group(1)
        m2 = RE_SCH.search(desc)
        if m2:
            classe = (classe + " " if classe else "") + f"SCH{m2.group(1)}"

    elif familia == "CADEIA ANCORAGEM":
        if "ANCORAGEM" in desc:
            subfamilia = "ANCORAGEM"
        elif "SUSPENS" in desc:
            subfamilia = "SUSPENSAO"
        m = RE_TENSAO.search(desc)
        if m:
            classe = f"{m.group(1)}KV"

    elif familia == "PERFILADO":
        if "INOX" in desc:
            subfamilia = "ACO INOX"
        elif "GALV" in desc:
            subfamilia = "GALVANIZADO A FOGO"
        elif "CAIXA" in desc:
            subfamilia = "ACESSORIO"
        m = RE_DIMENSOES.search(desc)
        if m:
            classe = m.group(1) + "MM"

    elif familia == "ISOLADORES":
        subfamilia = "PEDESTAL PORCELANA"
        m = RE_TENSAO.search(desc)
        if m:
            classe = f"{m.group(1)}KV"

    elif familia == "SEGURANCA":
        if "EXTINTOR" in desc:
            if "CO2" in desc:
                subfamilia = "EXTINTOR CO2"
            elif "H2O" in desc or "AGUA" in desc or "ÁGUA" in desc:
                subfamilia = "EXTINTOR AGUA"
            else:
                subfamilia = "EXTINTOR PO QUIMICO"

    return subfamilia.strip(), classe.strip()


def extrair():
    print(f"Abrindo {FONTE.name}...")

    wb_val = openpyxl.load_workbook(str(FONTE), read_only=True, data_only=True)
    ws_val = wb_val[NOME_ABA]

    wb_form = openpyxl.load_workbook(str(FONTE), read_only=True, data_only=False)
    ws_form = wb_form[NOME_ABA]

    # Datas de cotação
    datas_cotacao = []
    for row in ws_val.iter_rows(min_row=5, max_row=5, min_col=18, max_col=42, values_only=True):
        for v in row:
            if isinstance(v, datetime):
                datas_cotacao.append(v)
    data_ref = max(datas_cotacao).strftime("%Y-%m-%d") if datas_cotacao else ""

    # Fórmulas
    formulas = {}
    for i, row in enumerate(ws_form.iter_rows(min_row=7, values_only=False), 7):
        row_formulas = {}
        for c in row:
            if c.value and isinstance(c.value, str) and c.value.startswith("="):
                if c.column in (5, 8, 16):
                    row_formulas[c.column] = c.value
        if row_formulas:
            formulas[i] = row_formulas
    wb_form.close()

    registros = []
    stats = {"total": 0, "sem_codigo": 0, "sem_descricao": 0, "sem_preco": 0,
             "revisar": 0, "por_familia": {}, "subfamilia_preenchida": 0, "classe_preenchida": 0}

    for i, row in enumerate(ws_val.iter_rows(min_row=7, values_only=False), 7):
        vals = {c.column: c.value for c in row if c.value is not None}

        if 2 not in vals and 3 not in vals:
            continue
        if 2 not in vals and 4 not in vals and 5 not in vals:
            continue

        stats["total"] += 1

        codigo = str(vals.get(2, "")).strip() if 2 in vals else ""
        descricao_raw = vals.get(3, "")
        unidade_raw = vals.get(4, "")
        preco_unit = vals.get(5)
        preco_fabricante = vals.get(7)
        hh_raw = vals.get(8)
        valor_medio = vals.get(16)
        familia_raw = vals.get(1, "")

        nome = normalizar_texto(descricao_raw)
        familia = normalizar_familia(familia_raw)
        unidade, fator_conv = normalizar_unidade(unidade_raw)

        # Subfamilia e classe
        subfamilia, classe = extrair_subfamilia_classe(familia, nome)
        if subfamilia:
            stats["subfamilia_preenchida"] += 1
        if classe:
            stats["classe_preenchida"] += 1

        # Preços
        preco_unit_final = round(preco_unit, 4) if isinstance(preco_unit, (int, float)) and preco_unit > 0 else None
        preco_fab_final = round(preco_fabricante, 4) if isinstance(preco_fabricante, (int, float)) and preco_fabricante > 0 else None
        hh_final = round(hh_raw * fator_conv, 4) if isinstance(hh_raw, (int, float)) else None
        valor_medio_final = round(valor_medio, 4) if isinstance(valor_medio, (int, float)) and not isinstance(valor_medio, bool) else None

        # Fórmulas
        row_formulas = formulas.get(i, {})
        formula_parts = []
        if 5 in row_formulas:
            formula_parts.append(f"preco_unit: {row_formulas[5]}")
        if 8 in row_formulas:
            formula_parts.append(f"hh: {row_formulas[8]}")
        if 16 in row_formulas:
            formula_parts.append(f"valor_medio: {row_formulas[16]}")
        formula_str = " | ".join(formula_parts) if formula_parts else ""

        # Revisão
        motivos_revisao = []
        if not codigo:
            motivos_revisao.append("SEM CODIGO")
            stats["sem_codigo"] += 1
        if not nome:
            motivos_revisao.append("SEM DESCRICAO")
            stats["sem_descricao"] += 1
        if preco_unit_final is None:
            stats["sem_preco"] += 1
        if unidade == "0":
            motivos_revisao.append("UNIDADE INVALIDA (0)")

        revisar = "SIM" if motivos_revisao else ""
        if revisar:
            stats["revisar"] += 1

        # Observação
        obs_parts = []
        if fator_conv != 1:
            obs_parts.append("Convertido de DIAS para H (x8)")
        if valor_medio_final and valor_medio_final > 0:
            obs_parts.append(f"Preco medio fornecedores: R${valor_medio_final:.2f}")
        if motivos_revisao:
            obs_parts.append(f"Motivo revisao: {', '.join(motivos_revisao)}")

        stats["por_familia"][familia] = stats["por_familia"].get(familia, 0) + 1

        registro = {
            "codigo_fonte": codigo,
            "nome": nome,
            "unidade": unidade,
            "familia": familia,
            "subfamilia": subfamilia,
            "classe": classe,
            "preco_unitario": preco_unit_final,
            "preco_fabricante": preco_fab_final,
            "hh_unitario": hh_final,
            "data_referencia": data_ref,
            "fabricante": "",
            "ncm": "",
            "_fonte": NOME_FONTE,
            "_aba_origem": NOME_ABA,
            "_linha_origem": i,
            "_confianca": "baixa",
            "_revisar": revisar,
            "_inclui_imposto": "",
            "_formula_original": formula_str,
            "_observacao": " | ".join(obs_parts) if obs_parts else "",
        }
        registros.append(registro)

    wb_val.close()
    print(f"Extraidos {len(registros)} registros")

    gerar_excel(registros, stats)

    revisar_registros = [r for r in registros if r["_revisar"] == "SIM"]
    if revisar_registros:
        gerar_excel_revisao(revisar_registros)

    gerar_relatorio(stats, registros)


def gerar_excel(registros, stats):
    colunas = [
        "codigo_fonte", "nome", "unidade", "familia", "subfamilia", "classe",
        "preco_unitario", "preco_fabricante", "hh_unitario",
        "data_referencia", "fabricante", "ncm",
        "_fonte", "_aba_origem", "_linha_origem", "_confianca",
        "_revisar", "_inclui_imposto", "_formula_original", "_observacao",
    ]

    try:
        if SAIDA.exists():
            wb = openpyxl.load_workbook(str(SAIDA))
            if "Insumos" in wb.sheetnames:
                del wb["Insumos"]
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
    except PermissionError:
        print(f"AVISO: {SAIDA.name} está aberto. Criando fallback.")
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet("Insumos", 0)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    audit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    revisar_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, reg in enumerate(registros, 2):
        for col_idx, col_name in enumerate(colunas, 1):
            val = reg[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_name.startswith("_"):
                cell.fill = audit_fill
            if reg["_revisar"] == "SIM":
                cell.fill = revisar_fill

    larguras = {
        "codigo_fonte": 16, "nome": 50, "unidade": 8,
        "familia": 22, "subfamilia": 30, "classe": 18,
        "preco_unitario": 14, "preco_fabricante": 16, "hh_unitario": 12,
        "data_referencia": 14, "fabricante": 14, "ncm": 12,
        "_fonte": 45, "_aba_origem": 20, "_linha_origem": 12,
        "_confianca": 10, "_revisar": 8, "_inclui_imposto": 14,
        "_formula_original": 40, "_observacao": 50,
    }
    for col_idx, col_name in enumerate(colunas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larguras.get(col_name, 12)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(registros) + 1}"
    ws.freeze_panes = "A2"

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(SAIDA))
        print(f"Salvo: {SAIDA}")
    except PermissionError:
        fallback = SAIDA.parent / "EXTRACAO_Pequenas_Obras_LP_Insumos.xlsx"
        wb.save(str(fallback))
        print(f"Salvo (fallback): {fallback}")


def gerar_excel_revisao(registros):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Revisar"

    colunas = [
        "codigo_fonte", "nome", "unidade", "familia", "subfamilia", "classe",
        "preco_unitario", "_linha_origem", "_revisar", "_observacao",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")

    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, reg in enumerate(registros, 2):
        for col_idx, col_name in enumerate(colunas, 1):
            ws.cell(row=row_idx, column=col_idx, value=reg[col_name])

    larguras = {"codigo_fonte": 16, "nome": 50, "unidade": 8, "familia": 22,
                "subfamilia": 30, "classe": 18, "preco_unitario": 14,
                "_linha_origem": 12, "_revisar": 8, "_observacao": 60}
    for col_idx, col_name in enumerate(colunas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larguras.get(col_name, 12)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(registros) + 1}"
    ws.freeze_panes = "A2"

    REVISAR.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(REVISAR))
    print(f"Salvo: {REVISAR}")


def gerar_relatorio(stats, registros):
    print("\n" + "=" * 60)
    print("RELATÓRIO DE EXTRAÇÃO — Base de Materiais (v2)")
    print("=" * 60)
    print(f"Fonte: {NOME_FONTE}")
    print(f"Aba:   {NOME_ABA}")
    print(f"Data:  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("-" * 60)
    print(f"Total registros:        {stats['total']}")
    print(f"Sem código:             {stats['sem_codigo']}")
    print(f"Sem descrição:          {stats['sem_descricao']}")
    print(f"Sem preço:              {stats['sem_preco']}")
    print(f"Para revisão:           {stats['revisar']}")
    print(f"Subfamília preenchida:  {stats['subfamilia_preenchida']}/{stats['total']} ({100*stats['subfamilia_preenchida']/stats['total']:.0f}%)")
    print(f"Classe preenchida:      {stats['classe_preenchida']}/{stats['total']} ({100*stats['classe_preenchida']/stats['total']:.0f}%)")
    print("-" * 60)
    print("Por família:")
    for fam, qtd in sorted(stats["por_familia"].items(), key=lambda x: -x[1]):
        # Contar subfamilias e classes por família
        sf_count = sum(1 for r in registros if r["familia"] == fam and r["subfamilia"])
        cl_count = sum(1 for r in registros if r["familia"] == fam and r["classe"])
        print(f"  {fam:30s} {qtd:>5d}  (subfam: {sf_count}, classe: {cl_count})")
    print("=" * 60)


if __name__ == "__main__":
    extrair()
