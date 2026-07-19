"""
Extração: Encargos + Impostos → Tabela EncargosBDI
Fontes:
  - Encargos (encargos sociais horista e mensalista)
  - Impostos-industrial (alíquotas de impostos por cenário)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
FONTE = BASE_DIR / "fontes" / "Or xxx - Planilhas Pequenas Obras_Lucro Presumido.xlsx"
SAIDA = BASE_DIR / "intermediarios" / "EXTRACAO_Pequenas_Obras_LP.xlsx"

NOME_FONTE = FONTE.name

# Mapeamento da aba Encargos — extraído manualmente do diagnóstico
ENCARGOS_HORISTA = [
    # (codigo, nome, grupo, aliquota, base_calculo, observacao, linha)
    ("A1", "PREVIDENCIA SOCIAL", "ENCARGO_SOCIAL", 0.20, "salario_base", "", 4),
    ("A2", "SESI / SESC", "ENCARGO_SOCIAL", 0.015, "salario_base", "", 5),
    ("A3", "SENAI / SENAC", "ENCARGO_SOCIAL", 0.01, "salario_base", "", 6),
    ("A4", "INCRA", "ENCARGO_SOCIAL", 0.002, "salario_base", "", 7),
    ("A5", "SEBRAE", "ENCARGO_SOCIAL", 0.006, "salario_base", "", 8),
    ("A6", "FGTS", "ENCARGO_SOCIAL", 0.08, "salario_base", "", 9),
    ("A7", "SEGURO ACIDENTES DE TRABALHO", "ENCARGO_SOCIAL", 0.03, "salario_base", "", 10),
    ("A8", "SALARIO EDUCACAO", "ENCARGO_SOCIAL", 0.025, "salario_base", "", 11),
    ("B1", "DESCANSO SEMANAL REMUNERADO", "ENCARGO_SOCIAL", 0.1665909090909091, "salario_base", "7.33/44 - Verificar turno (/36) ou administrativo (/44)", 14),
    ("B2", "FERIADOS NO ANO", "ENCARGO_SOCIAL", 0.0385, "salario_base", "10 feriados / 300 dias uteis", 15),
    ("B3", "FERIAS + GRATIFICACAO (1/3)", "ENCARGO_SOCIAL", 0.1264, "salario_base", "1 mes a cada 12 + 1/3", 16),
    ("B4", "AUXILIO DOENCA", "ENCARGO_SOCIAL", 0.01, "salario_base", "Media historica CONCEPT", 17),
    ("B5", "LICENCA PATERNIDADE", "ENCARGO_SOCIAL", 0.0001388888888888889, "salario_base", "Media CONCEPT 2007", 18),
    ("B6", "FALTAS LEGAIS", "ENCARGO_SOCIAL", 0.002777777777777778, "salario_base", "1 falta/ano/funcionario", 19),
    ("B7", "ACIDENTE DE TRABALHO COM AFASTAMENTO", "ENCARGO_SOCIAL", 0.0006785714285714286, "salario_base", "57 dias em toda empresa em 2007", 20),
    ("B8", "EXAMES MEDICOS PERIODICOS", "ENCARGO_SOCIAL", 0.0033333333333333335, "salario_base", "1 dia/ano/funcionario", 21),
    ("B9", "13 SALARIO", "ENCARGO_SOCIAL", 0.08333333333333333, "salario_base", "1 mes a cada 12 meses", 22),
    ("C1", "MULTA 50% FGTS SOBRE RESCISAO", "ENCARGO_SOCIAL", 0.04, "salario_base", "50% de cada parcela depositada", 25),
    ("C3", "AVISO PREVIO INDENIZADO", "ENCARGO_SOCIAL", 0.06363636363636363, "salario_base", "Media 9 meses na empresa", 26),
    ("C4", "INDENIZACAO 7 DIAS AVISO PREVIO", "ENCARGO_SOCIAL", 0.004242424242424243, "salario_base", "7 dias indenizados", 27),
    ("C5", "INDENIZACAO ADICIONAL PRE-CONVENCAO", "ENCARGO_SOCIAL", 0.0008333333333333333, "salario_base", "2% dos funcionarios", 28),
    ("D1", "INCIDENCIA GRUPO A SOBRE GRUPO B", "ENCARGO_SOCIAL", 0.15888503549783556, "grupo_b", "Encargos sobre encargos", 32),
]

ENCARGOS_MENSALISTA = [
    ("A1M", "PREVIDENCIA SOCIAL (MENSALISTA)", "ENCARGO_SOCIAL", 0.20, "salario_base", "", 44),
    ("A2M", "SESI / SESC (MENSALISTA)", "ENCARGO_SOCIAL", 0.015, "salario_base", "", 45),
    ("A3M", "SENAI / SENAC (MENSALISTA)", "ENCARGO_SOCIAL", 0.01, "salario_base", "", 46),
    ("A4M", "INCRA (MENSALISTA)", "ENCARGO_SOCIAL", 0.002, "salario_base", "", 47),
    ("A5M", "SEBRAE (MENSALISTA)", "ENCARGO_SOCIAL", 0.006, "salario_base", "", 48),
    ("A6M", "FGTS (MENSALISTA)", "ENCARGO_SOCIAL", 0.08, "salario_base", "", 49),
    ("A7M", "SEGURO ACIDENTES DE TRABALHO (MENSALISTA)", "ENCARGO_SOCIAL", 0.03, "salario_base", "", 50),
    ("A8M", "SALARIO EDUCACAO (MENSALISTA)", "ENCARGO_SOCIAL", 0.025, "salario_base", "", 51),
    ("B3M", "FERIAS + GRATIFICACAO 1/3 (MENSALISTA)", "ENCARGO_SOCIAL", 0.1264, "salario_base", "", 56),
    ("B4M", "AUXILIO DOENCA (MENSALISTA)", "ENCARGO_SOCIAL", 0.01, "salario_base", "", 57),
    ("B5M", "LICENCA PATERNIDADE (MENSALISTA)", "ENCARGO_SOCIAL", 0.0001388888888888889, "salario_base", "", 58),
    ("B6M", "FALTAS LEGAIS (MENSALISTA)", "ENCARGO_SOCIAL", 0.002777777777777778, "salario_base", "", 59),
    ("B7M", "ACIDENTE DE TRABALHO (MENSALISTA)", "ENCARGO_SOCIAL", 0.0006785714285714286, "salario_base", "", 60),
    ("B8M", "EXAMES MEDICOS (MENSALISTA)", "ENCARGO_SOCIAL", 0.0033333333333333335, "salario_base", "", 61),
    ("B9M", "13 SALARIO (MENSALISTA)", "ENCARGO_SOCIAL", 0.08333333333333333, "salario_base", "", 62),
    ("C1M", "MULTA 50% FGTS (MENSALISTA)", "ENCARGO_SOCIAL", 0.04, "salario_base", "", 65),
    ("C3M", "AVISO PREVIO INDENIZADO (MENSALISTA)", "ENCARGO_SOCIAL", 0.06363636363636363, "salario_base", "", 66),
    ("C4M", "INDENIZACAO 7 DIAS (MENSALISTA)", "ENCARGO_SOCIAL", 0.004242424242424243, "salario_base", "", 67),
    ("C5M", "INDENIZACAO ADICIONAL (MENSALISTA)", "ENCARGO_SOCIAL", 0.0008333333333333333, "salario_base", "", 68),
    ("D1M", "INCIDENCIA GRUPO A SOBRE B (MENSALISTA)", "ENCARGO_SOCIAL", 0.08341158095238096, "grupo_b", "", 72),
]

# Totais
TOTAIS_ENCARGOS = [
    ("TOTAL_HORISTA", "TOTAL ENCARGOS HORISTA", "ENCARGO_SOCIAL", 1.0673499705627707, "salario_base", "Soma A+B+C+D horista", 34),
    ("TOTAL_MENSALISTA", "TOTAL ENCARGOS MENSALISTA", "ENCARGO_SOCIAL", 0.786785606926407, "salario_base", "Soma A+B+C+D mensalista. Sem DSR (b1) e feriados (b2) pois ja inclusos no salario mensal", 74),
]

# Impostos (da aba Impostos-industrial, linha 8-9 headers, dados nas alíquotas)
IMPOSTOS = [
    ("COFINS", "COFINS", "IMPOSTO", 0.03, "receita_bruta", "Aliquota cumulativa (Lucro Presumido)", 8),
    ("PIS", "PIS", "IMPOSTO", 0.0065, "receita_bruta", "Aliquota cumulativa (Lucro Presumido)", 8),
    ("ICMS_MAT", "ICMS SOBRE MATERIAIS", "IMPOSTO", 0.18, "valor_materiais", "Incide apenas sobre materiais de montagem", 8),
    ("ISS", "ISS SOBRE SERVICOS", "IMPOSTO", 0.03, "valor_servicos", "Incide sobre servicos", 8),
    ("INSS_OBRA", "RETENCAO SEGURIDADE SOCIAL (CPRB)", "IMPOSTO", 0.035, "nota_fiscal_servicos", "Lei 12.546/2011 paragrafo 6 artigo 9", 25),
]


def extrair():
    print(f"Extraindo Encargos e Impostos...")

    registros = []

    # Encargos horista
    for cod, nome, grupo, aliq, base, obs, linha in ENCARGOS_HORISTA:
        registros.append({
            "codigo": cod,
            "nome": nome,
            "grupo": grupo,
            "regime": "HORISTA",
            "aliquota": aliq,
            "base_calculo": base,
            "observacao": obs,
            "_fonte": NOME_FONTE,
            "_aba_origem": "Encargos",
            "_linha_origem": linha,
        })

    # Encargos mensalista
    for cod, nome, grupo, aliq, base, obs, linha in ENCARGOS_MENSALISTA:
        registros.append({
            "codigo": cod,
            "nome": nome,
            "grupo": grupo,
            "regime": "MENSALISTA",
            "aliquota": aliq,
            "base_calculo": base,
            "observacao": obs,
            "_fonte": NOME_FONTE,
            "_aba_origem": "Encargos",
            "_linha_origem": linha,
        })

    # Totais
    for cod, nome, grupo, aliq, base, obs, linha in TOTAIS_ENCARGOS:
        registros.append({
            "codigo": cod,
            "nome": nome,
            "grupo": grupo,
            "regime": "TOTAL",
            "aliquota": aliq,
            "base_calculo": base,
            "observacao": obs,
            "_fonte": NOME_FONTE,
            "_aba_origem": "Encargos",
            "_linha_origem": linha,
        })

    # Impostos
    for cod, nome, grupo, aliq, base, obs, linha in IMPOSTOS:
        registros.append({
            "codigo": cod,
            "nome": nome,
            "grupo": grupo,
            "regime": "LUCRO_PRESUMIDO",
            "aliquota": aliq,
            "base_calculo": base,
            "observacao": obs,
            "_fonte": NOME_FONTE,
            "_aba_origem": "Impostos-industrial",
            "_linha_origem": linha,
        })

    print(f"  {len(registros)} registros")

    salvar_excel(registros)
    gerar_relatorio(registros)


def salvar_excel(registros):
    colunas = [
        "codigo", "nome", "grupo", "regime", "aliquota",
        "base_calculo", "observacao",
        "_fonte", "_aba_origem", "_linha_origem",
    ]

    try:
        if SAIDA.exists():
            wb = openpyxl.load_workbook(str(SAIDA))
            if "EncargosBDI" in wb.sheetnames:
                del wb["EncargosBDI"]
        else:
            wb = openpyxl.Workbook()
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
    except PermissionError:
        print(f"AVISO: {SAIDA.name} aberto. Criando fallback.")
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet("EncargosBDI")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    audit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, reg in enumerate(registros, 2):
        for col_idx, col_name in enumerate(colunas, 1):
            val = reg.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_name.startswith("_"):
                cell.fill = audit_fill

    larguras = {
        "codigo": 16, "nome": 50, "grupo": 18, "regime": 18,
        "aliquota": 12, "base_calculo": 18, "observacao": 60,
        "_fonte": 45, "_aba_origem": 20, "_linha_origem": 12,
    }
    for col_idx, col_name in enumerate(colunas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larguras.get(col_name, 12)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(registros) + 1}"
    ws.freeze_panes = "A2"

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(str(SAIDA))
        print(f"Salvo: {SAIDA} (aba EncargosBDI)")
    except PermissionError:
        fallback = SAIDA.parent / "EXTRACAO_Pequenas_Obras_LP_EncargosBDI.xlsx"
        wb.save(str(fallback))
        print(f"Salvo (fallback): {fallback}")


def gerar_relatorio(registros):
    print("\n" + "=" * 60)
    print("RELATÓRIO DE EXTRAÇÃO — Encargos e Impostos")
    print("=" * 60)
    print(f"Fonte: {NOME_FONTE}")
    print(f"Abas:  Encargos + Impostos-industrial")
    print(f"Data:  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("-" * 60)
    print(f"Total registros: {len(registros)}")

    por_grupo = {}
    for r in registros:
        g = r["grupo"]
        por_grupo[g] = por_grupo.get(g, 0) + 1
    print(f"Por grupo: {por_grupo}")

    por_regime = {}
    for r in registros:
        g = r["regime"]
        por_regime[g] = por_regime.get(g, 0) + 1
    print(f"Por regime: {por_regime}")

    # Totais
    print("\nTotais:")
    for r in registros:
        if r["regime"] == "TOTAL":
            print(f"  {r['nome']}: {r['aliquota']:.4%}")

    print("\nImpostos (Lucro Presumido):")
    for r in registros:
        if r["grupo"] == "IMPOSTO":
            print(f"  {r['nome']}: {r['aliquota']:.2%} sobre {r['base_calculo']}")

    print("=" * 60)


if __name__ == "__main__":
    extrair()
