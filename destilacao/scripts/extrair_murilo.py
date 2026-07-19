"""
Extração completa da planilha MURILO (ex-FAZZER ZIP).
Gera intermediarios/EXTRACAO_MURILO.xlsx com abas:
  - Insumos (de Base Dados)
  - MaoDeObra (de Profissionais + Valores Custo-Venda)
  - EncargosBDI (de Encargos Sociais + Incidências MOI/MOD + Memória Cálculo PV)
  - Equipamentos (de Máq. e Ferramentas + Equipamentos)
  - Composicoes (de Composição MOI + MOD)
  - ItensComposicao (breakdown das composições)
"""
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import pandas as pd
import numpy as np
import os
import re

BASE = r"D:\00 - CLAUDE_CODE\01. PROJETOS\04. ORCAMENTACAO_POR_MODELAGEM_FINANCEIRA\01_DESTILACAO"
FONTE = os.path.join(BASE, "fontes", "PLANILHA ORCAMENTOS - MURILO.xlsm")
SAIDA = os.path.join(BASE, "intermediarios", "EXTRACAO_MURILO.xlsx")
NOME_FONTE = "PLANILHA ORCAMENTOS - MURILO.xlsm"

# Normalização de unidades (UPPERCASE, sem espaços)
UNIDADE_MAP = {
    'pç': 'PC', 'pc': 'PC', 'pç ': 'PC',
    'un': 'UN', 'unid.': 'UN', 'unid': 'UN',
    'm': 'M', 'mt': 'M', 'm1': 'M',
    'm2': 'M2', 'm²': 'M2', ' m2': 'M2',
    'm3': 'M3', 'm³': 'M3', '5m3': 'M3',
    'kg': 'KG',
    'h': 'H', 'hh': 'H',
    'vb': 'VB',
    'cj': 'CJ',
    'l': 'L', 'gl': 'L',
    'br': 'BR',
    'al': 'AL',
    'fr': 'FR',
    'kw': 'KW',
    'pcte': 'PCTE',
    '0': '',
}

def norm_unidade(u):
    if pd.isna(u) or u is None:
        return ''
    s = str(u).strip().lower()
    return UNIDADE_MAP.get(s, s.upper())

def norm_texto(t):
    if pd.isna(t) or t is None:
        return ''
    s = str(t).strip().upper()
    s = re.sub(r'\s+', ' ', s)
    return s

def norm_familia(f):
    if pd.isna(f) or f is None or str(f).strip().upper() == 'FAMÍLIA':
        return ''
    return norm_texto(f)


# =====================================================================
# PASSO 1: BASE DADOS → INSUMOS
# =====================================================================
print("=" * 60)
print("PASSO 1: (2) Base Dados → Insumos")
print("=" * 60)

df_bd = pd.read_excel(FONTE, sheet_name='(2) Base Dados', header=None, engine='openpyxl')

# Dados começam na linha index 8 (Excel L9)
# Colunas: B=1(QTDE), C=2(FAMÍLIA), D=3(CÓDIGO), E=4(DESCRIÇÃO), F=5(Unid), G=6(Hh), H=7(Preço)
# Filtrar: tem código E descrição (exclui headers de seção e linhas vazias)
mask = df_bd.iloc[:, 3].notna() & df_bd.iloc[:, 4].notna()
dados = df_bd[mask].copy()
# Excluir a linha de header "FAMÍLIA/CÓDIGO" (geralmente index 1 e 6)
dados = dados[dados.iloc[:, 3].astype(str).str.strip() != 'CÓDIGO']

# Identificar seções (subheaders) para enriquecer subfamília
# Seções são linhas onde col 4 tem texto mas col 3 é NaN
secoes_df = df_bd[(df_bd.iloc[:, 4].notna()) & (df_bd.iloc[:, 3].isna())]
secoes = []
for idx, row in secoes_df.iterrows():
    txt = str(row.iloc[4]).strip()
    if txt and txt not in ('CAIXA DE PESQUISA NO BANCO DE DADOS', 'DESCRIÇÃO'):
        secoes.append((idx, txt))

# Atribuir seção a cada linha de dados
def get_secao(row_idx):
    """Encontra a seção mais recente antes desta linha."""
    last = ''
    for sec_idx, sec_name in secoes:
        if sec_idx < row_idx:
            last = sec_name
        else:
            break
    return last

insumos = []
for idx, row in dados.iterrows():
    codigo = str(row.iloc[3]).strip()
    descricao = norm_texto(row.iloc[4])
    familia = norm_familia(row.iloc[2])
    unidade = norm_unidade(row.iloc[5])
    hh = row.iloc[6] if pd.notna(row.iloc[6]) else None
    preco = row.iloc[7] if pd.notna(row.iloc[7]) else None
    secao = get_secao(idx)

    # HH: se for numérico, usar; senão, vazio
    if hh is not None:
        try:
            hh = float(hh)
        except (ValueError, TypeError):
            hh = None

    if preco is not None:
        try:
            preco = float(preco)
        except (ValueError, TypeError):
            preco = None

    # Confiança
    confianca = 'alta'
    revisar = ''
    obs = ''

    if not familia:
        confianca = 'media'
        revisar = 'SIM'
        obs = 'Sem família atribuída'

    if preco is not None and preco <= 0:
        confianca = 'baixa'
        revisar = 'SIM'
        obs = (obs + '; ' if obs else '') + f'Preço zero ou negativo: {preco}'

    if hh is not None and hh > 500:
        revisar = 'SIM'
        obs = (obs + '; ' if obs else '') + f'HH muito alto: {hh}'

    insumos.append({
        'codigo_fonte': codigo,
        'nome': descricao,
        'unidade': unidade,
        'familia': familia,
        'subfamilia': norm_texto(secao) if secao else '',
        'classe': '',
        'preco_unitario': preco,
        'preco_fabricante': None,
        'hh_unitario': hh,
        'data_referencia': '',
        'fabricante': '',
        'ncm': '',
        '_fonte': NOME_FONTE,
        '_aba_origem': '(2) Base Dados',
        '_linha_origem': idx + 1,  # +1 para Excel (1-based, mas pandas já é 0-based do read)
        '_confianca': confianca,
        '_revisar': revisar,
        '_inclui_imposto': '',
        '_observacao': obs,
    })

df_insumos = pd.DataFrame(insumos)
print(f"  Total: {len(df_insumos)} insumos")
print(f"  Com preço: {df_insumos['preco_unitario'].notna().sum()}")
print(f"  Com HH: {df_insumos['hh_unitario'].notna().sum()}")
print(f"  Para revisão: {(df_insumos['_revisar'] == 'SIM').sum()}")
print(f"  Famílias: {df_insumos['familia'].nunique()}")


# =====================================================================
# PASSO 2: PROFISSIONAIS (SALÁRIOS) → MÃO DE OBRA
# =====================================================================
print("\n" + "=" * 60)
print("PASSO 2: (6) Profissionais (Salários) → Mão de Obra")
print("=" * 60)

df_prof = pd.read_excel(FONTE, sheet_name='(6) Profissionais (Salários)', header=None, engine='openpyxl')

# Estrutura: L7+ → C2=nº, C3=função, C5=nºhoras(220), C6=R$/hora, C7=salário
# Também ler Valores Custo-Venda para custo_hora_com_encargos
df_vcv = pd.read_excel(FONTE, sheet_name='(13) Valores Custo-Venda_HH', header=None, engine='openpyxl')

# Profissionais: dados a partir de index 6 (L7 Excel)
mao_de_obra = []
for idx in range(6, len(df_prof)):
    row = df_prof.iloc[idx]
    num = row.iloc[1]
    funcao = row.iloc[2]

    if pd.isna(funcao) or str(funcao).strip() == '':
        continue
    funcao_str = str(funcao).strip()
    if funcao_str.upper() in ('TABELA DE CARGOS', 'TABELA DE SALÁRIOS', ''):
        continue

    horas = row.iloc[4] if pd.notna(row.iloc[4]) else None
    custo_hora = row.iloc[5] if pd.notna(row.iloc[5]) else None
    salario = row.iloc[6] if pd.notna(row.iloc[6]) else None

    try:
        horas = float(horas) if horas else None
    except:
        horas = None
    try:
        custo_hora = float(custo_hora) if custo_hora else None
    except:
        custo_hora = None
    try:
        salario = float(salario) if salario else None
    except:
        salario = None

    # Buscar custo com encargos na aba Valores Custo-Venda
    # Mesma ordem de funções, linha correspondente
    custo_com_encargos = None
    venda_hora = None
    vcv_idx = idx  # mesma posição relativa
    if vcv_idx < len(df_vcv):
        vcv_row = df_vcv.iloc[vcv_idx]
        try:
            custo_com_encargos = float(vcv_row.iloc[4]) if pd.notna(vcv_row.iloc[4]) else None
        except:
            pass
        try:
            venda_hora = float(vcv_row.iloc[5]) if pd.notna(vcv_row.iloc[5]) else None
        except:
            pass

    # Classificar MOI/MOD pelo nome da função
    funcao_upper = funcao_str.upper()
    tipo_mo = 'MOI'
    funcoes_mod = ['SOLDADOR', 'MONTADOR', 'CALDEIREIRO', 'ENCANADOR', 'ELETRICISTA',
                   'AJUDANTE', 'MEIO OFICIAL', 'INSTRUMENTISTA', 'ISOLADOR', 'PINTOR',
                   'JATISTA', 'MAÇARIQUEIRO', 'ESMERILHADOR', 'OPERADOR', 'PEDREIRO',
                   'CARPINTEIRO', 'ARMADOR', 'SERVENTE', 'TORNEIRO', 'FRESADOR',
                   'MECÂNICO', 'FUNILEIRO', 'CABLISTA', 'RIGGER']
    for mod in funcoes_mod:
        if mod in funcao_upper:
            tipo_mo = 'MOD'
            break

    # Funções claramente MOI
    funcoes_moi = ['GERENTE', 'ENGENHEIRO', 'COORDENADOR', 'SUPERVISOR', 'TÉCNICO DE SEGURANÇA',
                   'MÉDICO', 'ENFERMEIRO', 'AUXILIAR DE ENFERMAGEM', 'OBSERVADOR',
                   'ALMOXARIFE', 'ADMINISTRATIVO', 'PLANEJADOR', 'TOPÓGRAFO',
                   'ENCARREGADO', 'MESTRE', 'INSPETOR']
    for moi in funcoes_moi:
        if moi in funcao_upper:
            tipo_mo = 'MOI'
            break

    confianca = 'alta'
    revisar = ''
    obs = ''
    if custo_hora is None and salario is None:
        confianca = 'baixa'
        revisar = 'SIM'
        obs = 'Sem salário nem custo/hora'

    mao_de_obra.append({
        'codigo_fonte': str(int(num)) if pd.notna(num) and isinstance(num, (int, float)) else '',
        'funcao': norm_texto(funcao_str),
        'tipo': 'CLT',
        'tipo_mo': tipo_mo,
        'salario_mensal': salario,
        'custo_hora': custo_hora,
        'periculosidade': None,
        'encargos_percentual': None,
        'custo_hora_com_encargos': custo_com_encargos,
        'preco_venda_hora': venda_hora,
        'dissidio_referencia': '2025',
        'carga_horaria_mensal': horas,
        '_fonte': NOME_FONTE,
        '_aba_origem': '(6) Profissionais (Salários)',
        '_linha_origem': idx + 1,
        '_confianca': confianca,
        '_revisar': revisar,
        '_formula_original': '',
        '_observacao': obs,
    })

df_mo = pd.DataFrame(mao_de_obra)
# Remover linhas sem função válida
df_mo = df_mo[df_mo['funcao'].str.len() > 0].reset_index(drop=True)
print(f"  Total: {len(df_mo)} funções")
print(f"  MOI: {(df_mo['tipo_mo'] == 'MOI').sum()} | MOD: {(df_mo['tipo_mo'] == 'MOD').sum()}")
print(f"  Com salário: {df_mo['salario_mensal'].notna().sum()}")
print(f"  Com custo/hora c/ encargos: {df_mo['custo_hora_com_encargos'].notna().sum()}")
print(f"  Para revisão: {(df_mo['_revisar'] == 'SIM').sum()}")


# =====================================================================
# PASSO 3: ENCARGOS + INCIDÊNCIAS + BDI → ENCARGOS_BDI
# =====================================================================
print("\n" + "=" * 60)
print("PASSO 3: Encargos + Incidências + Memória Cálc. PV → EncargosBDI")
print("=" * 60)

# 3a) Encargos Sociais
df_enc = pd.read_excel(FONTE, sheet_name='(8) Encargos Sociais', header=None, engine='openpyxl')
encargos = []
for idx in range(1, len(df_enc)):
    row = df_enc.iloc[idx]
    nome = row.iloc[1]
    aliquota = row.iloc[2]
    obs = row.iloc[3] if pd.notna(row.iloc[3]) else ''

    if pd.isna(nome) or str(nome).strip() == '':
        continue

    nome_str = str(nome).strip()
    if nome_str.upper() in ('ENCARGOS', 'TOTAL'):
        continue

    try:
        aliq = float(aliquota) if pd.notna(aliquota) else None
    except:
        aliq = None

    # Identificar código
    codigo = nome_str.split('(')[0].strip().replace(' ', '_').upper()[:20]

    encargos.append({
        'codigo': codigo,
        'nome': norm_texto(nome_str),
        'grupo': 'ENCARGO_SOCIAL',
        'aliquota': aliq,
        'base_calculo': 'salario_base',
        'observacao': str(obs).strip() if obs else '',
        '_fonte': NOME_FONTE,
        '_aba_origem': '(8) Encargos Sociais',
        '_linha_origem': idx + 1,
    })

# 3b) Incidências MOI
df_inc_moi = pd.read_excel(FONTE, sheet_name='(9) Incidências  MOI', header=None, engine='openpyxl')
for idx in range(4, len(df_inc_moi)):
    row = df_inc_moi.iloc[idx]
    codigo_inc = row.iloc[2] if pd.notna(row.iloc[2]) else ''
    nome_inc = row.iloc[3] if pd.notna(row.iloc[3]) else ''
    preco = row.iloc[4] if pd.notna(row.iloc[4]) else None
    rotatividade = row.iloc[5] if pd.notna(row.iloc[5]) else None
    custo_mensal = row.iloc[6] if pd.notna(row.iloc[6]) else None
    obs_inc = row.iloc[7] if pd.notna(row.iloc[7]) else ''

    if pd.isna(nome_inc) or str(nome_inc).strip() == '':
        continue
    nome_str = str(nome_inc).strip()
    if 'Custo Mensal por Pessoa' in nome_str or 'Detalhamento' in nome_str or 'Descrição' in nome_str:
        # É header de grupo ou subtotal
        if custo_mensal is not None and 'Custo Mensal' in nome_str:
            # Subtotal de grupo
            grupo_letra = str(codigo_inc).strip() if codigo_inc else ''
            encargos.append({
                'codigo': f'INCID_MOI_{grupo_letra}',
                'nome': f'TOTAL INCIDÊNCIAS MOI GRUPO {grupo_letra}',
                'grupo': 'INCIDENCIA_MOI',
                'aliquota': None,
                'base_calculo': f'custo_mensal_pessoa = {custo_mensal}',
                'observacao': f'Subtotal grupo {grupo_letra} - MOI',
                '_fonte': NOME_FONTE,
                '_aba_origem': '(9) Incidências  MOI',
                '_linha_origem': idx + 1,
            })
        continue

    try:
        custo_m = float(custo_mensal) if custo_mensal is not None else None
    except:
        custo_m = None
    try:
        preco_v = float(preco) if preco is not None else None
    except:
        preco_v = None

    encargos.append({
        'codigo': f'INCID_MOI_{str(codigo_inc).strip()}' if codigo_inc else '',
        'nome': norm_texto(nome_str),
        'grupo': 'INCIDENCIA_MOI',
        'aliquota': custo_m,
        'base_calculo': f'preco_bateria={preco_v}, rotatividade_meses={rotatividade}' if preco_v else '',
        'observacao': str(obs_inc).strip() if obs_inc else '',
        '_fonte': NOME_FONTE,
        '_aba_origem': '(9) Incidências  MOI',
        '_linha_origem': idx + 1,
    })

# 3c) Incidências MOD
df_inc_mod = pd.read_excel(FONTE, sheet_name='(11) Incidências  MOD', header=None, engine='openpyxl')
for idx in range(4, len(df_inc_mod)):
    row = df_inc_mod.iloc[idx]
    codigo_inc = row.iloc[2] if pd.notna(row.iloc[2]) else ''
    nome_inc = row.iloc[3] if pd.notna(row.iloc[3]) else ''
    preco = row.iloc[4] if pd.notna(row.iloc[4]) else None
    rotatividade = row.iloc[5] if pd.notna(row.iloc[5]) else None
    custo_mensal = row.iloc[6] if pd.notna(row.iloc[6]) else None
    obs_inc = row.iloc[7] if pd.notna(row.iloc[7]) else ''

    if pd.isna(nome_inc) or str(nome_inc).strip() == '':
        continue
    nome_str = str(nome_inc).strip()
    if 'Custo Mensal por Pessoa' in nome_str or 'Detalhamento' in nome_str or 'Descrição' in nome_str:
        if custo_mensal is not None and 'Custo Mensal' in nome_str:
            grupo_letra = str(codigo_inc).strip() if codigo_inc else ''
            encargos.append({
                'codigo': f'INCID_MOD_{grupo_letra}',
                'nome': f'TOTAL INCIDÊNCIAS MOD GRUPO {grupo_letra}',
                'grupo': 'INCIDENCIA_MOD',
                'aliquota': None,
                'base_calculo': f'custo_mensal_pessoa = {custo_mensal}',
                'observacao': f'Subtotal grupo {grupo_letra} - MOD',
                '_fonte': NOME_FONTE,
                '_aba_origem': '(11) Incidências  MOD',
                '_linha_origem': idx + 1,
            })
        continue

    try:
        custo_m = float(custo_mensal) if custo_mensal is not None else None
    except:
        custo_m = None
    try:
        preco_v = float(preco) if preco is not None else None
    except:
        preco_v = None

    encargos.append({
        'codigo': f'INCID_MOD_{str(codigo_inc).strip()}' if codigo_inc else '',
        'nome': norm_texto(nome_str),
        'grupo': 'INCIDENCIA_MOD',
        'aliquota': custo_m,
        'base_calculo': f'preco_bateria={preco_v}, rotatividade_meses={rotatividade}' if preco_v else '',
        'observacao': str(obs_inc).strip() if obs_inc else '',
        '_fonte': NOME_FONTE,
        '_aba_origem': '(11) Incidências  MOD',
        '_linha_origem': idx + 1,
    })

# 3d) Memória Cálculo PV → BDI
df_bdi = pd.read_excel(FONTE, sheet_name='(26) Memória Cálculo - PV', header=None, engine='openpyxl')
for idx in range(1, len(df_bdi)):
    row = df_bdi.iloc[idx]
    # Varrer todas as colunas buscando dados
    id_val = row.iloc[1] if pd.notna(row.iloc[1]) else ''
    desc = row.iloc[2] if pd.notna(row.iloc[2]) else ''

    if pd.isna(desc) or str(desc).strip() == '':
        continue
    desc_str = str(desc).strip()
    if desc_str.upper() in ('MEMORIAL DE CÁLCULO', 'DESCRIÇÃO', ''):
        continue

    # Capturar valores numéricos nas colunas seguintes
    vals = {}
    for c in range(3, min(len(row), 13)):
        v = row.iloc[c]
        if pd.notna(v):
            try:
                vals[c] = float(v)
            except:
                vals[c] = str(v).strip()

    # Se tem algum valor numérico, registrar
    aliq = None
    for c in sorted(vals.keys()):
        if isinstance(vals[c], (int, float)):
            aliq = vals[c]
            break

    encargos.append({
        'codigo': f'BDI_{str(id_val).strip().replace(" ", "_").upper()[:15]}' if id_val else '',
        'nome': norm_texto(desc_str),
        'grupo': 'BDI_MARKUP',
        'aliquota': aliq,
        'base_calculo': 'custo_direto',
        'observacao': f'Valores por componente: {vals}' if vals else '',
        '_fonte': NOME_FONTE,
        '_aba_origem': '(26) Memória Cálculo - PV',
        '_linha_origem': idx + 1,
    })

df_encargos = pd.DataFrame(encargos)
print(f"  Total: {len(df_encargos)} registros")
enc_groups = df_encargos['grupo'].value_counts()
for g, c in enc_groups.items():
    print(f"    {g}: {c}")


# =====================================================================
# PASSO 4: MÁQ. E FERRAMENTAS + EQUIPAMENTOS → EQUIPAMENTOS
# =====================================================================
print("\n" + "=" * 60)
print("PASSO 4: Máq. e Ferramentas + Equipamentos → Equipamentos")
print("=" * 60)

equipamentos = []

# 4a) Máq. e Ferramentas
df_mf = pd.read_excel(FONTE, sheet_name='(23) Máq. e Ferramentas', header=None, engine='openpyxl')
for idx in range(3, len(df_mf)):
    row = df_mf.iloc[idx]
    num = row.iloc[1]
    desc = row.iloc[2]
    qtde = row.iloc[3] if pd.notna(row.iloc[3]) else None
    valor_mensal = row.iloc[4] if pd.notna(row.iloc[4]) else None
    meses = row.iloc[5] if pd.notna(row.iloc[5]) else None
    total = row.iloc[6] if pd.notna(row.iloc[6]) else None

    if pd.isna(desc) or str(desc).strip() == '':
        continue
    desc_str = str(desc).strip()
    if desc_str.upper() in ('DESCRIÇÃO', 'DESCRIÇÃO '):
        continue

    try:
        vm = float(valor_mensal) if valor_mensal is not None else None
    except:
        vm = None
    try:
        tot = float(total) if total is not None else None
    except:
        tot = None
    try:
        q = float(qtde) if qtde is not None else None
    except:
        q = None

    equipamentos.append({
        'codigo_fonte': str(int(num)) if pd.notna(num) and isinstance(num, (int, float)) else '',
        'nome': norm_texto(desc_str),
        'tipo': 'ALUGADO',
        'categoria': 'FERRAMENTA',
        'quantidade': q,
        'custo_mensal': vm,
        'custo_hora': None,
        'depreciacao_meses': None,
        'valor_aquisicao': None,
        'meses_uso': meses,
        'custo_total': tot,
        '_fonte': NOME_FONTE,
        '_aba_origem': '(23) Máq. e Ferramentas',
        '_linha_origem': idx + 1,
        '_confianca': 'alta' if vm else 'media',
        '_revisar': '',
        '_observacao': 'Template sem valores preenchidos' if not vm and not tot else '',
    })

# 4b) Equipamentos
df_eq = pd.read_excel(FONTE, sheet_name='(24) Equipamentos', header=None, engine='openpyxl')
for idx in range(3, len(df_eq)):
    row = df_eq.iloc[idx]
    num = row.iloc[1]
    desc = row.iloc[2]
    qtde = row.iloc[3] if pd.notna(row.iloc[3]) else None
    valor_mensal = row.iloc[4] if pd.notna(row.iloc[4]) else None
    meses = row.iloc[5] if pd.notna(row.iloc[5]) else None
    total = row.iloc[6] if pd.notna(row.iloc[6]) else None

    if pd.isna(desc) or str(desc).strip() == '':
        continue
    desc_str = str(desc).strip()
    if desc_str.upper() in ('DESCRIÇÃO', 'DESCRIÇÃO '):
        continue

    try:
        vm = float(valor_mensal) if valor_mensal is not None else None
    except:
        vm = None
    try:
        tot = float(total) if total is not None else None
    except:
        tot = None
    try:
        q = float(qtde) if qtde is not None else None
    except:
        q = None

    equipamentos.append({
        'codigo_fonte': str(int(num)) if pd.notna(num) and isinstance(num, (int, float)) else '',
        'nome': norm_texto(desc_str),
        'tipo': 'ALUGADO',
        'categoria': 'EQUIPAMENTO',
        'quantidade': q,
        'custo_mensal': vm,
        'custo_hora': None,
        'depreciacao_meses': None,
        'valor_aquisicao': None,
        'meses_uso': meses,
        'custo_total': tot,
        '_fonte': NOME_FONTE,
        '_aba_origem': '(24) Equipamentos',
        '_linha_origem': idx + 1,
        '_confianca': 'alta' if vm else 'media',
        '_revisar': '',
        '_observacao': 'Template sem valores preenchidos' if not vm and not tot else '',
    })

df_equip = pd.DataFrame(equipamentos)
print(f"  Total: {len(df_equip)} equipamentos")
print(f"  Ferramentas: {(df_equip['categoria'] == 'FERRAMENTA').sum()}")
print(f"  Equipamentos: {(df_equip['categoria'] == 'EQUIPAMENTO').sum()}")
print(f"  Com custo mensal: {df_equip['custo_mensal'].notna().sum()}")
print(f"  Para revisão: {(df_equip['_revisar'] == 'SIM').sum()}")


# =====================================================================
# PASSO 5: COMPOSIÇÃO MOI + MOD → COMPOSIÇÕES + ITENS
# =====================================================================
print("\n" + "=" * 60)
print("PASSO 5: Composição MOI + MOD → Composições + Itens")
print("=" * 60)

import openpyxl

# Ler com openpyxl para capturar fórmulas e valores
wb_val = openpyxl.load_workbook(FONTE, data_only=True, read_only=True)
wb_form = openpyxl.load_workbook(FONTE, data_only=False, read_only=True)

composicoes = []
itens_comp = []

def extrair_composicoes_mo(ws_val, ws_form, tipo_mo, aba_nome):
    """Extrai composições de MOI ou MOD.

    Estrutura real por bloco (~30 linhas):
    - L_start: " Função:" | " Salário Base:" | " Carga Horária:" | CLT/PJ
    - L_start+1: nº | nome_funcao | salário(C7) | carga(C10)
    - L_start+2: "DISCRIMINAÇÃO DE CUSTOS"
    - L_start+3: headers (Histórico | Custo Mensal | Custo Dia | Custo HN | Custo HE)
    - L_start+4: sub-headers (Item | Descrição | Und | Fator | HE50% | HE100%)
    - L_start+5...: itens (1.1 Básico, 1.2 ADN, 1.3 Periculosidade, 2.x Encargos, 3.x Custos Indiretos...)
    - L_total: "Custo Total [R$]" com valores em C7(mensal), C8(dia), C9(HN), C10(HE50%), C11(HE100%)
    - L_total+1: "Custo Total [R$] + ADICIONAL NOTURNO 20%" em C9, C10, C11
    """
    max_row = ws_val.max_row or 1

    # Detectar blocos: procurar "Função:" na coluna C3
    func_blocks = []
    current_start = None
    current_func = None

    for r in range(1, min(max_row + 1, 2000)):
        c3 = ws_val.cell(r, 3).value
        if c3 is not None and 'Função:' in str(c3):
            if current_func and current_start:
                func_blocks.append((current_func, current_start, r - 1))
            current_start = r
            current_func = None
        if current_start and r == current_start + 1 and c3 is not None:
            current_func = str(c3).strip()

    if current_func and current_start:
        func_blocks.append((current_func, current_start, min(current_start + 32, max_row)))

    print(f"    {aba_nome}: {len(func_blocks)} funções encontradas")

    for func_name, start, end in func_blocks:
        salario = ws_val.cell(start + 1, 7).value
        carga_h = ws_val.cell(start + 1, 10).value
        modalidade = ws_val.cell(start, 12).value  # CLT ou PJ

        codigo_comp = f'{tipo_mo}_{func_name.replace(" ", "_").upper()[:30]}'

        custo_hn = None
        custo_hn_adn = None
        custo_he50 = None
        custo_he100 = None
        custo_mensal = None

        itens_bloco = []
        for r in range(start + 2, min(end + 2, start + 35)):
            c3 = ws_val.cell(r, 3).value
            if c3 is None:
                continue
            c3_str = str(c3).strip()
            c3_upper = c3_str.upper()

            # Totais — "Custo Total [R$]" na C3, valores em C7-C11
            if 'CUSTO TOTAL' in c3_upper and 'ADICIONAL' not in c3_upper:
                try:
                    custo_mensal = float(ws_val.cell(r, 7).value)
                except:
                    pass
                try:
                    custo_hn = float(ws_val.cell(r, 9).value)
                except:
                    pass
                try:
                    custo_he50 = float(ws_val.cell(r, 10).value)
                except:
                    pass
                try:
                    custo_he100 = float(ws_val.cell(r, 11).value)
                except:
                    pass
                continue

            if 'CUSTO TOTAL' in c3_upper and 'ADICIONAL' in c3_upper:
                try:
                    custo_hn_adn = float(ws_val.cell(r, 9).value)
                except:
                    pass
                continue

            # Pular headers
            if c3_upper in ('DISCRIMINAÇÃO DE CUSTOS', 'HISTÓRICO', 'ITEM', ''):
                continue
            # Pular headers de grupo sem valor (ex: "1. Salário", "2. Encargos")
            if re.match(r'^\d+\.$', c3_str):
                continue

            # Item da composição — pegar valores de múltiplas colunas
            c4 = ws_val.cell(r, 4).value  # Descrição detalhada
            c5 = ws_val.cell(r, 5).value  # Unidade (R$, %)
            c6 = ws_val.cell(r, 6).value  # Fator
            c7 = ws_val.cell(r, 7).value  # Custo Mensal
            c9 = ws_val.cell(r, 9).value  # Custo Hora Normal

            # Usar C4 como descrição se disponível, senão C3
            item_nome = str(c4).strip() if pd.notna(c4) and c4 else c3_str
            item_codigo = c3_str  # Ex: "1.1", "2.1", "3.4"

            try:
                custo_mensal_item = float(c7) if c7 is not None else None
            except:
                custo_mensal_item = None
            try:
                custo_hora_item = float(c9) if c9 is not None else None
            except:
                custo_hora_item = None
            try:
                fator = float(c6) if c6 is not None else None
            except:
                fator = None

            unid_item = str(c5).strip() if pd.notna(c5) and c5 else ''

            # Fórmula
            c9_f = ws_form.cell(r, 9).value
            formula = str(c9_f)[:200] if c9_f and isinstance(c9_f, str) and c9_f.startswith('=') else ''

            itens_bloco.append({
                'composicao_codigo': codigo_comp,
                'tipo_item': 'MAO_DE_OBRA',
                'item_codigo': item_codigo,
                'item_nome': norm_texto(item_nome),
                'quantidade': fator,
                'unidade': unid_item.upper() if unid_item else '',
                'custo_unitario': custo_mensal_item,
                'custo_hora': custo_hora_item,
                '_fonte': NOME_FONTE,
                '_aba_origem': aba_nome,
                '_linha_origem': r,
                '_formula_original': formula,
            })

        composicoes.append({
            'codigo': codigo_comp,
            'nome': norm_texto(func_name),
            'unidade': 'H',
            'tipo': tipo_mo,
            'modalidade': str(modalidade).strip() if pd.notna(modalidade) else 'CLT',
            'custo_unitario_calculado': custo_hn,
            'custo_hn_com_adn': custo_hn_adn,
            'custo_he50': custo_he50,
            'custo_he100': custo_he100,
            'custo_mensal': custo_mensal,
            'salario_base': salario,
            'carga_horaria': carga_h,
            '_fonte': NOME_FONTE,
            '_aba_origem': aba_nome,
            '_linha_origem': start + 1,
            '_confianca': 'alta' if custo_hn else 'media',
            '_revisar': '' if custo_hn else 'SIM',
            '_formula_original': '',
        })

        itens_comp.extend(itens_bloco)

# MOI
ws_moi_v = wb_val['(10) Composição MOI']
ws_moi_f = wb_form['(10) Composição MOI']
extrair_composicoes_mo(ws_moi_v, ws_moi_f, 'MOI', '(10) Composição MOI')

# MOD
ws_mod_v = wb_val['(12) Composição MOD']
ws_mod_f = wb_form['(12) Composição MOD']
extrair_composicoes_mo(ws_mod_v, ws_mod_f, 'MOD', '(12) Composição MOD')

wb_val.close()
wb_form.close()

df_comp = pd.DataFrame(composicoes)
df_itens = pd.DataFrame(itens_comp) if itens_comp else pd.DataFrame()
print(f"  Total composições: {len(df_comp)}")
print(f"  MOI: {(df_comp['tipo'] == 'MOI').sum()} | MOD: {(df_comp['tipo'] == 'MOD').sum()}")
print(f"  Total itens: {len(df_itens)}")
print(f"  Com custo HN calculado: {df_comp['custo_unitario_calculado'].notna().sum()}")
print(f"  Com custo HE50%: {df_comp['custo_he50'].notna().sum()}")

# Mostrar amostra
if len(df_comp) > 0:
    sample = df_comp[df_comp['custo_unitario_calculado'].notna()][['nome','tipo','custo_unitario_calculado','custo_he50','salario_base']].head(5)
    print(f"\n  Amostra:")
    for _, r in sample.iterrows():
        print(f"    {r['tipo']} {r['nome']}: HN={r['custo_unitario_calculado']}, HE50={r['custo_he50']}, Sal={r['salario_base']}")


# =====================================================================
# SALVAR TUDO
# =====================================================================
print("\n" + "=" * 60)
print("SALVANDO EXTRACAO_MURILO.xlsx")
print("=" * 60)

with pd.ExcelWriter(SAIDA, engine='openpyxl') as writer:
    df_insumos.to_excel(writer, sheet_name='Insumos', index=False)
    df_mo.to_excel(writer, sheet_name='MaoDeObra', index=False)
    df_encargos.to_excel(writer, sheet_name='EncargosBDI', index=False)
    df_equip.to_excel(writer, sheet_name='Equipamentos', index=False)
    df_comp.to_excel(writer, sheet_name='Composicoes', index=False)
    if len(df_itens) > 0:
        df_itens.to_excel(writer, sheet_name='ItensComposicao', index=False)

print(f"\nArquivo salvo: {SAIDA}")
print(f"Tamanho: {os.path.getsize(SAIDA) / 1024:.0f} KB")

# Resumo final
print("\n" + "=" * 60)
print("RESUMO FINAL")
print("=" * 60)
print(f"  Insumos:          {len(df_insumos):>6} registros")
print(f"  Mão de Obra:      {len(df_mo):>6} funções")
print(f"  Encargos/BDI:     {len(df_encargos):>6} registros")
print(f"  Equipamentos:     {len(df_equip):>6} itens")
print(f"  Composições:      {len(df_comp):>6} composições")
print(f"  Itens Composição: {len(df_itens):>6} itens")
print(f"  TOTAL:            {len(df_insumos)+len(df_mo)+len(df_encargos)+len(df_equip)+len(df_comp)+len(df_itens):>6} registros")

# Revisões pendentes
total_rev = (
    (df_insumos['_revisar'] == 'SIM').sum() +
    (df_mo['_revisar'] == 'SIM').sum() +
    (df_equip['_revisar'] == 'SIM').sum() +
    (df_comp['_revisar'] == 'SIM').sum()
)
print(f"\n  Para revisão humana: {total_rev} itens")
