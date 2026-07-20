"""
Extração Passo 4: Equipamentos
Fontes:
  - Equipamentos (ferramentas e automotores - custo/dia)
  - Cálculo Custo de Equipamentos (caminhões, guindastes, containers com depreciação)
  - EPI e Ferramentas (EPIs com depreciação)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import re
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
FONTE = BASE_DIR / "fontes" / "Or xxx - Planilhas Pequenas Obras_Lucro Presumido.xlsx"
SAIDA = BASE_DIR / "intermediarios" / "EXTRACAO_Pequenas_Obras_LP.xlsx"
NOME_FONTE = FONTE.name


def normalizar_texto(raw):
    if raw is None:
        return ""
    txt = str(raw).strip()
    txt = re.sub(r"\s+", " ", txt)
    return txt.upper()


def extrair_aba_equipamentos(wb):
    """Aba Equipamentos — ferramentas e automotores com custo/dia."""
    ws = wb["Equipamentos"]
    registros = []

    for i, row in enumerate(ws.iter_rows(min_row=9, max_row=48, values_only=False), 9):
        vals = {c.column: c.value for c in row if c.value is not None}
        if 2 not in vals:
            continue
        item = vals.get(1)
        if not isinstance(item, (int, float)):
            continue

        nome = normalizar_texto(vals.get(2, ""))
        custo_dia = vals.get(6)
        custo_mensal = None
        custo_hora = None

        if isinstance(custo_dia, (int, float)) and custo_dia > 0:
            custo_dia = round(custo_dia, 2)
            custo_mensal = round(custo_dia * 30, 2)
            custo_hora = round(custo_dia / 10, 4)  # 10h/dia
        else:
            custo_dia = None

        registros.append({
            "codigo_fonte": "",
            "nome": nome,
            "tipo": "ALUGADO" if "ALUGADO" in nome or "ALUG" in nome else "PROPRIO",
            "custo_mensal": custo_mensal,
            "custo_hora": custo_hora,
            "custo_diario": custo_dia,
            "depreciacao_meses": None,
            "valor_aquisicao": None,
            "_fonte": NOME_FONTE,
            "_aba_origem": "Equipamentos",
            "_linha_origem": i,
            "_confianca": "media",
            "_revisar": "",
            "_observacao": "",
        })

    return registros


def extrair_aba_calculo_equip(wb):
    """Aba Cálculo Custo de Equipamentos — depreciação detalhada."""
    ws = wb["C\xe1lculo Custo de Equipamentos"]
    registros = []

    for i, row in enumerate(ws.iter_rows(min_row=5, max_row=48, values_only=False), 5):
        vals = {c.column: c.value for c in row if c.value is not None}
        if 2 not in vals:
            continue
        item = vals.get(1)
        if not isinstance(item, (int, float)):
            continue

        nome_partes = [normalizar_texto(vals.get(2, ""))]
        acessorio = vals.get(3)
        if acessorio and isinstance(acessorio, str) and acessorio.strip():
            nome_partes.append(normalizar_texto(acessorio))
        nome = " + ".join(p for p in nome_partes if p)

        valor_equip = vals.get(4)
        valor_acess = vals.get(5)
        valor_total = 0
        if isinstance(valor_equip, (int, float)):
            valor_total += valor_equip
        if isinstance(valor_acess, (int, float)):
            valor_total += valor_acess

        custo_mensal = vals.get(6) if isinstance(vals.get(6), (int, float)) else None
        custo_hora = vals.get(12) if isinstance(vals.get(12), (int, float)) else None
        custo_mensal_calc = vals.get(13) if isinstance(vals.get(13), (int, float)) else None

        registros.append({
            "codigo_fonte": "",
            "nome": nome,
            "tipo": "PROPRIO",
            "custo_mensal": round(custo_mensal_calc, 2) if custo_mensal_calc else None,
            "custo_hora": round(custo_hora, 4) if custo_hora else None,
            "custo_diario": None,
            "depreciacao_meses": 60,
            "valor_aquisicao": round(valor_total, 2) if valor_total > 0 else None,
            "_fonte": NOME_FONTE,
            "_aba_origem": "Cálculo Custo de Equipamentos",
            "_linha_origem": i,
            "_confianca": "media",
            "_revisar": "",
            "_observacao": "Depreciação 60 meses + Rem.Invest.1.5%/mês + Manut.15%/ano" if custo_mensal else "",
        })

    return registros


def extrair_aba_epi(wb):
    """Aba EPI e Ferramentas — EPIs e caixas de ferramentas com depreciação."""
    ws = wb["EPI e Ferramentas"]
    registros = []

    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=18, values_only=False), 2):
        vals = {c.column: c.value for c in row if c.value is not None}
        if 2 not in vals:
            continue
        item = vals.get(1)
        if item is None:
            continue

        nome = normalizar_texto(vals.get(2, ""))
        if not nome:
            continue

        valor_aquisicao = vals.get(3) if isinstance(vals.get(3), (int, float)) else None
        depreciacao = vals.get(4) if isinstance(vals.get(4), (int, float)) else None
        custo_mensal = vals.get(6) if isinstance(vals.get(6), (int, float)) else None

        registros.append({
            "codigo_fonte": "",
            "nome": nome,
            "tipo": "PROPRIO",
            "custo_mensal": round(custo_mensal, 2) if custo_mensal else None,
            "custo_hora": round(custo_mensal / 220, 4) if custo_mensal else None,
            "custo_diario": None,
            "depreciacao_meses": depreciacao,
            "valor_aquisicao": round(valor_aquisicao, 2) if valor_aquisicao else None,
            "_fonte": NOME_FONTE,
            "_aba_origem": "EPI e Ferramentas",
            "_linha_origem": i,
            "_confianca": "alta",
            "_revisar": "",
            "_observacao": "",
        })

    return registros


def salvar_excel(registros):
    colunas = [
        "codigo_fonte", "nome", "tipo",
        "custo_mensal", "custo_hora", "custo_diario",
        "depreciacao_meses", "valor_aquisicao",
        "_fonte", "_aba_origem", "_linha_origem", "_confianca",
        "_revisar", "_observacao",
    ]

    try:
        wb = openpyxl.load_workbook(str(SAIDA))
        if "Equipamentos" in wb.sheetnames:
            del wb["Equipamentos"]
    except (FileNotFoundError, PermissionError):
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = wb.create_sheet("Equipamentos")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    audit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, col_name in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for row_idx, reg in enumerate(registros, 2):
        for col_idx, col_name in enumerate(colunas, 1):
            val = reg.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if col_name.startswith("_"):
                cell.fill = audit_fill

    larguras = {
        "codigo_fonte": 14, "nome": 70, "tipo": 14,
        "custo_mensal": 14, "custo_hora": 12, "custo_diario": 12,
        "depreciacao_meses": 16, "valor_aquisicao": 16,
        "_fonte": 45, "_aba_origem": 28, "_linha_origem": 12,
        "_confianca": 10, "_revisar": 8, "_observacao": 50,
    }
    for col_idx, col_name in enumerate(colunas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = larguras.get(col_name, 12)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(registros) + 1}"
    ws.freeze_panes = "A2"

    try:
        wb.save(str(SAIDA))
        print(f"Salvo: {SAIDA} (aba Equipamentos)")
    except PermissionError:
        fallback = SAIDA.parent / "EXTRACAO_Pequenas_Obras_LP_Equipamentos.xlsx"
        wb.save(str(fallback))
        print(f"Salvo (fallback): {fallback}")


def extrair():
    print("Passo 4: Equipamentos...")
    wb = openpyxl.load_workbook(str(FONTE), read_only=True, data_only=True)

    r1 = extrair_aba_equipamentos(wb)
    print(f"  Equipamentos: {len(r1)} itens")

    r2 = extrair_aba_calculo_equip(wb)
    print(f"  Cálculo Custo Equipamentos: {len(r2)} itens")

    r3 = extrair_aba_epi(wb)
    print(f"  EPI e Ferramentas: {len(r3)} itens")

    wb.close()

    registros = r1 + r2 + r3
    print(f"  Total: {len(registros)} equipamentos")

    salvar_excel(registros)

    # Relatório
    print(f"\n{'='*60}")
    print("RELATÓRIO — Equipamentos")
    print(f"{'='*60}")
    por_aba = {}
    for r in registros:
        a = r["_aba_origem"]
        por_aba[a] = por_aba.get(a, 0) + 1
    for a, c in por_aba.items():
        print(f"  {a}: {c}")
    com_preco = sum(1 for r in registros if r["custo_mensal"])
    print(f"  Com custo mensal: {com_preco}/{len(registros)}")
    com_depre = sum(1 for r in registros if r["depreciacao_meses"])
    print(f"  Com depreciação: {com_depre}/{len(registros)}")
    print(f"{'='*60}")


if __name__ == "__main__":
    extrair()
