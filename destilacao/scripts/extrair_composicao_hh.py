"""
Extração Passo 5: COMPOSIÇÃO Hh → Composições + Itens da Composição
Composição salarial completa por função (salário + encargos + custos diversos + BDI).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import re
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
FONTE = BASE_DIR / "fontes" / "Or xxx - Planilhas Pequenas Obras_Lucro Presumido.xlsx"
SAIDA = BASE_DIR / "intermediarios" / "EXTRACAO_Pequenas_Obras_LP.xlsx"
NOME_FONTE = FONTE.name


def normalizar_texto(raw):
    if raw is None:
        return ""
    return re.sub(r"\s+", " ", str(raw).strip()).upper()


def extrair():
    print("Passo 5: COMPOSIÇÃO Hh...")
    wb = openpyxl.load_workbook(str(FONTE), read_only=True, data_only=True)
    ws = wb["COMPOSI\xc7\xc3O Hh"]

    # Linha 4: nomes das funções nas colunas 2-21
    funcoes = {}
    for row in ws.iter_rows(min_row=4, max_row=4, values_only=False):
        for c in row:
            if c.column >= 2 and c.value:
                funcoes[c.column] = normalizar_texto(c.value)

    # Componentes de custo (linhas 6-41)
    componentes_config = [
        (6, "SAL_MENSAL", "SALARIO MENSAL", "MAO_DE_OBRA"),
        (7, "PERIC_30", "PERICULOSIDADE 30%", "MAO_DE_OBRA"),
        (8, "ENCARGOS", "ENCARGOS SOCIAIS", "MAO_DE_OBRA"),
        (12, "CAFE", "CAFE MENSAL", "ALIMENTACAO"),
        (13, "ALMOCO", "ALMOCO MENSAL", "ALIMENTACAO"),
        (14, "JANTAR", "JANTAR MENSAL", "ALIMENTACAO"),
        (15, "VALE_ALIM", "VALE ALIMENTACAO SINDICATO", "ALIMENTACAO"),
        (20, "TRANSP_ONIBUS", "TRANSPORTE ONIBUS IDA/VOLTA", "TRANSPORTE"),
        (23, "TRANSP_LOCAL", "TRANSPORTE LOCAL", "TRANSPORTE"),
        (25, "EPI_UNIFORME", "EPI E UNIFORME", "EPI"),
        (28, "DESP_MEDICAS", "DESPESAS MEDICAS E EXAMES", "ADMISSIONAL"),
        (29, "SEGURO_VIDA", "SEGURO DE VIDA", "ADMISSIONAL"),
        (32, "FERRAM_INDIV", "FERRAMENTAS USO INDIVIDUAL", "FERRAMENTA"),
    ]

    # Ler valores de cada componente para cada função
    dados = {}
    for row_num, cod, nome, tipo in componentes_config:
        for row in ws.iter_rows(min_row=row_num, max_row=row_num, values_only=False):
            for c in row:
                if c.column in funcoes and isinstance(c.value, (int, float)):
                    col = c.column
                    if col not in dados:
                        dados[col] = {}
                    dados[col][cod] = round(c.value, 2)

    # Totais por função
    totais = {}
    # Linha 38: Custo funcionário/mês sem BDI
    for row in ws.iter_rows(min_row=38, max_row=38, values_only=False):
        for c in row:
            if c.column in funcoes and isinstance(c.value, (int, float)):
                totais.setdefault(c.column, {})["custo_mes_sem_bdi"] = round(c.value, 2)

    # Linha 40: BDI percentual
    for row in ws.iter_rows(min_row=40, max_row=40, values_only=False):
        for c in row:
            if c.column in funcoes and isinstance(c.value, (int, float)):
                totais.setdefault(c.column, {})["bdi_percentual"] = round(c.value, 6)

    # Linha 44: Custo HH sem BDI / 180
    for row in ws.iter_rows(min_row=44, max_row=44, values_only=False):
        for c in row:
            if c.column in funcoes and isinstance(c.value, (int, float)):
                totais.setdefault(c.column, {})["custo_hh_sem_bdi"] = round(c.value, 4)

    # Linha 45: Custo HH com BDI / 180
    for row in ws.iter_rows(min_row=45, max_row=45, values_only=False):
        for c in row:
            if c.column in funcoes and isinstance(c.value, (int, float)):
                totais.setdefault(c.column, {})["custo_hh_com_bdi"] = round(c.value, 4)

    wb.close()

    # Montar registros de Composição (uma por função)
    composicoes = []
    itens_composicao = []

    for col, funcao in funcoes.items():
        cod_comp = f"COMP_HH_{col-1:02d}"
        tot = totais.get(col, {})

        composicoes.append({
            "codigo": cod_comp,
            "nome": f"COMPOSICAO SALARIAL - {funcao}",
            "unidade": "H",
            "tipo": "MOI" if col <= 11 else "MOD",
            "custo_unitario_calculado": tot.get("custo_hh_sem_bdi"),
            "custo_unitario_com_bdi": tot.get("custo_hh_com_bdi"),
            "bdi_percentual": tot.get("bdi_percentual"),
            "custo_mensal_sem_bdi": tot.get("custo_mes_sem_bdi"),
            "_fonte": NOME_FONTE,
            "_aba_origem": "COMPOSIÇÃO Hh",
            "_linha_origem": 4,
            "_confianca": "alta",
            "_revisar": "",
            "_observacao": f"Divisor: 180h/mês",
        })

        # Itens da composição
        func_dados = dados.get(col, {})
        for row_num, cod_item, nome_item, tipo_item in componentes_config:
            valor = func_dados.get(cod_item)
            if valor and valor > 0:
                itens_composicao.append({
                    "composicao_codigo": cod_comp,
                    "tipo_item": tipo_item,
                    "item_codigo": cod_item,
                    "item_nome": nome_item,
                    "quantidade": 1,
                    "unidade": "MES",
                    "custo_unitario": valor,
                    "custo_total": valor,
                    "_fonte": NOME_FONTE,
                    "_aba_origem": "COMPOSIÇÃO Hh",
                    "_linha_origem": row_num,
                    "_formula_original": "",
                })

    print(f"  {len(composicoes)} composições")
    print(f"  {len(itens_composicao)} itens de composição")

    salvar(composicoes, itens_composicao)


def salvar(composicoes, itens_composicao):
    try:
        wb = openpyxl.load_workbook(str(SAIDA))
    except (FileNotFoundError, PermissionError):
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    audit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_sheet(name, registros, colunas, larguras):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        for ci, cn in enumerate(colunas, 1):
            cell = ws.cell(row=1, column=ci, value=cn)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border
        for ri, reg in enumerate(registros, 2):
            for ci, cn in enumerate(colunas, 1):
                val = reg.get(cn, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.border = thin_border
                if cn.startswith("_"):
                    cell.fill = audit_fill
        for ci, cn in enumerate(colunas, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = larguras.get(cn, 12)
        ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(registros) + 1}"
        ws.freeze_panes = "A2"

    write_sheet("Composicoes", composicoes,
        ["codigo", "nome", "unidade", "tipo", "custo_unitario_calculado",
         "custo_unitario_com_bdi", "bdi_percentual", "custo_mensal_sem_bdi",
         "_fonte", "_aba_origem", "_linha_origem", "_confianca", "_revisar", "_observacao"],
        {"codigo": 16, "nome": 50, "unidade": 8, "tipo": 8,
         "custo_unitario_calculado": 20, "custo_unitario_com_bdi": 20,
         "bdi_percentual": 14, "custo_mensal_sem_bdi": 20,
         "_fonte": 45, "_aba_origem": 20, "_linha_origem": 12,
         "_confianca": 10, "_revisar": 8, "_observacao": 30})

    write_sheet("ItensComposicao", itens_composicao,
        ["composicao_codigo", "tipo_item", "item_codigo", "item_nome",
         "quantidade", "unidade", "custo_unitario", "custo_total",
         "_fonte", "_aba_origem", "_linha_origem", "_formula_original"],
        {"composicao_codigo": 16, "tipo_item": 14, "item_codigo": 16,
         "item_nome": 40, "quantidade": 10, "unidade": 8,
         "custo_unitario": 14, "custo_total": 14,
         "_fonte": 45, "_aba_origem": 20, "_linha_origem": 12,
         "_formula_original": 30})

    try:
        wb.save(str(SAIDA))
        print(f"Salvo: {SAIDA} (abas Composicoes + ItensComposicao)")
    except PermissionError:
        fallback = SAIDA.parent / "EXTRACAO_Pequenas_Obras_LP_Composicoes.xlsx"
        wb.save(str(fallback))
        print(f"Salvo (fallback): {fallback}")

    # Relatório
    print(f"\n{'='*60}")
    print("RELATÓRIO — Composições Hh")
    print(f"{'='*60}")
    for c in composicoes:
        print(f"  {c['codigo']}: {c['nome'][:45]} — R${c['custo_unitario_calculado']:.2f}/h (sem BDI) | R${c['custo_unitario_com_bdi']:.2f}/h (com BDI)")
    print(f"{'='*60}")


if __name__ == "__main__":
    extrair()
