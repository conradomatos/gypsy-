"""
Extração Passo 6: Materiais - industrial + Alimentação-industrial
→ Composições implícitas (fórmulas com material + HH + preço)

Materiais - industrial: template de orçamento com ~92 linhas com dados reais.
  Cada linha = item com código, descrição, unidade, qtd, HH, preço.
  São composições implícitas (material + MO embutida via HH).

Alimentação-industrial: calculadora de alimentadores.
  14 alimentadores com cálculos automáticos de cabos/eletrodutos/terminais.
  Estrutura muito complexa — extrair como composição por alimentador.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
import re
from datetime import datetime

BASE_DIR = Path(__file__).resolve().parent.parent
FONTE = BASE_DIR / "fontes" / "Or xxx - Planilhas Pequenas Obras_Lucro Presumido.xlsx"
SAIDA = BASE_DIR / "intermediarios" / "EXTRACAO_Pequenas_Obras_LP.xlsx"
NOME_FONTE = FONTE.name


def normalizar_texto(raw):
    if raw is None:
        return ""
    return re.sub(r"\s+", " ", str(raw).strip()).upper()


def extrair_materiais_industrial(wb):
    """Extrai itens com dados reais da aba Materiais - industrial."""
    ws = wb["Materiais - industrial"]
    registros = []
    secao_atual = ""

    for i, row in enumerate(ws.iter_rows(min_row=11, values_only=False), 11):
        vals = {c.column: c.value for c in row if c.value is not None}
        if not vals:
            continue

        # Detectar headers de seção (col 2 com texto, sem col 3)
        if 2 in vals and 3 not in vals and isinstance(vals[2], str):
            secao_atual = normalizar_texto(vals[2])
            continue

        # Detectar sub-headers (col 1 tipo "1.1", "2.3" etc, col 2 = nome seção)
        item = vals.get(1)
        if isinstance(item, str) and "." in item and 3 not in vals:
            secao_atual = normalizar_texto(vals.get(2, secao_atual))
            continue

        # Item real: tem descrição (col 3) ou código (col 2 com formato de código)
        desc = vals.get(3)
        codigo = vals.get(2)
        if not desc and not codigo:
            continue
        if isinstance(desc, str) and desc.strip():
            pass
        elif isinstance(codigo, str) and len(codigo) > 3:
            desc = codigo  # usar código como desc se desc vazio
        else:
            continue

        nome = normalizar_texto(desc)
        cod = normalizar_texto(codigo) if codigo else ""
        unidade = str(vals.get(4, "")).strip().upper() if vals.get(4) else ""
        quantidade = vals.get(5) if isinstance(vals.get(5), (int, float)) else None
        fornecimento = normalizar_texto(vals.get(6, ""))
        hh_unit = vals.get(7) if isinstance(vals.get(7), (int, float)) else None
        hh_fator = vals.get(8) if isinstance(vals.get(8), (int, float)) else None
        hh_total = vals.get(9) if isinstance(vals.get(9), (int, float)) else None
        preco_unit = vals.get(10) if isinstance(vals.get(10), (int, float)) else None
        preco_total = vals.get(11) if isinstance(vals.get(11), (int, float)) else None
        preco_venda_unit = vals.get(13) if isinstance(vals.get(13), (int, float)) else None
        preco_venda_total = vals.get(14) if isinstance(vals.get(14), (int, float)) else None
        tipo_fat = normalizar_texto(vals.get(15, ""))
        curva_abc = normalizar_texto(vals.get(18, ""))

        # Pular linhas de MO dentro de seções (H/h como unidade sem código)
        if unidade in ("H/H", "HXH") and not cod:
            # É uma linha de MO dentro da composição, não um material
            registros.append({
                "codigo_fonte": cod,
                "nome": nome,
                "secao": secao_atual,
                "unidade": "H",
                "tipo_item": "MAO_DE_OBRA",
                "quantidade": quantidade,
                "hh_unitario": hh_unit,
                "hh_fator_dificuldade": hh_fator,
                "hh_total": hh_total,
                "preco_unitario": preco_unit,
                "preco_total": preco_total,
                "preco_venda_unit": preco_venda_unit,
                "fornecimento": fornecimento,
                "curva_abc": curva_abc,
                "_fonte": NOME_FONTE,
                "_aba_origem": "Materiais - industrial",
                "_linha_origem": i,
                "_confianca": "media",
                "_revisar": "",
                "_observacao": "Linha de MO dentro de composição implícita",
            })
            continue

        registros.append({
            "codigo_fonte": cod,
            "nome": nome,
            "secao": secao_atual,
            "unidade": unidade,
            "tipo_item": "INSUMO",
            "quantidade": quantidade,
            "hh_unitario": hh_unit,
            "hh_fator_dificuldade": hh_fator,
            "hh_total": hh_total,
            "preco_unitario": preco_unit,
            "preco_total": preco_total,
            "preco_venda_unit": preco_venda_unit,
            "fornecimento": fornecimento,
            "curva_abc": curva_abc,
            "_fonte": NOME_FONTE,
            "_aba_origem": "Materiais - industrial",
            "_linha_origem": i,
            "_confianca": "media",
            "_revisar": "SIM" if not cod and nome else "",
            "_observacao": "Composição implícita: material + HH embutido" if hh_unit else "",
        })

    return registros


def extrair_alimentacao(wb):
    """Extrai alimentadores da aba Alimentação-industrial."""
    ws = wb["Alimenta\xe7\xe3o-industrial"]
    registros = []

    for i, row in enumerate(ws.iter_rows(min_row=14, values_only=False), 14):
        vals = {c.column: c.value for c in row if c.value is not None}
        desc = vals.get(3)
        if not desc or not isinstance(desc, str) or not desc.strip():
            continue
        kw = vals.get(5)
        if not isinstance(kw, (int, float)):
            continue

        nome = normalizar_texto(desc)
        tensao = vals.get(6) if isinstance(vals.get(6), (int, float)) else None

        registros.append({
            "codigo_fonte": nome,
            "nome": f"ALIMENTADOR {nome}",
            "secao": "ALIMENTADORES",
            "unidade": "CJ",
            "tipo_item": "COMPOSICAO_ALIMENTADOR",
            "quantidade": 1,
            "hh_unitario": None,
            "hh_fator_dificuldade": None,
            "hh_total": None,
            "preco_unitario": None,
            "preco_total": None,
            "preco_venda_unit": None,
            "fornecimento": "",
            "curva_abc": "",
            "_fonte": NOME_FONTE,
            "_aba_origem": "Alimentação-industrial",
            "_linha_origem": i,
            "_confianca": "baixa",
            "_revisar": "SIM",
            "_observacao": f"KW={kw} V={tensao}. Composição calculada por fórmulas — itens decompostos em cabos/eletrodutos/terminais na planilha original",
        })

    return registros


def salvar(registros):
    colunas = [
        "codigo_fonte", "nome", "secao", "unidade", "tipo_item",
        "quantidade", "hh_unitario", "hh_fator_dificuldade", "hh_total",
        "preco_unitario", "preco_total", "preco_venda_unit",
        "fornecimento", "curva_abc",
        "_fonte", "_aba_origem", "_linha_origem", "_confianca",
        "_revisar", "_observacao",
    ]

    try:
        wb = openpyxl.load_workbook(str(SAIDA))
    except (FileNotFoundError, PermissionError):
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    if "MateriaisIndustrial" in wb.sheetnames:
        del wb["MateriaisIndustrial"]

    ws = wb.create_sheet("MateriaisIndustrial")

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    audit_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, cn in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=ci, value=cn)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    for ri, reg in enumerate(registros, 2):
        for ci, cn in enumerate(colunas, 1):
            val = reg.get(cn, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            if cn.startswith("_"):
                cell.fill = audit_fill

    larguras = {
        "codigo_fonte": 16, "nome": 55, "secao": 30, "unidade": 8,
        "tipo_item": 22, "quantidade": 12, "hh_unitario": 12,
        "hh_fator_dificuldade": 18, "hh_total": 10,
        "preco_unitario": 14, "preco_total": 14, "preco_venda_unit": 16,
        "fornecimento": 12, "curva_abc": 25,
        "_fonte": 45, "_aba_origem": 25, "_linha_origem": 12,
        "_confianca": 10, "_revisar": 8, "_observacao": 60,
    }
    for ci, cn in enumerate(colunas, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = larguras.get(cn, 12)

    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(registros) + 1}"
    ws.freeze_panes = "A2"

    try:
        wb.save(str(SAIDA))
        print(f"Salvo: {SAIDA} (aba MateriaisIndustrial)")
    except PermissionError:
        fallback = SAIDA.parent / "EXTRACAO_Pequenas_Obras_LP_MateriaisIndustrial.xlsx"
        wb.save(str(fallback))
        print(f"Salvo (fallback): {fallback}")


def extrair():
    print("Passo 6: Materiais - industrial + Alimentação...")
    wb = openpyxl.load_workbook(str(FONTE), read_only=True, data_only=True)

    r1 = extrair_materiais_industrial(wb)
    print(f"  Materiais - industrial: {len(r1)} itens")

    r2 = extrair_alimentacao(wb)
    print(f"  Alimentação-industrial: {len(r2)} alimentadores")

    wb.close()

    registros = r1 + r2
    print(f"  Total: {len(registros)}")

    salvar(registros)

    # Relatório
    print(f"\n{'='*60}")
    print("RELATÓRIO — Materiais Industrial + Alimentação")
    print(f"{'='*60}")
    por_tipo = {}
    for r in registros:
        t = r["tipo_item"]
        por_tipo[t] = por_tipo.get(t, 0) + 1
    for t, c in sorted(por_tipo.items(), key=lambda x: -x[1]):
        print(f"  {t}: {c}")
    revisar = sum(1 for r in registros if r["_revisar"] == "SIM")
    print(f"  Para revisão: {revisar}")
    com_hh = sum(1 for r in registros if r.get("hh_unitario"))
    print(f"  Com HH unitário: {com_hh}")
    com_preco = sum(1 for r in registros if r.get("preco_unitario"))
    print(f"  Com preço unitário: {com_preco}")
    print(f"{'='*60}")


if __name__ == "__main__":
    extrair()
