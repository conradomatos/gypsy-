"""
Gera relatório Excel mapeando todas as 37 abas da planilha Pequenas Obras.
Uma linha por aba com: propósito, inputs, outputs, dependências, utilidade CostAI.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
SAIDA = BASE_DIR / "intermediarios" / "MAPA_ABAS_Pequenas_Obras_LP.xlsx"

# Dados coletados do diagnóstico completo
ABAS = [
    {
        "num": 1,
        "nome": "Formula Reajuste Cabos",
        "linhas": 19,
        "formulas": 7,
        "proposito": "Fórmula de reajuste de preço de cabos com base na cotação do cobre (LME) e câmbio USD/BRL. Permite recalcular preço de cabos no momento do faturamento.",
        "inputs_usuario": "LMEi (cotação cobre inicial), Câmbio US$i, LMEf (cotação atual), Câmbio US$f, Kcobre (conteúdo metal kg/km)",
        "calcula_gera": "Preço realinhado do cabo (P = Pi + [(LMEf × US$f) - (LMEi × US$i)] × Kcobre / 1000)",
        "puxa_de": "(nenhuma — dados digitados)",
        "alimenta": "Base de Materiais (fator de reajuste para preços de cabos)",
        "utilidade_costai": "ALTA — Fórmula de reajuste de cabos é parametrizável no CostAI para atualização automática de preços",
        "extraido": "NAO",
        "tipo": "PARAMETRO",
    },
    {
        "num": 2,
        "nome": "Plan1",
        "linhas": 0,
        "formulas": 0,
        "proposito": "Aba vazia, sem conteúdo. Provavelmente residual.",
        "inputs_usuario": "(nenhum)",
        "calcula_gera": "(nada)",
        "puxa_de": "(nenhuma)",
        "alimenta": "(nenhuma)",
        "utilidade_costai": "NENHUMA — Aba vazia",
        "extraido": "IGNORADA",
        "tipo": "RESIDUAL",
    },
    {
        "num": 3,
        "nome": "LEVANTAMENTO",
        "linhas": 76,
        "formulas": 45,
        "proposito": "Checklist de parâmetros do orçamento: padrão do cliente (VALE/Klabin), flags (fibra óptica, proteção passiva painéis, lixo classe 2, plano saúde). Configura opções que afetam cálculos em outras abas.",
        "inputs_usuario": "Flags: Padrão VALE, Padrão Klabin, dias integração, fibra óptica (m), fusão fibra (pt), proteção passiva (m²)",
        "calcula_gera": "Parâmetros de configuração que alimentam cálculos de materiais e seguros",
        "puxa_de": "M.O.D. Elétrica, M.O.I, RESUMO PREÇOS Industrial",
        "alimenta": "(parâmetros usados indiretamente)",
        "utilidade_costai": "BAIXA — São flags de orçamento específicas por obra, não dados reutilizáveis",
        "extraido": "NAO",
        "tipo": "CONFIGURACAO",
    },
    {
        "num": 4,
        "nome": "Administração Central IND",
        "linhas": 62,
        "formulas": 121,
        "proposito": "Calcula o custo de administração central (overhead da CONCEPT) como % do valor total da obra. Distribui por item: materiais, MOD, MOI, seguros, mobilização, etc. Inclui BDI da administração central.",
        "inputs_usuario": "Alíquotas de BDI (ISS, PIS, COFINS, Adm CONCEPT 10%, Custo Financeiro, Desoneração, Lucro 15%, IR 4.8%, IR Adicional 3.2%, CSLL)",
        "calcula_gera": "Custo total da administração central com impostos e BDI. Curva ABC dos custos.",
        "puxa_de": "M.O.D. Elétrica, M.O.I, Materiais-industrial, Manutenção de Canteiro, Seguros e Outros, Despesas Viagens, Prog. Segurança, RESUMO PREÇOS",
        "alimenta": "RESUMO PREÇOS Industrial",
        "utilidade_costai": "MEDIA — Alíquotas de BDI e estrutura de overhead são referência para parametrização do CostAI (composição do BDI)",
        "extraido": "NAO",
        "tipo": "CALCULO_BDI",
    },
    {
        "num": 5,
        "nome": "Alimentação-industrial",
        "linhas": 631,
        "formulas": 9326,
        "proposito": "Calculadora de alimentadores elétricos. A partir de KW, V e distância, dimensiona automaticamente cabos (bitola), eletrodutos, leitos e terminais. Gera lista de materiais e quantidades por alimentador.",
        "inputs_usuario": "Por alimentador: descrição, tipo (motor/iluminação), KW, V, distância (m), fatores de agrupamento/temperatura/potência",
        "calcula_gera": "Corrente nominal, bitola de cabos, quantidades de cabos (tetrapolar, tripolar, unipolar, MT), eletrodutos, leitos, terminais. 9.326 fórmulas.",
        "puxa_de": "(nenhuma — cálculos internos com tabelas embutidas)",
        "alimenta": "Materiais - industrial (puxa quantidades calculadas dos alimentadores)",
        "utilidade_costai": "ALTA — Lógica de dimensionamento elétrico é o core do CostAI. As regras de cálculo (corrente → bitola → material) são parametrizáveis.",
        "extraido": "PARCIAL (14 alimentadores como composição, sem decomposição dos itens)",
        "tipo": "CALCULADORA",
    },
    {
        "num": 6,
        "nome": "Pintura",
        "linhas": 34,
        "formulas": 43,
        "proposito": "Calcula área de pintura (jateamento + fundo + acabamento) de perfis, eletrodutos, eletrocalhas e leitos a partir de peso/metro linear. Gera custo total de pintura.",
        "inputs_usuario": "Quantidades de perfis (kg/m), eletrodutos e eletrocalhas por bitola (m)",
        "calcula_gera": "Área total de pintura (m²), HH de pintura, custo de materiais de pintura",
        "puxa_de": "Materiais - industrial (quantidades de eletrodutos/eletrocalhas)",
        "alimenta": "Custo Obra_Industrial (custo de pintura entra como material de consumo)",
        "utilidade_costai": "MEDIA — Coeficientes de conversão kg/m → m²/m (área de pintura) são úteis como composição",
        "extraido": "NAO",
        "tipo": "CALCULO",
    },
    {
        "num": 7,
        "nome": "Base de Materiais",
        "linhas": 2879,
        "formulas": 7569,
        "proposito": "Base de preços de todos os materiais de montagem elétrica. 22 famílias (eletroduto, cabo, leito, eletrocalha, etc.). Preço unitário calculado por fator de reajuste × preço fabricante. Cotações históricas de fornecedores.",
        "inputs_usuario": "Preço fabricante (col G), fatores de reajuste por família (col O), cotações de fornecedores (cols Q-AJ)",
        "calcula_gera": "Preço unitário reajustado (col E = G × fator), valor médio de cotações (col P = AVERAGE), HH unitário por item",
        "puxa_de": "Formula Reajuste Cabos (fator de reajuste de cabos), Base de Materiais (auto-referência para médias)",
        "alimenta": "Materiais - industrial (preços unitários dos materiais)",
        "utilidade_costai": "ALTA — Tabela mãe de insumos com 2.771 itens, preços e HH por unidade. Fundação do CostAI.",
        "extraido": "SIM — Aba Insumos (2.771 registros, schema familia/subfamilia/classe)",
        "tipo": "BASE_DADOS",
    },
    {
        "num": 8,
        "nome": "Impostos-industrial",
        "linhas": 23,
        "formulas": 126,
        "proposito": "Planilha de impostos por cenário. Calcula COFINS, PIS, ICMS, IPI, ISS sobre cada item do orçamento (materiais, serviços, mobilização, etc.). Gera preço com e sem impostos.",
        "inputs_usuario": "Alíquotas de impostos (COFINS 3%, PIS 0.65%, ICMS 18%, ISS 3%), preços sem impostos por item",
        "calcula_gera": "Valor de cada imposto por item, preço total com impostos",
        "puxa_de": "Materiais - industrial, RESUMO PREÇOS Industrial",
        "alimenta": "(resultado final — não alimenta outras abas)",
        "utilidade_costai": "ALTA — Alíquotas de impostos por regime tributário (Lucro Presumido). Parametrizável no CostAI.",
        "extraido": "SIM — Aba EncargosBDI (5 impostos)",
        "tipo": "IMPOSTO",
    },
    {
        "num": 9,
        "nome": "Horas improdutivas",
        "linhas": 13,
        "formulas": 6,
        "proposito": "Calcula fator de improdutividade da MO. Tempo perdido com deslocamento interno, DDS (diálogo de segurança), liberação de frentes de trabalho.",
        "inputs_usuario": "Duração de cada atividade improdutiva (deslocamento, DDS, liberação frentes), distância casa-obra, velocidade média",
        "calcula_gera": "% de horas improdutivas sobre jornada (10.2% sobre 9h = 55min/dia). Fator de produtividade.",
        "puxa_de": "(nenhuma)",
        "alimenta": "histograma (fator de horas improdutivas aplicado ao histograma de MO)",
        "utilidade_costai": "ALTA — Fator de produtividade (0.898) é crucial para composição de custos. Parametrizável por tipo de obra.",
        "extraido": "NAO",
        "tipo": "PARAMETRO",
    },
    {
        "num": 10,
        "nome": "Materiais - industrial",
        "linhas": 966,
        "formulas": 10458,
        "proposito": "Planilha de orçamento de materiais para uma obra específica. Template com seções por tipo de serviço (painéis, transformadores, eletrodutos, cabos, etc.). Cada item tem: código, descrição, qtd, HH, preço. Fórmulas calculam totais e preços de venda com markup.",
        "inputs_usuario": "Quantidades de cada item (col 5), fornecimento (CONCEPT/DIRETO), fator de dificuldade para HH",
        "calcula_gera": "HH total por item, preço total materiais, preço total MO, preço de venda (com markup), critério curva ABC, tipo faturamento (direto/CONCEPT)",
        "puxa_de": "Base de Materiais (preços), Alimentação-industrial (qtds alimentadores), Instrumentação-industrial (qtds instrumentação), M.O.D. Elétrica (custo HH), TAB HH (preço HH de venda), RESUMO PREÇOS",
        "alimenta": "RESUMO PREÇOS Industrial, Impostos-industrial, Pintura, Administração Central",
        "utilidade_costai": "ALTA — Template de orçamento com composições implícitas (material + HH). Modelo de como a CONCEPT orça.",
        "extraido": "SIM — Aba MateriaisIndustrial (866 itens com HH e preços)",
        "tipo": "ORCAMENTO",
    },
    {
        "num": 11,
        "nome": "LM (2)",
        "linhas": 69,
        "formulas": 89,
        "proposito": "Lista de materiais de uma obra específica (versão 2). Itens com descrição, unidade, quantidade, preço unitário e preço total. Inclui materiais EX (à prova de explosão).",
        "inputs_usuario": "Quantidades e preços por item",
        "calcula_gera": "Preço total por item (qtd × preço unit)",
        "puxa_de": "(nenhuma — dados digitados)",
        "alimenta": "(nenhuma — lista avulsa)",
        "utilidade_costai": "BAIXA — Lista de materiais de obra específica, não é base reutilizável. Preços EX são referência.",
        "extraido": "NAO",
        "tipo": "LISTA_MATERIAIS",
    },
    {
        "num": 12,
        "nome": "LM",
        "linhas": 88,
        "formulas": 112,
        "proposito": "Lista de materiais de uma obra específica (versão original). Similar à LM (2) mas com mais itens. Materiais elétricos com preços e quantidades.",
        "inputs_usuario": "Quantidades e preços por item",
        "calcula_gera": "Preço total por item",
        "puxa_de": "(nenhuma)",
        "alimenta": "(nenhuma)",
        "utilidade_costai": "BAIXA — Lista de materiais de obra específica",
        "extraido": "NAO",
        "tipo": "LISTA_MATERIAIS",
    },
    {
        "num": 13,
        "nome": "Instrumentação- industrial",
        "linhas": 172,
        "formulas": 130,
        "proposito": "Calculadora de quantitativos de instrumentação. A partir do número de transmissores, válvulas e painéis, calcula necessidade de cabos de sinal, tubing inox/polietileno, eletrodutos, conectores e terminais.",
        "inputs_usuario": "Quantidade de transmissores (TT, PDT, FT, LT), válvulas (LV, PV, TV, FV), painéis. Distância média para cabos/eletrodutos/tubing.",
        "calcula_gera": "Quantidades de cabos de sinal, tubing inox, tubing polietileno, eletrodutos, conectores por tipo de instrumento",
        "puxa_de": "(nenhuma — cálculos internos)",
        "alimenta": "Materiais - industrial (quantidades de instrumentação)",
        "utilidade_costai": "ALTA — Regras de quantificação de instrumentação (instrumento → materiais) são parametrizáveis no CostAI",
        "extraido": "NAO",
        "tipo": "CALCULADORA",
    },
    {
        "num": 14,
        "nome": "RESUMO PREÇOS Industrial",
        "linhas": 71,
        "formulas": 149,
        "proposito": "Resumo consolidado do orçamento. Agrupa todos os custos por categoria (gerenciamento, engenharia, montagem, mobilização, etc.) com valores de serviço e material. Calcula preço de venda final.",
        "inputs_usuario": "Prazo de execução por item (meses/semanas)",
        "calcula_gera": "Valor total por categoria, HH total, preço de venda total da obra. É a 'folha de rosto' do orçamento.",
        "puxa_de": "M.O.D. Elétrica, M.O.I, Materiais-industrial, Manutenção Canteiro, Seguros e Outros, Despesas Viagens, Prog. Segurança, Custo Obra_Industrial",
        "alimenta": "Quase todas as abas puxam dados do RESUMO (é o hub central): histograma, Impostos, Adm Central, Cronograma, Equipamentos, Mobilização, Desmobilização, Seguros, PROJETOS, etc.",
        "utilidade_costai": "MEDIA — Estrutura de resumo de orçamento é referência para o formato de saída do CostAI. Não tem dados base novos.",
        "extraido": "NAO",
        "tipo": "RESUMO",
    },
    {
        "num": 15,
        "nome": "BASE DADOS PAINEIS",
        "linhas": 30557,
        "formulas": 0,
        "proposito": "Lista de preços Siemens completa (Jan/2019). 30.552 itens com código, MLFB, descrição, grupo de material, NCM, disponibilidade e preço líquido (sem PIS/COFINS/ICMS).",
        "inputs_usuario": "(nenhum — dados importados da Siemens)",
        "calcula_gera": "(nenhum cálculo — é base de dados pura)",
        "puxa_de": "(nenhuma — dados estáticos importados)",
        "alimenta": "Paineis (2) (busca preços por código MLFB)",
        "utilidade_costai": "MEDIA — Catálogo Siemens com preços, NCM e códigos. Útil para preços de referência de painéis, mas dados de Jan/2019 (defasados).",
        "extraido": "SIM — Arquivo separado EXTRACAO_Paineis.xlsx (30.552 itens)",
        "tipo": "BASE_DADOS",
    },
    {
        "num": 16,
        "nome": "Paineis (2)",
        "linhas": 299,
        "formulas": 5671,
        "proposito": "Configurador de painéis elétricos. Monta BOM (bill of materials) de painéis por tipo de partida (direta, estrela-triângulo, soft-starter, inversor) e potência. Busca preços na BASE DADOS PAINEIS e aplica desconto/impostos.",
        "inputs_usuario": "Tipo de partida, potência (kW), desconto de negociação",
        "calcula_gera": "Lista de componentes por painel (disjuntor, contator, relé, bornes), preço com impostos e sem IPI, desconto",
        "puxa_de": "BASE DADOS PAINEIS (preços por código MLFB)",
        "alimenta": "(resultado usado no orçamento de painéis)",
        "utilidade_costai": "ALTA — Composições de painéis por tipo de partida e potência. Modelo de configurador de painéis para o CostAI.",
        "extraido": "NAO",
        "tipo": "CONFIGURADOR",
    },
    {
        "num": 17,
        "nome": "histograma",
        "linhas": 96,
        "formulas": 837,
        "proposito": "Histograma de mão de obra: distribui as funções (MOI + MOD) ao longo das semanas do cronograma. Calcula HH total por semana, incluindo horas improdutivas. Gera curva de mobilização.",
        "inputs_usuario": "Quantidade de pessoas por função por semana (grid semanal), dias da semana trabalhados",
        "calcula_gera": "HH por semana, total de funcionários por semana, HH com improdutividade, curva de mobilização",
        "puxa_de": "Tabela de Entrada de Preços (salários), M.O.D. Elétrica (funções), Horas improdutivas (fator), RESUMO PREÇOS",
        "alimenta": "M.O.I (dias trabalhados/pessoas), M.O.D. Elétrica (dias/pessoas)",
        "utilidade_costai": "BAIXA — Histograma é específico por obra. A ESTRUTURA (função × semana) pode ser modelo.",
        "extraido": "NAO",
        "tipo": "PLANEJAMENTO",
    },
    {
        "num": 18,
        "nome": "Custos MOI - Equip.-mensal-Ind.",
        "linhas": 12,
        "formulas": 19,
        "proposito": "Resumo do custo mensal de MOI e Equipamentos. Agrega: Administração Local, Administração Central, Ferramentas/Equipamentos/Veículos, Manutenção de Canteiro. Custo por semana.",
        "inputs_usuario": "Quantidade de semanas",
        "calcula_gera": "Custo semanal e total de overhead (MOI + equipamentos + canteiro)",
        "puxa_de": "Manutenção de Canteiro, RESUMO PREÇOS Industrial",
        "alimenta": "(resumo auxiliar — não alimenta diretamente)",
        "utilidade_costai": "BAIXA — Resumo auxiliar de custos indiretos",
        "extraido": "NAO",
        "tipo": "RESUMO",
    },
    {
        "num": 19,
        "nome": "Cronograma de desembolso_Indust",
        "linhas": 33,
        "formulas": 156,
        "proposito": "Cronograma financeiro: distribui o valor total do orçamento em meses (curva de desembolso). Calcula avanço financeiro acumulado e média de funcionários por semana.",
        "inputs_usuario": "Pesos de distribuição por mês (% de cada item no período)",
        "calcula_gera": "Desembolso por mês por categoria, avanço financeiro acumulado, média de funcionários diretos",
        "puxa_de": "RESUMO PREÇOS Industrial",
        "alimenta": "(saída final)",
        "utilidade_costai": "BAIXA — Cronograma específico por obra",
        "extraido": "NAO",
        "tipo": "PLANEJAMENTO",
    },
    {
        "num": 20,
        "nome": "Cronograma macro",
        "linhas": 38,
        "formulas": 77,
        "proposito": "Cronograma macro da obra em formato Gantt simplificado. Lista de atividades com duração em dias e datas de início/fim. Inclui fases: engenharia, compras, mobilização, pré-parada, parada, comissionamento.",
        "inputs_usuario": "Data início, duração de cada atividade (dias), equipe e horas/dia",
        "calcula_gera": "Datas de início/fim de cada atividade, HH por atividade (equipe × horas × dias)",
        "puxa_de": "(nenhuma — dados digitados)",
        "alimenta": "(nenhuma)",
        "utilidade_costai": "BAIXA — Cronograma específico. A lista de fases/atividades padrão pode servir de template.",
        "extraido": "NAO",
        "tipo": "PLANEJAMENTO",
    },
    {
        "num": 21,
        "nome": "PROJETOS",
        "linhas": 29,
        "formulas": 54,
        "proposito": "Custo de projetos de engenharia (projeto de cabos, dimensionamento, identificação). Calcula custo por HH de engenheiro × quantidade de horas. Inclui BDI separado (ISS, PIS, COFINS, Adm 10%, Lucro 15%, IR).",
        "inputs_usuario": "Quantidade de HH de engenharia, preço unitário (custo)",
        "calcula_gera": "Custo total de projetos com encargos sobre faturamento",
        "puxa_de": "RESUMO PREÇOS Industrial (base de cálculo para impostos)",
        "alimenta": "RESUMO PREÇOS Industrial",
        "utilidade_costai": "BAIXA — Custo de projeto é específico. As alíquotas de BDI sobre projetos são referência.",
        "extraido": "NAO",
        "tipo": "CALCULO",
    },
    {
        "num": 22,
        "nome": "Manutenção de Canteiro",
        "linhas": 52,
        "formulas": 175,
        "proposito": "Custo de manutenção do canteiro de obras: containers, cobertura, acessórios, internet, vigilância, ambulatório, ETE. Calcula custo diário de cada item × dias de uso. Inclui insumos de canteiro (água, energia, limpeza).",
        "inputs_usuario": "Quantidade de cada item de canteiro, dias de uso, insumos mensais (m³ água, kW energia)",
        "calcula_gera": "Custo total do canteiro, custo com encargos/impostos, % sobre valor total da obra",
        "puxa_de": "Tabela de Entrada de Preços (preços de containers/equipamentos), RESUMO PREÇOS Industrial",
        "alimenta": "RESUMO PREÇOS Industrial, Administração Central, Custo Obra_Industrial, Custos MOI Mensal",
        "utilidade_costai": "MEDIA — Custos de canteiro são parametrizáveis (tipo de canteiro × duração da obra)",
        "extraido": "NAO",
        "tipo": "CALCULO",
    },
    {
        "num": 23,
        "nome": "Equipamentos",
        "linhas": 70,
        "formulas": 246,
        "proposito": "Custo de equipamentos e ferramentas para a obra. Automotores (guindastes, muncks, plataformas), ferramentas (furadeiras, esmerilhadeiras), geradores. Custo por DIA × dias de uso × quantidade. Inclui encargos sobre faturamento.",
        "inputs_usuario": "Quantidade de cada equipamento, dias de uso",
        "calcula_gera": "Custo total de equipamentos, valor com encargos (ISS, PIS, COFINS), % sobre custo direto",
        "puxa_de": "Tabela de Entrada de Preços (preços base), M.O.D./M.O.I (para % sobre custos), RESUMO PREÇOS",
        "alimenta": "RESUMO PREÇOS Industrial",
        "utilidade_costai": "ALTA — Tabela de equipamentos com custo/dia. Já extraído.",
        "extraido": "SIM — Aba Equipamentos (40 itens + 25 do Cálculo + 13 EPIs = 78 total)",
        "tipo": "BASE_DADOS",
    },
    {
        "num": 24,
        "nome": "Tabela de Entrada de Preços",
        "linhas": 116,
        "formulas": 145,
        "proposito": "Tabela mãe de salários e preços unitários. 4 seções: MÃO-DE-OBRA (41 funções com salário mensal, CLT/PJ, custo/hora), EQUIPAMENTOS (automotores com preço mensal), CANTEIRO (containers, coberturas), FERRAMENTAS (andaime, rosqueadeira, etc.).",
        "inputs_usuario": "Salário base de cada função, tipo CLT/PJ, flag periculosidade. Dissídio 2020.",
        "calcula_gera": "Custo/hora (CLT: salário/220; PJ: salário/220/(1+encargos)). Salário com periculosidade (×1.3). Fórmulas referenciam aba Encargos.",
        "puxa_de": "Cálculo Custo de Equipamentos (preços de containers/equipamentos depreciados)",
        "alimenta": "M.O.I, M.O.D. Elétrica, Equipamentos, Manutenção Canteiro, histograma, Mobilização, Desmobilização",
        "utilidade_costai": "ALTA — Tabela base de salários. Já extraída como Mão de Obra (41 funções).",
        "extraido": "SIM — Aba MaoDeObra (41 funções, seção de MO)",
        "tipo": "BASE_DADOS",
    },
    {
        "num": 25,
        "nome": "Cálculo Custo de Equipamentos",
        "linhas": 55,
        "formulas": 220,
        "proposito": "Cálculo detalhado de custo de propriedade de equipamentos: caminhões + guindastes (6 configurações), containers (7 tipos), geradores (6 potências), plataformas, betoneira, oficina completa. Fórmula: depreciação 60 meses + remuneração investimento 1.5%/mês + manutenção 15%/ano + combustível.",
        "inputs_usuario": "Preço de aquisição do caminhão, preço do guindaste/acessório, consumo diesel/gasolina por hora, preço combustível",
        "calcula_gera": "Depreciação mensal, custo/hora (com e sem operador), custo/mês. Inclui custo de combustível por hora.",
        "puxa_de": "(nenhuma — dados de aquisição digitados)",
        "alimenta": "Tabela de Entrada de Preços (preços mensais dos equipamentos)",
        "utilidade_costai": "ALTA — Modelo de TCO (Total Cost of Ownership) para equipamentos. Já extraído (25 itens com depreciação).",
        "extraido": "SIM — Aba Equipamentos (25 itens com depreciação 60 meses)",
        "tipo": "CALCULO",
    },
    {
        "num": 26,
        "nome": "M.O.I",
        "linhas": 111,
        "formulas": 343,
        "proposito": "Custo da Mão de Obra Indireta para a obra. 10 funções de gestão com HH normal + HE 50% + HE 100%. Calcula Hxh Venda (custo hora com encargos + alimentação + overhead / (1-margem)). Seção de alimentação e transporte.",
        "inputs_usuario": "Dias trabalhados e nº de pessoas por função (vêm do histograma), valor de refeições/transporte",
        "calcula_gera": "Custo total MOI, HxH de venda por função, custo com encargos sociais (78.7%), custo de alimentação/transporte",
        "puxa_de": "Tabela de Entrada de Preços (salários), EPI e Ferramentas (custo EPI/mês), M.O.D. Elétrica, RESUMO PREÇOS",
        "alimenta": "COMPOSIÇÃO Hh, TAB HH, RESUMO PREÇOS, Administração Central, Custo Obra, histograma, Equipamentos, Prog. Segurança",
        "utilidade_costai": "ALTA — Composição de custo hora de venda da MOI. Já extraída.",
        "extraido": "SIM — Aba MaoDeObra (cruzamento para classificação MOI + Hxh Venda)",
        "tipo": "CALCULO",
    },
    {
        "num": 27,
        "nome": "COMPOSIÇÃO Hh",
        "linhas": 36,
        "formulas": 421,
        "proposito": "Composição salarial completa por função (20 funções). Decompõe o custo HH em: salário, periculosidade, encargos sociais, alimentação (café+almoço+jantar+VA), transporte, EPI, despesas admissionais, seguro vida, ferramentas individuais. Calcula custo/mês e custo/hora com e sem BDI.",
        "inputs_usuario": "Fator BDI (0.74), divisor de horas (180h/mês)",
        "calcula_gera": "Custo funcionário/mês sem BDI, custo HH sem BDI (÷180), custo HH com BDI. 20 funções detalhadas. Adicionais de HE (50% e 100%) e noturno (20%).",
        "puxa_de": "M.O.I (salários e encargos MOI), M.O.D. Elétrica (salários e encargos MOD), EPI e Ferramentas (custo EPI/mês), Prog. Segurança (exames)",
        "alimenta": "TAB HH (preço HH de venda final)",
        "utilidade_costai": "ALTA — Composição detalhada do custo HH. Já extraída (20 composições + 232 itens).",
        "extraido": "SIM — Abas Composicoes (20) + ItensComposicao (232)",
        "tipo": "COMPOSICAO",
    },
    {
        "num": 28,
        "nome": "M.O.D. Elétrica",
        "linhas": 120,
        "formulas": 364,
        "proposito": "Custo da Mão de Obra Direta elétrica. 10 funções operacionais com HH normal + HE 50% + HE 100%. Periculosidade 30%, encargos 78.7%, transferência. Seção de alimentação e transporte.",
        "inputs_usuario": "Dias trabalhados e nº de pessoas por função (histograma), refeições/transporte",
        "calcula_gera": "Custo total MOD, custo com periculosidade e encargos, custo de alimentação/transporte, % sobre valor total",
        "puxa_de": "Tabela de Entrada de Preços (salários), EPI e Ferramentas, RESUMO PREÇOS",
        "alimenta": "COMPOSIÇÃO Hh, TAB HH, RESUMO PREÇOS, Materiais-industrial, Administração Central, Custo Obra, histograma, Impostos, Equipamentos, Mobilização, Desmobilização, Prog. Segurança",
        "utilidade_costai": "ALTA — Custo MOD com encargos. Já extraída.",
        "extraido": "SIM — Aba MaoDeObra (cruzamento para classificação MOD)",
        "tipo": "CALCULO",
    },
    {
        "num": 29,
        "nome": "TAB HH",
        "linhas": 33,
        "formulas": 46,
        "proposito": "Tabela resumo de preço do HH por função — versão 'para apresentação'. Lista 20 funções (10 MOI + 10 MOD) com preço unitário R$/HH de venda (com BDI). É a tabela que vai na proposta comercial.",
        "inputs_usuario": "(nenhum — valores calculados)",
        "calcula_gera": "Preço HH de venda por função (referência final para precificação de serviços)",
        "puxa_de": "COMPOSIÇÃO Hh (custo HH com BDI), M.O.I, M.O.D. Elétrica",
        "alimenta": "Materiais - industrial (preço HH para calcular custo de MO nos serviços)",
        "utilidade_costai": "ALTA — Tabela de preços HH de venda. Dados já estão na COMPOSIÇÃO Hh extraída.",
        "extraido": "INDIRETO — Dados presentes na aba Composicoes (custo_unitario_com_bdi)",
        "tipo": "RESUMO",
    },
    {
        "num": 30,
        "nome": "EPI e Ferramentas",
        "linhas": 16,
        "formulas": 15,
        "proposito": "Custo mensal de EPIs e ferramentas individuais por depreciação. EPIs: capacete, uniforme (calça+camiseta+jaleco+antichama), botina, óculos, luvas, protetor auricular. Ferramentas: caixa eletricista FC, caixa eletricista montador.",
        "inputs_usuario": "Valor de aquisição, tempo de depreciação (meses), quantidade",
        "calcula_gera": "Custo mensal por item (aquisição / depreciação × quantidade), custo total EPI/mês, custo hora EPI (/220)",
        "puxa_de": "(nenhuma — dados digitados)",
        "alimenta": "COMPOSIÇÃO Hh (custo EPI/uniforme por mês), M.O.I (custo EPI), M.O.D. Elétrica (custo EPI)",
        "utilidade_costai": "ALTA — Custos de EPI depreciados. Já extraído (13 itens).",
        "extraido": "SIM — Aba Equipamentos (13 itens EPI)",
        "tipo": "BASE_DADOS",
    },
    {
        "num": 31,
        "nome": "Seguros e Outros",
        "linhas": 43,
        "formulas": 99,
        "proposito": "Custos de seguros, garantias e outros custos diretos: proteção passiva, fusão fibra óptica, data book, performance bond, seguro RC, ART, ensaios, enchimento trafo, licenças, galvanização a frio.",
        "inputs_usuario": "Quantidade de cada item, valor unitário",
        "calcula_gera": "Custo total de seguros e outros, com encargos (ISS, PIS, COFINS), % sobre custo direto",
        "puxa_de": "RESUMO PREÇOS Industrial (base de cálculo para %)",
        "alimenta": "RESUMO PREÇOS Industrial, Administração Central, Custo Obra_Industrial",
        "utilidade_costai": "BAIXA — Custos de seguros são específicos por obra. Alíquotas de seguros (1.65% RC, 10% performance bond) podem ser referência.",
        "extraido": "NAO",
        "tipo": "CALCULO",
    },
    {
        "num": 32,
        "nome": "Custo Obra_Industrial",
        "linhas": 45,
        "formulas": 501,
        "proposito": "Planilha de controle de custos da obra. Compara ORÇADO × REALIZADO por categoria: estadia/viagens, alimentação, canteiro, locação equipamentos, ferramentas, transporte, salários MOD/MOI, assistência médica, EPI, materiais de consumo, materiais de fornecimento. Com colunas para rateio por sub-item.",
        "inputs_usuario": "Valores realizados por categoria (para controle durante a obra)",
        "calcula_gera": "Comparativo orçado × realizado, desvios por categoria, custos agrupados",
        "puxa_de": "M.O.D., M.O.I, Materiais-industrial, Manutenção Canteiro, Seguros, Despesas Viagens, Prog. Segurança, RESUMO PREÇOS",
        "alimenta": "RESUMO PREÇOS Industrial (custo total consolidado)",
        "utilidade_costai": "MEDIA — Estrutura de controle de custos (categorias de custo) é referência para relatórios do CostAI",
        "extraido": "NAO",
        "tipo": "CONTROLE",
    },
    {
        "num": 33,
        "nome": "Mobilização",
        "linhas": 115,
        "formulas": 208,
        "proposito": "Custo de mobilização da equipe: MO direta (5 funções × dias), alimentação, transporte, exames admissionais, treinamentos, fardamento. Estrutura idêntica à M.O.D. (HN + HE50% + HE100%).",
        "inputs_usuario": "Dias de mobilização, nº de pessoas, distância de deslocamento",
        "calcula_gera": "Custo total de mobilização com encargos e impostos",
        "puxa_de": "Tabela de Entrada de Preços (salários), RESUMO PREÇOS",
        "alimenta": "RESUMO PREÇOS Industrial",
        "utilidade_costai": "MEDIA — Modelo de composição de custo de mobilização. Coeficientes (dias/pessoa por tipo de obra) são parametrizáveis.",
        "extraido": "NAO",
        "tipo": "CALCULO",
    },
    {
        "num": 34,
        "nome": "Desmobilização",
        "linhas": 114,
        "formulas": 207,
        "proposito": "Custo de desmobilização (espelho da Mobilização). Mesma estrutura: MO, alimentação, transporte, exames demissionais.",
        "inputs_usuario": "Dias de desmobilização, nº de pessoas",
        "calcula_gera": "Custo total de desmobilização com encargos e impostos",
        "puxa_de": "Tabela de Entrada de Preços, RESUMO PREÇOS",
        "alimenta": "RESUMO PREÇOS Industrial",
        "utilidade_costai": "BAIXA — Espelho da mobilização",
        "extraido": "NAO",
        "tipo": "CALCULO",
    },
    {
        "num": 35,
        "nome": "Encargos",
        "linhas": 58,
        "formulas": 60,
        "proposito": "Tabela de encargos sociais detalhada: Grupo A (INSS 20%, SESI, SENAI, INCRA, SEBRAE, FGTS, SAT, Sal.Educação = 36.8%), Grupo B (DSR, feriados, férias+1/3, auxílio doença, licenças, 13º = 43.2% horista / 22.7% mensalista), Grupo C (multa FGTS, aviso prévio = 10.9%), Grupo D (incidência A sobre B). Total: Horista 106.7%, Mensalista 78.7%.",
        "inputs_usuario": "Alíquotas de cada encargo (parametrizáveis por legislação/dissídio)",
        "calcula_gera": "Total de encargos horista (106.7%) e mensalista (78.7%)",
        "puxa_de": "(nenhuma — alíquotas digitadas)",
        "alimenta": "Tabela de Entrada de Preços (% encargos para cálculo custo/hora PJ), implicitamente usado em M.O.I e M.O.D.",
        "utilidade_costai": "ALTA — Tabela detalhada de encargos sociais. Já extraída (44 encargos + 2 totais).",
        "extraido": "SIM — Aba EncargosBDI (44 encargos detalhados)",
        "tipo": "BASE_DADOS",
    },
    {
        "num": 36,
        "nome": "DESPESAS VIAGENS GERENCIAIS",
        "linhas": 32,
        "formulas": 66,
        "proposito": "Custo de viagens da gerência: carro (km × R$/km), estacionamento, táxi, passagens aéreas, alimentação, hospedagem. Com encargos sobre faturamento (ISS, PIS, COFINS, Adm 10%, Lucro 15%, IR).",
        "inputs_usuario": "Quantidade de viagens, distância, preços unitários",
        "calcula_gera": "Custo total de viagens com BDI completo",
        "puxa_de": "RESUMO PREÇOS Industrial",
        "alimenta": "RESUMO PREÇOS Industrial, Administração Central, Custo Obra_Industrial",
        "utilidade_costai": "BAIXA — Custos de viagem são específicos por obra",
        "extraido": "NAO",
        "tipo": "CALCULO",
    },
    {
        "num": 37,
        "nome": "Prog. segurança e exames",
        "linhas": 138,
        "formulas": 183,
        "proposito": "Programas de segurança do trabalho e exames médicos: plano saúde, PPRA, PCMAT, PCMSO, controle ambiental, CIPA, PPP. Exames: admissional, demissional, periódico, audiometria, espirometria, eletroencefalograma, etc. Aplica por função (quem precisa de qual exame).",
        "inputs_usuario": "Quantidade de exames por tipo, preço unitário, quais funções necessitam",
        "calcula_gera": "Custo total de segurança e exames, com encargos sobre faturamento",
        "puxa_de": "M.O.D. Elétrica, M.O.I, RESUMO PREÇOS Industrial",
        "alimenta": "COMPOSIÇÃO Hh (custo de exames por funcionário), RESUMO PREÇOS, Administração Central, Custo Obra",
        "utilidade_costai": "MEDIA — Lista de exames por função e preços unitários. Matriz função × exame é parametrizável.",
        "extraido": "NAO",
        "tipo": "CALCULO",
    },
]


def gerar():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mapa das Abas"

    colunas = [
        ("num", "#", 4),
        ("nome", "Nome da Aba", 35),
        ("tipo", "Tipo", 16),
        ("linhas", "Linhas", 8),
        ("formulas", "Fórmulas", 10),
        ("proposito", "Propósito / O que faz", 70),
        ("inputs_usuario", "Inputs do Usuário", 50),
        ("calcula_gera", "O que Calcula / Gera", 55),
        ("puxa_de", "Puxa Dados De (fórmulas)", 55),
        ("alimenta", "Alimenta Quais Abas", 55),
        ("utilidade_costai", "Utilidade para CostAI", 55),
        ("extraido", "Status Extração", 50),
    ]

    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Cores por tipo
    tipo_cores = {
        "BASE_DADOS": "C6EFCE",    # verde claro
        "COMPOSICAO": "C6EFCE",
        "CALCULO": "BDD7EE",       # azul claro
        "CALCULO_BDI": "BDD7EE",
        "CALCULADORA": "D9E2F3",   # azul mais claro
        "CONFIGURADOR": "D9E2F3",
        "ORCAMENTO": "FCE4EC",     # rosa claro
        "PARAMETRO": "FFF2CC",     # amarelo claro
        "IMPOSTO": "FFF2CC",
        "CONFIGURACAO": "FFF2CC",
        "RESUMO": "F2F2F2",        # cinza claro
        "PLANEJAMENTO": "F2F2F2",
        "CONTROLE": "F2F2F2",
        "LISTA_MATERIAIS": "F2F2F2",
        "RESIDUAL": "E0E0E0",      # cinza
    }

    # Header
    for ci, (key, label, width) in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = thin_border
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = width

    # Dados
    for ri, aba in enumerate(ABAS, 2):
        tipo = aba.get("tipo", "")
        cor = tipo_cores.get(tipo, "FFFFFF")
        fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")

        for ci, (key, label, width) in enumerate(colunas, 1):
            val = aba.get(key, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill = fill

    # Altura das linhas
    for ri in range(2, len(ABAS) + 2):
        ws.row_dimensions[ri].height = 60

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(colunas))}{len(ABAS) + 1}"
    ws.freeze_panes = "A2"

    # --- Aba 2: Grafo de dependências ---
    ws2 = wb.create_sheet("Dependências")
    ws2.cell(row=1, column=1, value="Aba Origem").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="→ Alimenta").font = Font(bold=True)
    ws2.cell(row=1, column=3, value="Tipo Origem").font = Font(bold=True)
    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 16

    ri = 2
    for aba in ABAS:
        alimenta = aba.get("alimenta", "")
        if alimenta and "(nenhuma" not in alimenta and "(resultado" not in alimenta and "(saída" not in alimenta and "(resumo" not in alimenta and "(parâm" not in alimenta:
            ws2.cell(row=ri, column=1, value=aba["nome"])
            ws2.cell(row=ri, column=2, value=alimenta)
            ws2.cell(row=ri, column=3, value=aba["tipo"])
            ri += 1

    SAIDA.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(SAIDA))
    print(f"Salvo: {SAIDA}")
    print(f"  Aba 1: Mapa das Abas ({len(ABAS)} linhas)")
    print(f"  Aba 2: Dependências")


if __name__ == "__main__":
    gerar()
