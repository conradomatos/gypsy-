"""
Gera MAPA_ABAS_FAZZER.xlsx com descrições refinadas baseadas na inspeção real.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import os, re

BASE = r"D:\00 - CLAUDE_CODE\01. PROJETOS\04. ORCAMENTACAO_POR_MODELAGEM_FINANCEIRA\01_DESTILACAO"
FONTE = os.path.join(BASE, "fontes", "FAZZER ZIP.xlsm")
SAIDA = os.path.join(BASE, "intermediarios", "MAPA_ABAS_FAZZER.xlsx")

# Dados refinados de cada aba (baseado na inspeção real)
ABAS = [
    {
        "nome": "(1) Atividades",
        "dim": "211x4",
        "linhas_dados": "~209",
        "descricao": "Lista mestre de atividades/serviços. Estrutura hierárquica: Escopo → Disciplina → Atividade. É o cadastro de serviços que organiza o orçamento.",
        "inputs": "Escopos (pacotes), disciplinas (elétrica, instrumentação, etc.), atividades (serviços específicos)",
        "calcula": "Nada — é cadastro base. Gera códigos de atividade usados em Plan. Quant. e Equipes.",
        "puxa_de": "(nenhuma)",
        "alimenta": "(3) Plan. Quant., (4) Plan. Quant. (Resumo), Equipes, Cronograma, Histograma",
        "costai": "Composições (estrutura de serviços — código, nome, unidade)",
        "prioridade": "ALTA",
    },
    {
        "nome": "(2) Base Dados",
        "dim": "10000x8",
        "linhas_dados": "~5000+ (capacidade 10000)",
        "descricao": "Base de dados de insumos/materiais com código, família, descrição, unidade, HH e preço unitário. Inclui caixa de pesquisa. Famílias: Painéis, Cabos, Eletrodutos, etc.",
        "inputs": "Código, família, descrição, unidade, Hh por unidade, preço unitário",
        "calcula": "Nada — é cadastro base. VLOOKUP em Plan. Quant. busca dados daqui.",
        "puxa_de": "(nenhuma)",
        "alimenta": "(3) Plan. Quant. (via VLOOKUP por código)",
        "costai": "Insumos (código, nome, família, unidade, hh_unitario, preco_unitario). DADO CRÍTICO — é a base de preços e HH.",
        "prioridade": "ALTA",
    },
    {
        "nome": "(3) Plan. Quant.",
        "dim": "9919x20",
        "linhas_dados": "~variável (template 9919 linhas)",
        "descricao": "Planilha de quantitativos detalhada. Cada linha: Pacote/Disciplina/Atividade/Código/Descrição/Unid/Qtde/Hh. VLOOKUP busca descrição, unidade e HH da Base Dados pelo código.",
        "inputs": "Pacote, disciplina, atividade, código do item (busca automática na Base Dados), quantidade, Hh inserido, multiplicador",
        "calcula": "HH Total = Hh Inserido (ou Hh Base) × Multiplicador × Qtde. Totaliza HH por atividade.",
        "puxa_de": "(2) Base Dados (VLOOKUP por código → descrição, unidade, Hh)",
        "alimenta": "(4) Plan. Quant. (Resumo), (5) Tab. Dinâm., Equipes (HH total por atividade)",
        "costai": "Composições + ItensComposicao (quantitativos e coeficientes por serviço). Útil para extrair composições de serviço com seus insumos e quantidades.",
        "prioridade": "ALTA",
    },
    {
        "nome": "(4) Plan. Quant. (Resumo)",
        "dim": "5003x6",
        "linhas_dados": "~variável",
        "descricao": "Resumo consolidado dos quantitativos. Agrupamento por atividade com totais.",
        "inputs": "Nenhum — resume automaticamente da Plan. Quant.",
        "calcula": "Totais de HH e quantidades por atividade",
        "puxa_de": "(3) Plan. Quant.",
        "alimenta": "(5) Tab. Dinâm. Resumo",
        "costai": "Referência (validação de totais)",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(5) Tab. Dinâm. Resumo",
        "dim": "35x6",
        "linhas_dados": "~33",
        "descricao": "Tabela dinâmica (pivot) — resumo cruzado dos quantitativos para visualização rápida.",
        "inputs": "Nenhum — tabela dinâmica automática",
        "calcula": "Pivot de HH por disciplina/atividade",
        "puxa_de": "(4) Plan. Quant. (Resumo)",
        "alimenta": "(nenhuma)",
        "costai": "Nenhuma — é visualização",
        "prioridade": "IGNORAR",
    },
    {
        "nome": "(6) Profissionais (Salários)",
        "dim": "147x7",
        "linhas_dados": "~140 funções",
        "descricao": "Tabela de cargos e salários. Cada função: nº horas (220h), R$/hora, salário mensal. Referência: 2025. Inclui funções MOI e MOD (Gerente, Engenheiros, Técnicos, Soldadores, Montadores, etc.).",
        "inputs": "Função, nº horas trabalhadas (220h), R$/hora, salário mensal",
        "calcula": "R$/hora = Salário / Nº Horas",
        "puxa_de": "(nenhuma)",
        "alimenta": "Equipes (nomes das funções), (10) Composição MOI (salário base), (12) Composição MOD (salário base), (13) Valores Custo-Venda_HH (função)",
        "costai": "Mão de Obra (função, salário_mensal, custo_hora). BASE SALARIAL — todas as composições de MO derivam daqui.",
        "prioridade": "ALTA",
    },
    {
        "nome": "Equipes",
        "dim": "1006x161",
        "linhas_dados": "~1000 linhas (até 1006 atividades × 140+ funções nas colunas)",
        "descricao": "Dimensionamento de equipes por atividade. Linhas = atividades (pacote/disciplina/atividade + Hh previsto + datas). Colunas = funções (MOI e MOD). Cada célula = qtde de profissionais daquela função naquela atividade. Calcula equipe média e HH realizado.",
        "inputs": "Para cada atividade: data início, data término, quantidade de cada profissional. Nomes das funções vêm de Profissionais (Salários).",
        "calcula": "HH total realizado = Qtde Prof × Carga Horária × Dias. Diferença entre HH previsto e realizado.",
        "puxa_de": "(6) Profissionais (Salários) (nomes de funções), (3) Plan. Quant. (HH previsto por atividade)",
        "alimenta": "Histograma, Cronograma, MOI-MOD(2), (1)-(3) Histograma 01-03",
        "costai": "Composições (equipes por atividade) + ItensComposicao (funções e quantidades por equipe). Dado MUITO RICO — permite montar composições de equipe por tipo de serviço.",
        "prioridade": "ALTA",
    },
    {
        "nome": "Histograma",
        "dim": "149x1102",
        "linhas_dados": "~147 funções × períodos semanais",
        "descricao": "Histograma geral de MO — distribuição semanal de profissionais ao longo do cronograma. Matriz gigante: funções × semanas.",
        "inputs": "Nenhum — calculado de Equipes + Cronograma",
        "calcula": "Nº de profissionais por função por semana (curva de mobilização)",
        "puxa_de": "Equipes, Cronograma",
        "alimenta": "(1)-(3) Histograma 01-03, (7) Resumo Geral HH",
        "costai": "Nenhuma — é planejamento operacional (específico de obra, não dado de referência)",
        "prioridade": "IGNORAR",
    },
    {
        "nome": "Cronograma",
        "dim": "1006x1142",
        "linhas_dados": "~1000 atividades × semanas",
        "descricao": "Cronograma físico (Gantt) — duração e sequência de atividades por semana.",
        "inputs": "Datas de início/fim (vêm de Equipes)",
        "calcula": "Barras de Gantt (visual) e flags por semana",
        "puxa_de": "Equipes (datas)",
        "alimenta": "Histograma",
        "costai": "Nenhuma — é planejamento",
        "prioridade": "IGNORAR",
    },
    {
        "nome": "MOI - MOD (2)",
        "dim": "133x128",
        "linhas_dados": "~130 funções × meses (semanas)",
        "descricao": "Distribuição mensal/semanal de profissionais separados em MOI e MOD. Similar ao Histograma mas organizado por tipo de MO e mês.",
        "inputs": "Nenhum — derivado de Equipes",
        "calcula": "Profissionais MOI e MOD por semana/mês",
        "puxa_de": "Equipes",
        "alimenta": "(nenhuma diretamente — é visualização alternativa)",
        "costai": "Referência (classificação MOI vs MOD por função)",
        "prioridade": "MEDIA",
    },
    {
        "nome": "(1) Histograma - 01",
        "dim": "3816x40",
        "linhas_dados": "~3800",
        "descricao": "Histograma detalhado cenário 01 — HH por função por semana. Separa MOI e MOD. Inclui nomes das funções e HH acumulado.",
        "inputs": "Nenhum — calculado de Equipes e Cronograma",
        "calcula": "HH semanal por função. Alimenta Composição MOI/MOD com a função de cada profissional.",
        "puxa_de": "Equipes, Cronograma",
        "alimenta": "(4) Resumo Geral HH_01, (10) Composição MOI (nome da função), (12) Composição MOD (nome da função), (14) Resumo Preço Custo HH_01",
        "costai": "Referência (estrutura de funções por categoria MOI/MOD)",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(2) Histograma - 02",
        "dim": "3816x40",
        "linhas_dados": "~3800",
        "descricao": "Histograma detalhado cenário 02 — variação do cenário 1.",
        "inputs": "Nenhum — calculado",
        "calcula": "HH semanal por função no cenário 2",
        "puxa_de": "Equipes, Cronograma",
        "alimenta": "(5) Resumo Geral HH_02, (15) Resumo Preço Custo HH_02",
        "costai": "Nenhuma — cenário alternativo",
        "prioridade": "IGNORAR",
    },
    {
        "nome": "(3) Histograma - 03",
        "dim": "3816x40",
        "linhas_dados": "~3800",
        "descricao": "Histograma detalhado cenário 03 — variação do cenário 1.",
        "inputs": "Nenhum — calculado",
        "calcula": "HH semanal por função no cenário 3",
        "puxa_de": "Equipes, Cronograma",
        "alimenta": "(6) Resumo Geral HH_03, (16) Resumo Preço Custo HH_03",
        "costai": "Nenhuma — cenário alternativo",
        "prioridade": "IGNORAR",
    },
    {
        "nome": "(4) Resumo Geral - HH_01",
        "dim": "1753x17",
        "linhas_dados": "~1750",
        "descricao": "Totalização de HH do cenário 01 por função e período.",
        "inputs": "Nenhum — soma do Histograma 01",
        "calcula": "Total HH por função acumulado",
        "puxa_de": "(1) Histograma - 01",
        "alimenta": "(7) Resumo Geral - HH, (14) Resumo Preço Custo HH_01",
        "costai": "Referência (totais HH)",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(5) Resumo Geral - HH_02",
        "dim": "1753x17",
        "linhas_dados": "~1750",
        "descricao": "Totalização de HH do cenário 02.",
        "inputs": "Nenhum",
        "calcula": "Total HH cenário 2",
        "puxa_de": "(2) Histograma - 02",
        "alimenta": "(7) Resumo Geral - HH, (15) Resumo Preço Custo HH_02",
        "costai": "Referência",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(6) Resumo Geral - HH_03",
        "dim": "1753x17",
        "linhas_dados": "~1750",
        "descricao": "Totalização de HH do cenário 03.",
        "inputs": "Nenhum",
        "calcula": "Total HH cenário 3",
        "puxa_de": "(3) Histograma - 03",
        "alimenta": "(7) Resumo Geral - HH, (16) Resumo Preço Custo HH_03",
        "costai": "Referência",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(7) Resumo Geral - HH",
        "dim": "55x16384",
        "linhas_dados": "~50",
        "descricao": "Consolidação final dos 3 cenários — por mês: Nº Prof, HH, R$ Custo, R$ Venda. Só o mês de Janeiro está preenchido no exemplo (38 prof, 3344 HH, R$138k custo, R$266k venda).",
        "inputs": "Nenhum — consolida os 3 cenários",
        "calcula": "Comparativo mensal: profissionais, HH, custo e venda",
        "puxa_de": "(4)-(6) Resumo Geral HH_01-03, (13) Valores Custo-Venda_HH",
        "alimenta": "(25) Preço de Custo Total, (27) Preço de Venda Total",
        "costai": "Referência (resumo financeiro consolidado — validação)",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(8) Encargos Sociais",
        "dim": "21x4",
        "linhas_dados": "~19",
        "descricao": "Tabela de encargos sociais: INSS 20%, FGTS 8%, 13º Salário 8,33%, Férias+1/3 11,11%, Multa Rescisória 3,2%, Seguro Desemprego 2,4%, etc. Total ~60%. Cada linha tem alíquota e observação legal.",
        "inputs": "Encargo, alíquota (%), observações",
        "calcula": "Total de encargos sociais (soma das alíquotas). Usado como fator multiplicador nas Composições MOI/MOD.",
        "puxa_de": "(nenhuma)",
        "alimenta": "(10) Composição MOI, (12) Composição MOD",
        "costai": "EncargosBDI (encargos trabalhistas detalhados com alíquotas e base legal)",
        "prioridade": "ALTA",
    },
    {
        "nome": "Planilha1",
        "dim": "8x16",
        "linhas_dados": "~1",
        "descricao": "Aba auxiliar/rascunho — praticamente vazia. Apenas uma referência 'Preço Hora =' na célula O8. Provavelmente usada para testes durante o desenvolvimento da planilha.",
        "inputs": "Nenhum",
        "calcula": "Nada",
        "puxa_de": "(nenhuma)",
        "alimenta": "(nenhuma)",
        "costai": "Nenhuma",
        "prioridade": "IGNORAR",
    },
    {
        "nome": "(9) Incidências  MOI",
        "dim": "57x10",
        "linhas_dados": "~55",
        "descricao": "Detalhamento de custos adicionais (incidências) sobre MOI. Grupos: A) Despesas Admissionais (exames R$500, treinamentos), B) Benefícios (VA, VT, hospedagem, EPI), C) Custos Operacionais. Calcula custo mensal por pessoa por grupo.",
        "inputs": "Preço por item, rotatividade em meses, observações de obrigatoriedade",
        "calcula": "Custo mensal por pessoa de cada incidência. Total por grupo (A, B, C).",
        "puxa_de": "(nenhuma)",
        "alimenta": "(10) Composição MOI (custo de incidências por grupo A/B/C)",
        "costai": "EncargosBDI (incidências sobre MO — exames, benefícios, custos operacionais)",
        "prioridade": "ALTA",
    },
    {
        "nome": "(10) Composição MOI",
        "dim": "912x24",
        "linhas_dados": "~900",
        "descricao": "Composição completa de custo da MOI por função. Para cada função: salário base (de Profissionais) + encargos (de Encargos Sociais) + incidências A/B/C (de Incidências MOI) + hora extra + adicional noturno + periculosidade. Calcula custo/hora normal, HE 50%, HE 100%, com e sem adicional noturno.",
        "inputs": "Modalidade do contrato (CLT/PJ, pacote 220h/176h), percentuais de HE, adicional noturno, periculosidade, insalubridade",
        "calcula": "Custo/hora final por função MOI em 4+ modalidades (HN, HN+ADN, HE50%, HE50%+ADN, HE100%, etc.)",
        "puxa_de": "(6) Profissionais (Salários), (8) Encargos Sociais, (9) Incidências MOI, (1) Histograma - 01 (nome das funções)",
        "alimenta": "(13) Valores Custo-Venda_HH (custo/hora MOI)",
        "costai": "Composições (custo/hora MOI completo) + ItensComposicao (breakdown: salário, encargos, incidências A/B/C, HE, periculosidade). DADO MUITO RICO.",
        "prioridade": "ALTA",
    },
    {
        "nome": "(11) Incidências  MOD",
        "dim": "57x10",
        "linhas_dados": "~55",
        "descricao": "Detalhamento de custos adicionais (incidências) sobre MOD. Mesma estrutura da MOI: A) Admissionais, B) Benefícios, C) Operacionais. Valores podem diferir dos de MOI.",
        "inputs": "Preço por item, rotatividade em meses, observações",
        "calcula": "Custo mensal por pessoa de cada incidência MOD",
        "puxa_de": "(nenhuma)",
        "alimenta": "(12) Composição MOD (custo de incidências por grupo A/B/C)",
        "costai": "EncargosBDI (incidências sobre MOD)",
        "prioridade": "ALTA",
    },
    {
        "nome": "(12) Composição MOD",
        "dim": "1212x24",
        "linhas_dados": "~1200",
        "descricao": "Composição completa de custo da MOD por função. Mesma lógica da Composição MOI: salário + encargos + incidências + HE + adicionais. Mais funções que MOI (Encarregado Mecânica R$4400, Soldador, Montador, etc.).",
        "inputs": "Modalidade CLT/PJ, percentuais de HE/ADN/periculosidade/insalubridade, salário mínimo (R$1039 — defasado)",
        "calcula": "Custo/hora final por função MOD em múltiplas modalidades",
        "puxa_de": "(6) Profissionais (Salários), (8) Encargos Sociais, (11) Incidências MOD, (1) Histograma - 01 (nome das funções)",
        "alimenta": "(13) Valores Custo-Venda_HH (custo/hora MOD)",
        "costai": "Composições (custo/hora MOD completo) + ItensComposicao. DADO MUITO RICO.",
        "prioridade": "ALTA",
    },
    {
        "nome": "Macro1",
        "dim": "1x1",
        "linhas_dados": "0",
        "descricao": "Aba técnica de macro VBA — contém referências de código de automação. Sem dados.",
        "inputs": "Nenhum",
        "calcula": "Nada (código VBA)",
        "puxa_de": "(nenhuma)",
        "alimenta": "(nenhuma)",
        "costai": "Nenhuma",
        "prioridade": "IGNORAR",
    },
    {
        "nome": "(13) Valores Custo-Venda_HH",
        "dim": "147x21",
        "linhas_dados": "~140 funções",
        "descricao": "Resumo de valores R$/HH por função — custo E venda. Colunas: Custo HN, Venda HN, Custo HN+ADN, Venda HN+ADN, Custo HE50%, Venda HE50%, etc. Gerente Obras: R$176,89 custo → R$340,48 venda. Markup ~1,92x (vem da Memória Cálculo PV).",
        "inputs": "Nenhum — consolidação automática",
        "calcula": "Custo HH (de Composição MOI/MOD) × Fator de Venda (de Memória Cálculo PV) = Preço Venda HH",
        "puxa_de": "(6) Profissionais (Salários) (funções), (10) Composição MOI (custo/hora), (12) Composição MOD (custo/hora), (26) Memória Cálculo - PV (fator de markup)",
        "alimenta": "(14)-(16) Resumo Preço Custo HH, (17)-(19) Resumo Preço Venda HH, (7) Resumo Geral HH",
        "costai": "Mão de Obra (custo_hora_com_encargos + preço de venda por função). TABELA-CHAVE para preços de HH.",
        "prioridade": "ALTA",
    },
    {
        "nome": "(14) Resumo Preço Custo HH_01",
        "dim": "1755x16379",
        "linhas_dados": "~1750",
        "descricao": "Valoração do histograma cenário 01 a preço de CUSTO — multiplica HH por semana × custo/hora de cada função.",
        "inputs": "Nenhum — calculado",
        "calcula": "Custo total de MO por função por semana (cenário 1)",
        "puxa_de": "(1) Histograma - 01, (13) Valores Custo-Venda_HH",
        "alimenta": "(25) Preço de Custo Total (SUMIF MOD/MOI)",
        "costai": "Referência (custo MO valorado)",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(15) Resumo Preço Custo HH_02",
        "dim": "1755x16379",
        "linhas_dados": "~1750",
        "descricao": "Valoração do histograma cenário 02 a preço de custo.",
        "inputs": "Nenhum",
        "calcula": "Custo MO cenário 2",
        "puxa_de": "(2) Histograma - 02, (13) Valores Custo-Venda_HH",
        "alimenta": "(25) Preço de Custo Total",
        "costai": "Referência",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(16) Resumo Preço Custo HH_03",
        "dim": "1755x16379",
        "linhas_dados": "~1750",
        "descricao": "Valoração do histograma cenário 03 a preço de custo.",
        "inputs": "Nenhum",
        "calcula": "Custo MO cenário 3",
        "puxa_de": "(3) Histograma - 03, (13) Valores Custo-Venda_HH",
        "alimenta": "(25) Preço de Custo Total",
        "costai": "Referência",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(17) Resumo Preço Venda HH_01",
        "dim": "1755x16379",
        "linhas_dados": "~1750",
        "descricao": "Valoração do histograma cenário 01 a preço de VENDA.",
        "inputs": "Nenhum",
        "calcula": "Receita MO cenário 1",
        "puxa_de": "(1) Histograma - 01, (13) Valores Custo-Venda_HH",
        "alimenta": "(27) Preço de Venda Total",
        "costai": "Referência",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(18) Resumo Preço Venda HH_02",
        "dim": "1755x16379",
        "linhas_dados": "~1750",
        "descricao": "Valoração do histograma cenário 02 a preço de venda.",
        "inputs": "Nenhum",
        "calcula": "Receita MO cenário 2",
        "puxa_de": "(2) Histograma - 02, (13) Valores Custo-Venda_HH",
        "alimenta": "(27) Preço de Venda Total",
        "costai": "Referência",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(19) Resumo Preço Venda HH_03",
        "dim": "1755x16379",
        "linhas_dados": "~1750",
        "descricao": "Valoração do histograma cenário 03 a preço de venda.",
        "inputs": "Nenhum",
        "calcula": "Receita MO cenário 3",
        "puxa_de": "(3) Histograma - 03, (13) Valores Custo-Venda_HH",
        "alimenta": "(27) Preço de Venda Total",
        "costai": "Referência",
        "prioridade": "BAIXA",
    },
    {
        "nome": "(20) Equipe Mobilização - Custo",
        "dim": "36x7",
        "linhas_dados": "~34",
        "descricao": "Equipe designada para mobilização. Cada linha: função, valor HH (de Composição MOI/MOD), carga horária (8,8h), dias, valor total. Gerente Obras: R$176,89/h × 8,8h × 30 dias = R$46.699.",
        "inputs": "Função, carga horária diária, quantidade de dias",
        "calcula": "Custo de mobilização por profissional e total",
        "puxa_de": "(13) Valores Custo-Venda_HH (valor HH)",
        "alimenta": "(25) Preço de Custo Total (custo mobilização equipe)",
        "costai": "Composições (mobilização — equipe e custo)",
        "prioridade": "MEDIA",
    },
    {
        "nome": "(21) Mob. & Desm.",
        "dim": "51x10",
        "linhas_dados": "~48",
        "descricao": "Recursos de mobilização e desmobilização — containers (escritório, almoxarifado, banheiro), fretes, caminhões munck. Cada item: quantidade, valor mensal, meses, total. Ex: 10 containers × R$2.500/mês × 8 meses = R$200.000.",
        "inputs": "Descrição do recurso, quantidade, valor mensal de locação, qtde meses",
        "calcula": "Total = Qtde × Valor Mensal × Qtde Meses",
        "puxa_de": "(nenhuma — inputs manuais)",
        "alimenta": "(25) Preço de Custo Total (custo mob/desmob recursos)",
        "costai": "Composições (infraestrutura de canteiro) + referência de custos de mobilização",
        "prioridade": "MEDIA",
    },
    {
        "nome": "(22) Despesas Operacionais",
        "dim": "61x8",
        "linhas_dados": "~58",
        "descricao": "Despesas operacionais mensais — mesma estrutura de Mob/Desmob (containers, fretes, equipamentos de canteiro). Separada porque é custo mensal contínuo vs mobilização que é pontual.",
        "inputs": "Descrição, quantidade, valor mensal, meses de operação",
        "calcula": "Total = Qtde × Valor Mensal × Meses",
        "puxa_de": "(nenhuma — inputs manuais)",
        "alimenta": "(25) Preço de Custo Total (despesas operacionais)",
        "costai": "Composições (despesas operacionais de obra) / EncargosBDI (overhead)",
        "prioridade": "MEDIA",
    },
    {
        "nome": "(23) Máq. e Ferramentas",
        "dim": "61x10",
        "linhas_dados": "~58",
        "descricao": "Máquinas e ferramentas — cadastro com locação mensal. Esmerilhadeira 7\" (100 un × R$100/mês × 36 meses = R$360.000), conj. oxicorte, estufas, máquinas de solda, talhas, tifors, etc.",
        "inputs": "Descrição, quantidade, valor mensal locação, meses",
        "calcula": "Total = Qtde × Valor Mensal × Meses",
        "puxa_de": "(nenhuma — inputs manuais)",
        "alimenta": "(25) Preço de Custo Total (custo ferramentas)",
        "costai": "Equipamentos (nome, tipo=ALUGADO, custo_mensal, período)",
        "prioridade": "ALTA",
    },
    {
        "nome": "(24) Equipamentos",
        "dim": "81x10",
        "linhas_dados": "~78",
        "descricao": "Equipamentos pesados — empilhadeiras (gasolina/GLP/elétrica 1-8t), plataformas elevatórias (diesel/elétrica), guindastes, caminhões munck/carroceria, ônibus, vans. Locação mensal.",
        "inputs": "Descrição, quantidade, valor mensal locação, meses",
        "calcula": "Total = Qtde × Valor Mensal × Meses",
        "puxa_de": "(nenhuma — inputs manuais)",
        "alimenta": "(25) Preço de Custo Total (custo equipamentos)",
        "costai": "Equipamentos (nome, tipo=ALUGADO, custo_mensal, período). Complementa Máq. e Ferramentas.",
        "prioridade": "ALTA",
    },
    {
        "nome": "(25) Preço de Custo Total",
        "dim": "65x51",
        "linhas_dados": "~63",
        "descricao": "Composição de preços — consolidação TOTAL de custo. Itens: 1) MOD (R$52.712), 2) MOI (R$85.710), 3) Mob/Desmob (R$246.699), 4) Despesas Operacionais (R$250.000), 5) Máq/Ferramentas, 6) Equipamentos. Compara 3 cenários de histograma. Cliente: Demuth. Obra: Pátio de Madeiras - LD Celulose.",
        "inputs": "Nenhum — consolida automaticamente. Dados do projeto (cliente, local, proposta) são inputs manuais.",
        "calcula": "Preço de custo total = soma de todos os componentes por cenário",
        "puxa_de": "(14)-(16) Resumo Preço Custo HH (MOD/MOI), (20) Equipe Mob, (21) Mob&Desm, (22) Desp.Oper, (23) Máq.Ferr, (24) Equipamentos",
        "alimenta": "(26) Memória Cálculo PV (custo base para markup)",
        "costai": "Referência (composição de custo total — validação). Mostra estrutura de custos de um orçamento completo.",
        "prioridade": "MEDIA",
    },
    {
        "nome": "(26) Memória Cálculo - PV",
        "dim": "38x13",
        "linhas_dados": "~36",
        "descricao": "Memorial de cálculo para formação do preço de venda. Define BDI por componente: MO, Materiais, Desp.Operacionais, Máq/Ferramentas, Equipamentos. Cada um com percentuais de ISS, PIS, COFINS, IRPJ, CSLL, lucro. O fator de markup (~1,92x) é aplicado ao custo para obter preço de venda.",
        "inputs": "Percentuais de impostos (ISS, PIS, COFINS, IRPJ, CSLL), margem de lucro, administração central",
        "calcula": "Fator BDI por componente de custo. Preço Venda = Custo × Fator BDI.",
        "puxa_de": "(nenhuma — percentuais são inputs manuais. Pode referenciar Preço de Custo Total para validação)",
        "alimenta": "(13) Valores Custo-Venda_HH (fator markup para converter custo→venda), (27) Preço de Venda Total",
        "costai": "EncargosBDI (BDI completo — impostos, lucro, administração central). DADO CRÍTICO para formação de preço.",
        "prioridade": "ALTA",
    },
    {
        "nome": "(27) Preço de Venda Total",
        "dim": "65x51",
        "linhas_dados": "~63",
        "descricao": "Composição de preço de venda — espelho do Preço de Custo Total, mas com markup aplicado. MOD R$101.460, MOI R$164.975, Mob/Desmob R$387.886, Desp.Oper R$372.500. Mesma estrutura de 3 cenários.",
        "inputs": "Nenhum — custo × BDI",
        "calcula": "Preço de venda total = custo × fator BDI por componente",
        "puxa_de": "(25) Preço de Custo Total (ou componentes individuais), (26) Memória Cálculo PV (fator BDI), (17)-(19) Resumo Preço Venda HH",
        "alimenta": "(nenhuma — é o resultado final do orçamento)",
        "costai": "Referência (preço de venda final — validação de markup)",
        "prioridade": "MEDIA",
    },
]

# Criar Excel
wb_out = openpyxl.Workbook()
ws = wb_out.active
ws.title = "Mapa de Abas"

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
wrap = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

fills = {
    "ALTA": PatternFill("solid", fgColor="C6EFCE"),
    "MEDIA": PatternFill("solid", fgColor="FFEB9C"),
    "BAIXA": PatternFill("solid", fgColor="D9E2F3"),
    "IGNORAR": PatternFill("solid", fgColor="F2F2F2"),
}

# Headers
cols = ["Nº", "Nome da Aba", "Dimensão", "Linhas c/ Dados",
        "Descrição / Propósito", "Inputs do Usuário",
        "O que Calcula / Gera", "Puxa Dados De (fórmulas)",
        "Alimenta Quais Abas", "Tabela CostAI Destino",
        "Prioridade CostAI"]

for c, h in enumerate(cols, 1):
    cell = ws.cell(1, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Dados
for i, aba in enumerate(ABAS, 2):
    data = [
        i - 1, aba["nome"], aba["dim"], aba["linhas_dados"],
        aba["descricao"], aba["inputs"], aba["calcula"],
        aba["puxa_de"], aba["alimenta"], aba["costai"], aba["prioridade"]
    ]
    for c, val in enumerate(data, 1):
        cell = ws.cell(i, c, val)
        cell.alignment = wrap
        cell.border = thin_border

    # Cor por prioridade
    pri = aba["prioridade"]
    if pri in fills:
        for c in range(1, len(cols) + 1):
            ws.cell(i, c).fill = fills[pri]

# Larguras
widths = [5, 28, 12, 14, 55, 40, 50, 45, 50, 50, 15]
for c, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(c)].width = w

ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(ABAS)+1}"
ws.freeze_panes = "A2"

# === Aba 2: Fluxo de dados ===
ws2 = wb_out.create_sheet("Fluxo de Dados")
ws2.cell(1, 1, "FLUXO DE DADOS DA PLANILHA FAZZER").font = Font(bold=True, size=14)
ws2.merge_cells("A1:F1")

flows = [
    ("", ""),
    ("CAMADA 1 — CADASTROS BASE (inputs manuais)", ""),
    ("(1) Atividades", "Lista de serviços/atividades"),
    ("(2) Base Dados", "Insumos: código, família, descrição, unidade, HH, preço"),
    ("(6) Profissionais (Salários)", "Funções e salários base"),
    ("(8) Encargos Sociais", "Alíquotas trabalhistas (INSS, FGTS, 13°, férias...)"),
    ("(9) Incidências MOI", "Custos adicionais MOI (exames, benefícios, operacional)"),
    ("(11) Incidências MOD", "Custos adicionais MOD"),
    ("(26) Memória Cálculo - PV", "BDI: impostos + lucro + administração central"),
    ("", ""),
    ("CAMADA 2 — COMPOSIÇÕES DE CUSTO (calculadas)", ""),
    ("(10) Composição MOI", "Salário + Encargos + Incidências → Custo/hora MOI"),
    ("(12) Composição MOD", "Salário + Encargos + Incidências → Custo/hora MOD"),
    ("(13) Valores Custo-Venda_HH", "Custo HH × BDI → Preço de Venda HH"),
    ("", ""),
    ("CAMADA 3 — QUANTITATIVOS (específicos de obra)", ""),
    ("(3) Plan. Quant.", "Atividades × Insumos × Quantidades"),
    ("Equipes", "Atividades × Funções × Profissionais"),
    ("", ""),
    ("CAMADA 4 — RECURSOS E DESPESAS (inputs manuais)", ""),
    ("(23) Máq. e Ferramentas", "Ferramentas com locação mensal"),
    ("(24) Equipamentos", "Equipamentos pesados com locação"),
    ("(20) Equipe Mobilização", "Equipe para mob/desmob"),
    ("(21) Mob. & Desm.", "Infraestrutura de canteiro (containers, fretes)"),
    ("(22) Despesas Operacionais", "Custos fixos mensais de operação"),
    ("", ""),
    ("CAMADA 5 — HISTOGRAMAS E VALORAÇÃO (calculados)", ""),
    ("Histograma + (1)-(3) Histograma 01-03", "HH por função por semana (3 cenários)"),
    ("(4)-(6) Resumo Geral HH 01-03", "Totalização HH por cenário"),
    ("(14)-(16) Resumo Preço Custo HH", "HH × Custo/hora = Custo MO"),
    ("(17)-(19) Resumo Preço Venda HH", "HH × Venda/hora = Receita MO"),
    ("", ""),
    ("CAMADA 6 — CONSOLIDAÇÃO FINAL", ""),
    ("(25) Preço de Custo Total", "Soma: MOD + MOI + Mob + Desp + Máq + Equip"),
    ("(27) Preço de Venda Total", "Custo × BDI = Preço de Venda"),
    ("(7) Resumo Geral - HH", "Resumo mensal: Prof, HH, R$ Custo, R$ Venda"),
]

for r, (aba, desc) in enumerate(flows, 3):
    cell1 = ws2.cell(r, 1, aba)
    cell2 = ws2.cell(r, 2, desc)
    if desc == "":
        cell1.font = Font(bold=True, size=11, color="2F5496")
    else:
        cell2.alignment = wrap

ws2.column_dimensions['A'].width = 40
ws2.column_dimensions['B'].width = 60

# === Aba 3: Resumo CostAI ===
ws3 = wb_out.create_sheet("Mapeamento CostAI")
headers3 = ["Tabela CostAI", "Abas Fonte (ALTA)", "Abas Fonte (MEDIA/BAIXA)", "O que extrair", "Notas"]
for c, h in enumerate(headers3, 1):
    cell = ws3.cell(1, c, h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align

mapa_costai = [
    ("Insumos", "(2) Base Dados", "", "código, família, descrição, unidade, hh_unitario, preco_unitario", "Base de ~5000+ itens com HH por unidade. DADO CRÍTICO — fundação de todo o orçamento."),
    ("Mão de Obra", "(6) Profissionais (Salários), (13) Valores Custo-Venda_HH", "MOI-MOD(2)", "função, salário_mensal, custo_hora, custo_hora_com_encargos, tipo (MOI/MOD)", "~140 funções com salários 2025. Custo/hora completo inclui encargos+incidências."),
    ("Composições (MO)", "(10) Composição MOI, (12) Composição MOD", "", "Composição salarial: salário + encargos + incidências A/B/C + HE + adicionais → custo/hora", "Breakdown detalhado do custo/hora — 4+ modalidades (HN, HE50%, HE100%, com/sem ADN)."),
    ("Composições (Serviço)", "(3) Plan. Quant., Equipes", "", "Quantitativos por atividade + composição de equipes", "Específico de obra — útil como TEMPLATE de composições típicas."),
    ("Composições (Mob/Desp)", "", "(20) Equipe Mob, (21) Mob&Desm, (22) Desp.Oper", "Itens de mobilização, infraestrutura de canteiro, despesas mensais", "Custos fixos e de mobilização — referência para futuras obras."),
    ("Equipamentos", "(23) Máq. e Ferramentas, (24) Equipamentos", "", "nome, tipo=ALUGADO, custo_mensal, período", "~136 itens: ferramentas leves (esmerilhadeiras, soldas) + pesados (empilhadeiras, guindastes, plataformas)."),
    ("EncargosBDI", "(8) Encargos Sociais, (9) Incidências MOI, (11) Incidências MOD, (26) Memória Cálculo PV", "", "encargos trabalhistas + incidências (exames, benefícios) + BDI (impostos, lucro)", "Encargos ~60% + Incidências variáveis + BDI ~1,92x. Referência para formação de preço."),
]

for r, (tab, alta, media, extrair, notas) in enumerate(mapa_costai, 2):
    ws3.cell(r, 1, tab)
    ws3.cell(r, 2, alta)
    ws3.cell(r, 3, media)
    ws3.cell(r, 4, extrair)
    ws3.cell(r, 5, notas)
    for c in range(1, 6):
        ws3.cell(r, c).alignment = wrap
        ws3.cell(r, c).border = thin_border

for c, w in zip(range(1, 6), [25, 45, 40, 55, 60]):
    ws3.column_dimensions[get_column_letter(c)].width = w

# Salvar
wb_out.save(SAIDA)
print(f"Salvo: {SAIDA}")
print(f"Abas no arquivo: {wb_out.sheetnames}")
print(f"Total de abas mapeadas: {len(ABAS)}")
print(f"ALTA: {sum(1 for a in ABAS if a['prioridade'] == 'ALTA')}")
print(f"MEDIA: {sum(1 for a in ABAS if a['prioridade'] == 'MEDIA')}")
print(f"BAIXA: {sum(1 for a in ABAS if a['prioridade'] == 'BAIXA')}")
print(f"IGNORAR: {sum(1 for a in ABAS if a['prioridade'] == 'IGNORAR')}")
